import logging
from datetime import date, datetime, timedelta
import calendar
import json
import re
from collections import defaultdict
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from urllib.parse import quote
from io import BytesIO
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth import get_user_model
from django.utils.text import slugify
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count, OuterRef, Subquery, Exists
from django.db.models.functions import TruncMonth
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.conf import settings
from django.http import HttpResponse, JsonResponse
from django.core.files.base import ContentFile
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt

from . import forms, models, services, totalpass_service
from .signals import ensure_profissional_for_user
from shared.ai.openai_client import GeminiError, extract_address_from_proof, extract_student_from_document, improve_evolution_text
from .whatsapp_service import WhatsappService, WhatsappMessageType
from .whatsapp_scheduler import (
    _send_class_reminders, _send_professor_schedule, _send_contract_renewals,
    _send_birthdays, _send_payment_due, _send_payment_overdue, _send_three_months,
    reservas_lembrete_por_aluno, montar_mensagem_lembrete,
    _log_key, _already_sent, _store_log, _is_session_down,
)

logger = logging.getLogger(__name__)


def _contrato_assinatura_link(contrato, request=None):
    token = services.gerar_token_contrato(contrato)
    safe_token = quote(token, safe="")
    base_url = (settings.SITE_BASE_URL or "").rstrip("/")
    if request and (not base_url or "localhost" in base_url):
        base_url = request.build_absolute_uri("/").rstrip("/")
    return f"{base_url}/contratos/assinar/{safe_token}/"


def _mensagem_contrato_whatsapp(contrato, link, is_new=False):
    aluno = contrato.cdAluno
    plano = contrato.cdPlano
    unidade = contrato.cdUnidade
    prefix = "Seu contrato foi criado com sucesso" if is_new else "Seu contrato foi gerado"
    linhas = [
        f"Oi {aluno.dsNome}!",
        f"{prefix} no Mayris Pilates.",
        "",
        f"Contrato: #{contrato.cdContrato}",
        f"Plano: {plano}",
        f"Unidade: {unidade}",
        "",
        "Para ler e assinar, clique no link abaixo:",
        link,
        "",
        "Qualquer duvida, estamos a disposicao.",
    ]
    return "\n".join(linhas)


def _shift_month(value, months):
    offset = (value.month - 1) + months
    year = value.year + (offset // 12)
    month = (offset % 12) + 1
    return value.replace(year=year, month=month, day=1)


def _get_profissional_for_user(user):
    if not user or not user.is_authenticated:
        return None
    return models.Profissional.objects.select_related("cdPerfilAcesso").filter(user=user).first()


def _is_professor_user(user):
    profissional = _get_profissional_for_user(user)
    if not profissional or not profissional.cdPerfilAcesso_id:
        return False
    perfil = (profissional.cdPerfilAcesso.dsPerfilAcesso or "").strip().lower()
    return "professor" in perfil


def _professor_block(request):
    """Area restrita a admin. Retorna um redirect se o usuario for professor, senao None."""
    if _is_professor_user(request.user):
        messages.error(request, "Sem permissao para acessar esta area.")
        return redirect("aulas_list")
    return None


def _aplica_filtro_professor(request, profissional_id):
    """Professor so enxerga as proprias aulas: forca o filtro de profissional para o dele."""
    if _is_professor_user(request.user):
        prof = _get_profissional_for_user(request.user)
        return str(prof.id) if prof else "0"
    return profissional_id


def _admin_only_models():
    """Modelos cujo CRUD generico nao pode ser acessado por professor."""
    return (
        models.Plano, models.Contrato,
        models.ContasReceber, models.ContasPagar,
        models.ContaBancaria, models.MovimentoConta,
        models.Reserva,
    )


def _enviar_contrato_whatsapp(request, contrato, is_new=False):
    service = WhatsappService()
    telefone = service.get_aluno_phone(contrato.cdAluno)
    if not telefone:
        messages.warning(request, "Aluno sem telefone valido para WhatsApp.")
        return False
    link = _contrato_assinatura_link(contrato, request=request)
    pdf_link = _contrato_pdf_link(contrato, request=request)
    mensagem = _mensagem_contrato_whatsapp(contrato, link, is_new=is_new)
    if pdf_link:
        _salvar_documento_contrato(contrato)
        service.send_document(
            contrato.cdAluno,
            telefone,
            pdf_link,
            filename=f"Contrato-{contrato.cdContrato}.pdf",
            caption=f"Contrato {contrato.cdContrato} em PDF",
            contrato=contrato,
        )
    resp = service.send(contrato.cdAluno, telefone, mensagem, WhatsappMessageType.CONTRACT_LINK, contrato=contrato)
    if resp.get("error"):
        messages.warning(request, "Nao foi possivel enviar o contrato por WhatsApp.")
        return False
    messages.success(request, "Contrato enviado por WhatsApp.")
    return True


def _contrato_pdf_link(contrato, request=None):
    token = services.gerar_token_contrato(contrato)
    base_url = (settings.SITE_BASE_URL or "").rstrip("/")
    if request and (not base_url or "localhost" in base_url):
        base_url = request.build_absolute_uri("/").rstrip("/")
    return f"{base_url}/contratos/pdf/{token}/"


def _salvar_documento_contrato(contrato, pdf_bytes=None):
    pdf_bytes = pdf_bytes or services.render_contrato_pdf(contrato)
    max_cd = models.AlunoDocumento.objects.order_by("-cdDocumento").values_list("cdDocumento", flat=True).first() or 0
    documento, _ = models.AlunoDocumento.objects.update_or_create(
        contrato=contrato,
        origem="CONTRATO_PDF",
        defaults={
            "cdDocumento": max_cd + 1,
            "aluno": contrato.cdAluno,
            "titulo": f"Contrato #{contrato.cdContrato}",
            "descricao": f"Contrato gerado para {contrato.cdAluno.dsNome}.",
        },
    )
    documento.arquivo.save(
        f"contrato-{contrato.cdContrato}.pdf",
        ContentFile(pdf_bytes),
        save=True,
    )
    return documento


def _contrato_precificacao(plano, valor_aula=None):
    # Contratos sao cobrados mensalmente: o valor do plano e o valor MENSAL (parcela).
    # Gera 1 parcela por mes de duracao. O "valor por aula" e apenas informativo:
    # valor mensal / numero de aulas no mes (aulas_por_semana x 4 semanas).
    recorrencia = "MENSAL"
    valor_plano = Decimal(str(getattr(plano, "valor", 0) or 0))
    aulas_por_semana = int(getattr(plano, "aulas_por_semana", 0) or 0)
    duracao_meses = int(getattr(plano, "duracao_meses", 0) or 0) or 1
    aulas_no_mes = aulas_por_semana * 4

    valor_parcela = valor_plano
    valor_total = valor_plano * duracao_meses
    if aulas_no_mes > 0:
        valor_aula_decimal = (valor_plano / Decimal(aulas_no_mes)).quantize(
            Decimal("0.01"), rounding=ROUND_HALF_UP
        )
    else:
        valor_aula_decimal = None
    return recorrencia, valor_aula_decimal, valor_parcela, valor_total


def _aula_avulsa_precificacao(plano, quantidade, valor_aula=None):
    recorrencia = getattr(plano, "recorrencia", "SEMANAL") or "SEMANAL"
    valor_plano = Decimal(str(getattr(plano, "valor", 0) or 0))
    try:
        quantidade = max(int(quantidade or 1), 1)
    except (TypeError, ValueError):
        quantidade = 1
    try:
        valor_base = Decimal(str(valor_aula)) if valor_aula not in (None, "") else valor_plano
    except Exception:
        valor_base = valor_plano
    valor_total = valor_base
    return recorrencia, valor_base, valor_total, quantidade


def _advance_recorrencia_date(base_date, recorrencia, idx):
    if recorrencia == "MENSAL":
        return _add_months(base_date, idx)
    return base_date + timedelta(days=7 * idx)


def _align_date_to_weekday(base_date, weekday):
    delta = (weekday - base_date.weekday()) % 7
    return base_date + timedelta(days=delta)


def _competencia_avulsa(dt_value, recorrencia):
    if recorrencia == "MENSAL":
        return dt_value.strftime("%Y-%m")
    iso_year, iso_week, _ = dt_value.isocalendar()
    return f"{iso_year}-{iso_week:02d}"


def _active_menu(path: str) -> str:
    if path.startswith("/cadastros/alunos") or path.startswith("/cadastros/profissionais"):
        return "pessoas"
    if path.startswith("/agenda"):
        return "agenda"
    if path.startswith("/financeiro"):
        return "financeiro"
    if path.startswith("/configuracoes"):
        return "configuracoes"
    if path.startswith("/evolucoes"):
        return "evolucao"
    if path.startswith("/contratos") or path.startswith("/wizard"):
        return "contratos"
    if path.startswith("/cadastros"):
        return "cadastros"
    return "dashboard"


PHONE_CLEAN_REGEX = re.compile(r"\D+")

CONTRACT_TEMPLATE_VARIABLES = [
    {"key": "ALUNO_NOME", "label": "Nome do aluno"},
    {"key": "ALUNO_CPF", "label": "CPF do aluno"},
    {"key": "ALUNO_RG", "label": "RG do aluno"},
    {"key": "ALUNO_NASCIMENTO", "label": "Nascimento do aluno"},
    {"key": "ALUNO_ESTADO_CIVIL", "label": "Estado civil do aluno"},
    {"key": "ALUNO_PROFISSAO", "label": "Profissao do aluno"},
    {"key": "ALUNO_EMAIL", "label": "Email do aluno"},
    {"key": "ALUNO_TELEFONE", "label": "Telefone do aluno"},
    {"key": "ALUNO_ENDERECO", "label": "Endereco completo"},
    {"key": "ENDERECO_LOGRADOURO", "label": "Logradouro"},
    {"key": "ENDERECO_NUMERO", "label": "Numero"},
    {"key": "ENDERECO_BAIRRO", "label": "Bairro"},
    {"key": "ENDERECO_CIDADE", "label": "Cidade"},
    {"key": "ENDERECO_CEP", "label": "CEP"},
    {"key": "PROFISSIONAL_NOME", "label": "Profissional"},
    {"key": "PROFISSIONAL_CREFITO", "label": "Crefito"},
    {"key": "UNIDADE_NOME", "label": "Unidade"},
    {"key": "UNIDADE_CAPACIDADE", "label": "Capacidade"},
    {"key": "PLANO_NOME", "label": "Plano"},
    {"key": "PLANO_AULAS_SEMANA", "label": "Aulas por semana"},
    {"key": "PLANO_DURACAO_MESES", "label": "Duracao (meses)"},
    {"key": "PLANO_RECORRENCIA", "label": "Recorrencia do plano"},
    {"key": "TIPO_SERVICO", "label": "Tipo de servico"},
    {"key": "CONTRATO_NUMERO", "label": "Numero do contrato"},
    {"key": "CONTRATO_INICIO", "label": "Inicio do contrato"},
    {"key": "CONTRATO_FIM", "label": "Fim do contrato"},
    {"key": "CONTRATO_RECORRENCIA", "label": "Recorrencia do contrato"},
    {"key": "CONTRATO_MODO_PAGAMENTO", "label": "Modo de pagamento"},
    {"key": "CONTRATO_VALOR_PARCELA", "label": "Valor da parcela"},
    {"key": "CONTRATO_VALOR_TOTAL", "label": "Valor total"},
    {"key": "DATA_HOJE", "label": "Data de hoje"},
]


def _render_whatsapp_template(template, **context):
    class _SafeDict(dict):
        def __missing__(self, key):
            return "{" + key + "}"

    text = (template or "").strip()
    if not text:
        return ""
    try:
        return text.format_map(_SafeDict({k: "" if v is None else v for k, v in context.items()}))
    except Exception:
        return text


def _format_whatsapp_number(telefones):
    for tel in telefones:
        cleaned = PHONE_CLEAN_REGEX.sub("", tel or "")
        if not cleaned:
            continue
        if cleaned.startswith("55"):
            return cleaned
        if len(cleaned) in (10, 11):
            return f"55{cleaned}"
        return cleaned
    return None


def _round_to_next_hour(dt_value):
    if dt_value.minute == 0 and dt_value.second == 0 and dt_value.microsecond == 0:
        return dt_value
    return dt_value.replace(minute=0, second=0, microsecond=0) + timedelta(hours=1)


AULA_INTERVALO_MINUTOS = 10


def _load_funcionamento(unidade_id, tipo_servico_id=None):
    qs = models.HorarioFuncionamento.objects.filter(unidade_id=unidade_id, ativo=True)
    if tipo_servico_id:
        qs = qs.filter(
            Q(tipoServico_id=tipo_servico_id)
            | Q(tipos_servico=tipo_servico_id)
            | Q(tipoServico__isnull=True, tipos_servico__isnull=True)
        )
    horarios = list(qs.order_by("diaSemana", "horaInicio"))
    by_day = {}
    if tipo_servico_id:
        specific_days = {}
        for item in horarios:
            item_service_ids = {item.tipoServico_id} if item.tipoServico_id else set()
            if hasattr(item, "_prefetched_objects_cache") and "tipos_servico" in item._prefetched_objects_cache:
                item_service_ids.update(servico.id for servico in item.tipos_servico.all())
            else:
                item_service_ids.update(item.tipos_servico.values_list("id", flat=True))
            if tipo_servico_id in item_service_ids:
                specific_days.setdefault(item.diaSemana, []).append(item)
            else:
                by_day.setdefault(item.diaSemana, []).append(item)
        for day, items in specific_days.items():
            by_day[day] = items
    else:
        for item in horarios:
            by_day.setdefault(item.diaSemana, []).append(item)
    return by_day


def _parse_horario_servicos(post_data):
    raw_ids = post_data.getlist("tipos_servico")
    ids = []
    for raw_id in raw_ids:
        try:
            ids.append(int(raw_id))
        except (TypeError, ValueError):
            continue
    if not ids:
        raw_single = post_data.get("tipoServico") or ""
        if raw_single:
            try:
                ids.append(int(raw_single))
            except (TypeError, ValueError):
                pass
    return ids


def _load_horarios_ativos(unidade_id, tipo_servico_id=None):
    funcionamento = _load_funcionamento(unidade_id, tipo_servico_id)
    if funcionamento:
        return funcionamento, True

    qs = models.HorarioStudio.objects.filter(unidade_id=unidade_id)
    if tipo_servico_id is not None:
        qs = qs.filter(Q(tipoServico_id=tipo_servico_id) | Q(tipoServico__isnull=True))
    horarios = list(qs.order_by("diaSemana", "horaInicio"))
    by_day = {}
    for item in horarios:
        by_day.setdefault(item.diaSemana, []).append(item)
    return by_day, bool(by_day)


def _load_bloqueios(unidade_id, tipo_servico_id, profissional_id, start_date, end_date):
    qs = models.BloqueioAgenda.objects.filter(unidade_id=unidade_id, ativo=True)
    if tipo_servico_id is not None:
        qs = qs.filter(Q(tipoServico_id=tipo_servico_id) | Q(tipoServico__isnull=True))
    if profissional_id:
        qs = qs.filter(Q(profissional_id=profissional_id) | Q(profissional__isnull=True))
    else:
        qs = qs.filter(profissional__isnull=True)
    qs = qs.filter(Q(dataInicio__lte=end_date) & (Q(dataFim__isnull=True) | Q(dataFim__gte=start_date)))
    return list(qs)


def _is_slot_blocked(blocks, slot_date, slot_start, slot_end):
    for block in blocks:
        if block.recorrente:
            if block.diaSemana is None or block.diaSemana != slot_date.weekday():
                continue
            if block.dataInicio and slot_date < block.dataInicio:
                continue
            if block.dataFim and slot_date > block.dataFim:
                continue
        else:
            if slot_date < block.dataInicio:
                continue
            if block.dataFim and slot_date > block.dataFim:
                continue
        if slot_start < block.horaFim and slot_end > block.horaInicio:
            return True
    return False


def _generate_slots_for_date(target_date, horario_windows, duracao_minutos):
    slots = []
    for window in horario_windows:
        start_dt = _round_to_next_hour(datetime.combine(target_date, window.horaInicio))
        end_dt = datetime.combine(target_date, window.horaFim)
        while start_dt + timedelta(minutes=duracao_minutos) <= end_dt:
            slot_end = start_dt + timedelta(minutes=duracao_minutos)
            slots.append((start_dt.time(), slot_end.time()))
            start_dt += timedelta(minutes=duracao_minutos + AULA_INTERVALO_MINUTOS)
    return slots


def _build_reserva_slots(reservas, days_ahead=60):
    reservas_list = list(reservas or [])
    if not reservas_list:
        return {}
    today = timezone.now().date()
    end_date = today + timedelta(days=days_ahead)
    slots_map = {}
    for reserva in reservas_list:
        aula_atual = reserva.aulaSessao
        if not aula_atual:
            slots_map[reserva.id] = []
            continue
        # So geramos horarios (e criamos AulaSessao) para aulas futuras; nao ha como
        # remarcar uma aula passada. Isso evita varrer/criar sessoes para todo o
        # historico do aluno em cada abertura da ficha (causava WORKER TIMEOUT).
        if aula_atual.data < today:
            slots_map[reserva.id] = []
            continue
        unidade = aula_atual.unidade
        if not unidade:
            slots_map[reserva.id] = []
            continue
        start_date = min(today, aula_atual.data)
        duracao = unidade.duracao_aula_minutos or 50
        funcionamento = _load_funcionamento(unidade.id, aula_atual.tipoServico_id)
        if not funcionamento:
            slots_map[reserva.id] = []
            continue
        blocks = _load_bloqueios(unidade.id, aula_atual.tipoServico_id, aula_atual.profissional_id, start_date, end_date)
        existing_aulas = list(
            models.AulaSessao.objects.select_related("unidade", "profissional", "tipoServico")
            .filter(
                data__range=(start_date, end_date),
                unidade_id=unidade.id,
                tipoServico_id=aula_atual.tipoServico_id,
                profissional_id=aula_atual.profissional_id,
            )
            .order_by("data", "horaInicio")
        )
        existing_by_key = {(a.data, a.horaInicio): a for a in existing_aulas}
        current_date = start_date
        while current_date <= end_date:
            windows = funcionamento.get(current_date.weekday(), [])
            if windows:
                for hora_inicio, hora_fim in _generate_slots_for_date(current_date, windows, duracao):
                    if _is_slot_blocked(blocks, current_date, hora_inicio, hora_fim):
                        continue
                    key = (current_date, hora_inicio)
                    if key in existing_by_key:
                        continue
                    aula = models.AulaSessao.objects.create(
                        unidade_id=unidade.id,
                        tipoServico_id=aula_atual.tipoServico_id,
                        profissional_id=aula_atual.profissional_id,
                        data=current_date,
                        horaInicio=hora_inicio,
                        horaFim=hora_fim,
                    )
                    existing_by_key[key] = aula
            current_date += timedelta(days=1)

        aulas = list(existing_by_key.values())
        if aula_atual not in aulas:
            aulas.append(aula_atual)
        aula_ids = [a.id for a in aulas]
        reserva_counts = {
            row["aulaSessao_id"]: row["total"]
            for row in models.Reserva.objects.filter(aulaSessao_id__in=aula_ids, status="RESERVADA")
            .values("aulaSessao_id")
            .annotate(total=Count("id"))
        }
        slots = []
        for aula in sorted(aulas, key=lambda x: (x.data, x.horaInicio)):
            capacidade = aula.capacidade if aula.capacidade is not None else unidade.capacidade
            reservadas = reserva_counts.get(aula.id, 0)
            is_current = aula.id == aula_atual.id
            if not is_current and (capacidade is None or capacidade <= 0 or reservadas >= capacidade):
                continue
            vagas = max((capacidade or 0) - reservadas, 0)
            label = f"{aula.horaInicio.strftime('%H:%M')} - {aula.horaFim.strftime('%H:%M')}"
            if aula.profissional_id:
                label = f"{label} | {aula.profissional}"
            if capacidade and capacidade > 0:
                label = f"{label} ({vagas} vaga(s))"
            slots.append(
                {
                    "id": aula.id,
                    "date": aula.data.strftime("%Y-%m-%d"),
                    "time_start": aula.horaInicio.strftime("%H:%M"),
                    "time_end": aula.horaFim.strftime("%H:%M"),
                    "label": label,
                }
            )
        slots_map[reserva.id] = slots
    return slots_map



def login_view(request):
    if request.method == "POST":
        user = authenticate(request, username=request.POST.get("username"), password=request.POST.get("password"))
        if user:
            ensure_profissional_for_user(user)
            login(request, user)
            if _is_professor_user(user):
                return redirect("aulas_list")
            return redirect("dashboard")
        messages.error(request, "Login invalido")
    return render(request, "login.html")


def logout_view(request):
    logout(request)
    return redirect("login")


@login_required
def perfil_view(request):
    profissional = models.Profissional.objects.filter(user=request.user).first()
    if not profissional:
        profissional = ensure_profissional_for_user(request.user)
    if not profissional:
        messages.error(request, "Nao foi possivel localizar o seu cadastro.")
        return redirect("dashboard")
    return edit_view(request, models.Profissional, forms.ProfissionalForm, "perfil", profissional.pk)


@login_required
def dashboard(request):
    if _is_professor_user(request.user):
        return redirect("aulas_list")
    today = timezone.localdate()
    start_month = _shift_month(today.replace(day=1), -5)
    aniversariantes_mes = (
        models.Aluno.objects.filter(dtNascimento__month=today.month)
        .exclude(dtNascimento__isnull=True)
        .order_by("dtNascimento__day", "dsNome")[:8]
    )
    proximas_aulas = (
        models.Reserva.objects.select_related("aluno", "aulaSessao", "aulaSessao__profissional", "aulaSessao__unidade")
        .filter(aulaSessao__data__gte=today, status__in=["RESERVADA", "PENDENTE"])
        .order_by("aulaSessao__data", "aulaSessao__horaInicio")[:8]
    )
    contas_atrasadas = models.ContasReceber.objects.filter(
        status__in=["ABERTO", "ATRASADO"],
        dtVencimento__lt=today,
    ).count()
    contratos_vencendo = models.Contrato.objects.filter(
        dtFimContrato__gte=today,
        dtFimContrato__lte=today + timedelta(days=7),
    ).count()
    alertas = []
    if contas_atrasadas:
        alertas.append(f"{contas_atrasadas} faturas em atraso")
    if contratos_vencendo:
        alertas.append(f"{contratos_vencendo} contratos vencendo em 7 dias")

    aulas_por_mes_raw = (
        models.Reserva.objects.filter(aulaSessao__data__gte=start_month)
        .annotate(month=TruncMonth("aulaSessao__data"))
        .values("month")
        .annotate(total=Count("id"))
        .order_by("month")
    )
    aulas_por_mes_map = {item["month"]: item["total"] for item in aulas_por_mes_raw if item["month"]}
    aulas_por_mes = []
    for i in range(0, 6):
        month_start = _shift_month(start_month, i)
        total = aulas_por_mes_map.get(month_start, 0)
        aulas_por_mes.append(
            {
                "label": f"{calendar.month_abbr[month_start.month]}/{str(month_start.year)[-2:]}",
                "total": total,
            }
        )

    # Percentual para as barras do grafico de aulas por mes
    max_mes = max((item["total"] for item in aulas_por_mes), default=0) or 1
    for item in aulas_por_mes:
        item["pct"] = round(item["total"] / max_mes * 100)

    aulas_por_professor = list(
        models.Reserva.objects.filter(aulaSessao__data__gte=start_month, aulaSessao__profissional__isnull=False)
        .values("aulaSessao__profissional__profissional")
        .annotate(total=Count("id"))
        .order_by("-total")[:6]
    )
    max_prof = max((item["total"] for item in aulas_por_professor), default=0) or 1
    for item in aulas_por_professor:
        item["nome"] = item["aulaSessao__profissional__profissional"]
        item["pct"] = round(item["total"] / max_prof * 100)

    total_reservas_periodo = sum(item["total"] for item in aulas_por_mes)
    context = {
        "alunos": models.Aluno.objects.count(),
        "contratos": models.Contrato.objects.count(),
        "reservas_hoje": models.Reserva.objects.filter(aulaSessao__data=today).exclude(status="CANCELADA").count(),
        "receber_aberto": models.ContasReceber.objects.filter(status__in=["ABERTO", "ATRASADO"]).count(),
        "aniversariantes_mes": aniversariantes_mes,
        "proximas_aulas": proximas_aulas,
        "alertas": alertas,
        "aulas_por_mes": aulas_por_mes,
        "aulas_por_professor": aulas_por_professor,
        "total_reservas_periodo": total_reservas_periodo,
        "contas_atrasadas": contas_atrasadas,
        "contratos_vencendo": contratos_vencendo,
        "breadcrumbs": [("Home", "#")],
        "active_menu": "dashboard",
    }
    return render(request, "dashboard.html", context)

@login_required
@login_required
@require_POST
def contrato_whatsapp(request, pk):
    contrato = get_object_or_404(models.Contrato, pk=pk)
    next_url = request.POST.get("next")
    _enviar_contrato_whatsapp(request, contrato, is_new=False)
    if next_url:
        return redirect(next_url)
    return redirect(f"{reverse('alunos_detail', args=[contrato.cdAluno_id])}?tab=contratos")


def aluno_detail(request, pk):
    services.concluir_aulas_passadas()
    is_professor = _is_professor_user(request.user)
    aluno = get_object_or_404(models.Aluno, pk=pk)
    endereco = aluno.cdEndereco
    telefones = list(aluno.telefones.values_list("dsTelefone", flat=True))
    whatsapp_number = _format_whatsapp_number(telefones)
    contratos = models.Contrato.objects.filter(cdAluno=aluno).select_related("cdPlano", "cdUnidade")
    contrato_forms = {contrato.id: forms.ContratoForm(instance=contrato) for contrato in contratos}
    aulas_avulsas = (
        models.AulaAvulsa.objects.filter(aluno=aluno)
        .select_related("plano", "unidade", "profissional")
        .order_by("-dtCadastro")
    )
    aulas_avaliativas = (
        models.AulaAvaliativa.objects.filter(aluno=aluno)
        .select_related("unidade", "profissional", "tipoServico")
        .order_by("-data", "-horaInicio")
    )
    reservas_base = models.Reserva.objects.filter(aluno=aluno).select_related(
        "aulaSessao",
        "aulaSessao__profissional",
        "aulaSessao__unidade",
        "aulaSessao__tipoServico",
        "pacote_avulso",
        "pacote_avulso__plano",
    )
    # Filtro por status das aulas (padrao: apenas Reservadas)
    status_validos = {valor for valor, _ in models.Reserva.STATUS_CHOICES}
    status_filtro = [s for s in request.GET.getlist("status") if s in status_validos]
    if not status_filtro:
        status_filtro = ["RESERVADA"]
    contagem_status = {
        row["status"]: row["total"]
        for row in reservas_base.values("status").annotate(total=Count("id"))
    }
    # "Aulas que faltam": reservadas que ainda vao acontecer
    aulas_restantes = reservas_base.filter(
        status="RESERVADA", aulaSessao__data__gte=timezone.localdate()
    ).count()
    reservas = reservas_base.filter(status__in=status_filtro).order_by(
        "aulaSessao__data", "aulaSessao__horaInicio"
    )
    reserva_forms = {reserva.id: forms.ReservaForm(instance=reserva) for reserva in reservas}
    reserva_slots = _build_reserva_slots(reservas)
    evolucoes = (
        models.EvolucaoAluno.objects.filter(reserva__aluno=aluno)
        .select_related("profissional", "reserva")
        .order_by("-dtEvolucao")
    )
    avaliacoes = (
        models.AvaliacaoAluno.objects.filter(reserva__aluno=aluno)
        .select_related("profissional", "reserva")
        .order_by("-dtAvaliacao")
    )
    contas_receber = _filtrar_contas_receber(
        models.ContasReceber.objects.filter(
            Q(contrato__cdAluno=aluno) | Q(reserva__aluno=aluno) | Q(aluno=aluno)
        ).select_related(
            "contrato",
            "contrato__cdPlano",
            "contrato__cdPlano__subcategoria_receita",
            "reserva",
            "reserva__pacote_avulso",
            "reserva__pacote_avulso__plano",
        ),
        request,
    )
    planos = models.Plano.objects.select_related("cdTipoServico").filter(is_avulso=False)
    planos_avulsos = models.Plano.objects.select_related("cdTipoServico").filter(is_avulso=True)
    unidades = models.Unidade.objects.all()
    profissionais = models.Profissional.objects.all()
    whatsapp_messages = aluno.whatsapp_messages.select_related("contrato").all()
    whatsapp_form = forms.WhatsappMessageForm()
    wpp_config = models.WhatsappConfiguracao.objects.filter(unidade=aluno.cdUnidade_id).first() or models.WhatsappConfiguracao()
    _wpp_all = {
        "aniversario": (wpp_config.avisar_aniversario, wpp_config.template_aniversario),
        "fim_contrato": (wpp_config.avisar_fim_contrato, wpp_config.template_fim_contrato),
        "vencimento_proximo": (wpp_config.avisar_vencimento, wpp_config.template_vencimento_proximo),
        "mensalidade_atraso": (wpp_config.avisar_atraso, wpp_config.template_mensalidade_atraso),
        "tres_meses": (wpp_config.avisar_tres_meses, wpp_config.template_tres_meses),
    }
    # so mostra no modal os templates LIGADOS (toggle) e com texto
    whatsapp_templates = {k: txt for k, (ativo, txt) in _wpp_all.items() if ativo and txt}
    termos_uso = models.TermoUso.objects.all()
    documentos = aluno.documentos.select_related("contrato").all()
    documento_form = forms.AlunoDocumentoForm()
    aula_avulsa_form = forms.AulaAvulsaForm(
        initial={
            "unidade": aluno.cdUnidade_id,
            "dtInicio": timezone.localdate(),
            "quantidade": 1,
        }
    )
    aula_avulsa_preview = {}
    plano_avulso_padrao = planos_avulsos.first()
    if plano_avulso_padrao:
        qtd_preview = 1
        dt_inicio_preview = timezone.localdate()
        recorrencia_preview = (getattr(plano_avulso_padrao, "recorrencia", "SEMANAL") or "SEMANAL").upper()
        valor_aula_preview = Decimal(str(getattr(plano_avulso_padrao, "valor", 0) or 0))
        if recorrencia_preview == "MENSAL":
            dt_fim_preview = _add_months(dt_inicio_preview, qtd_preview)
        else:
            dt_fim_preview = dt_inicio_preview + timedelta(days=7 * qtd_preview)
        aula_avulsa_preview = {
            "plano_id": plano_avulso_padrao.id,
            "recorrencia": recorrencia_preview,
            "quantidade": qtd_preview,
            "valor_aula": valor_aula_preview,
            "valor_total": valor_aula_preview,
            "dtInicio": dt_inicio_preview,
            "dtFim": dt_fim_preview,
        }
    context = {
        "aluno": aluno,
        "endereco": endereco,
        "telefones": telefones,
        "whatsapp_number": whatsapp_number,
        "whatsapp_templates": whatsapp_templates,
        "termos_uso": termos_uso,
        "contratos": contratos,
        "contrato_forms": contrato_forms,
        "aulas_avulsas": aulas_avulsas,
        "aula_avulsa_form": aula_avulsa_form,
        "aulas_avaliativas": aulas_avaliativas,
        "tipos_servico": models.TipoServico.objects.all(),
        "reservas": reservas,
        "reserva_forms": reserva_forms,
        "reserva_slots": reserva_slots,
        "status_choices": models.Reserva.STATUS_CHOICES,
        "status_filtro": status_filtro,
        "contagem_status": contagem_status,
        "aulas_restantes": aulas_restantes,
        "evolucoes": evolucoes,
        "avaliacoes": avaliacoes,
        "contas_receber": contas_receber,
        "lancamento_avulso_form": forms.LancamentoAvulsoForm(
            initial={"dtVencimento": timezone.localdate(), "competencia": timezone.localdate().strftime("%Y-%m")}
        ),
        "filtros_financeiro": _get_filtros_financeiro(request),
        "today": timezone.now().date().strftime("%Y-%m-%d"),
        "today_date": timezone.now().date(),
        "planos": planos,
        "planos_avulsos": planos_avulsos,
        "aula_avulsa_preview": aula_avulsa_preview,
        "unidades": unidades,
        "profissionais": profissionais,
        "whatsapp_messages": whatsapp_messages,
        "whatsapp_form": whatsapp_form,
        "documentos": documentos,
        "documento_form": documento_form,
        "edit_form": forms.AlunoForm(instance=aluno),
        "breadcrumbs": [("Home", reverse("dashboard")), ("Alunos", reverse("alunos_list")), ("Ficha", "#")],
        "active_menu": "cadastros",
        "can_view_contratos": not is_professor,
        "can_view_financeiro": not is_professor,
        "is_professor": is_professor,
    }
    return render(request, "alunos/detail.html", context)


@login_required
@require_POST
def aluno_autoriza_imagem_toggle(request, pk):
    aluno = get_object_or_404(models.Aluno, pk=pk)
    next_url = request.POST.get("next", "").strip()
    if next_url and not next_url.startswith("/"):
        next_url = ""
    aluno.autoriza_imagem = request.POST.get("autoriza_imagem") in ("1", "on", "true", "True")
    aluno.save(update_fields=["autoriza_imagem"])
    messages.success(request, "Permissao de uso da imagem atualizada.")
    return redirect(next_url or reverse("alunos_detail", args=[aluno.pk]))


@login_required
@require_POST
def aluno_termo_registrar(request, pk):
    aluno = get_object_or_404(models.Aluno, pk=pk)
    termo = models.TermoUso.objects.filter(pk=request.POST.get("termo_id")).first()
    if not termo:
        messages.error(request, "Selecione um termo valido.")
        return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=imagem")
    aluno.cdTermoUso = termo
    aluno.termo_aceite_em = timezone.now()
    aluno.autoriza_imagem = True
    aluno.save(update_fields=["cdTermoUso", "termo_aceite_em", "autoriza_imagem"])
    messages.success(request, "Termo de uso de imagem registrado para o aluno.")
    return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=imagem")


@login_required
def aluno_termo_documento(request, pk):
    aluno = get_object_or_404(models.Aluno, pk=pk)
    termo = models.TermoUso.objects.filter(pk=request.GET.get("termo")).first() or aluno.cdTermoUso or models.TermoUso.objects.first()
    if not termo:
        messages.warning(request, "Nenhum termo cadastrado. Cadastre em Cadastros > Modelos de Termo de Uso.")
        return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=imagem")
    return render(
        request,
        "alunos/termo_documento.html",
        {"aluno": aluno, "termo": termo, "conteudo": services.render_termo_html(termo, aluno), "active_menu": "cadastros"},
    )


@login_required
@require_POST
def aluno_termo_whatsapp(request, pk):
    from django.utils.html import strip_tags

    aluno = get_object_or_404(models.Aluno, pk=pk)
    termo = aluno.cdTermoUso or models.TermoUso.objects.first()
    if not termo:
        messages.warning(request, "Nenhum termo cadastrado para enviar.")
        return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=imagem")
    service = WhatsappService()
    telefone = service.get_aluno_phone(aluno)
    if not telefone:
        messages.warning(request, "Aluno sem telefone valido cadastrado.")
        return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=imagem")
    texto = strip_tags(services.render_termo_html(termo, aluno)).strip()
    mensagem = f"Olá {aluno.dsNome}! Segue o nosso Termo de Autorização de Uso de Imagem:\n\n{texto}"
    resp = service.send(aluno, telefone, mensagem, WhatsappMessageType.MANUAL)
    if resp.get("error"):
        messages.warning(request, "Nao foi possivel enviar o termo via WhatsApp.")
    else:
        messages.success(request, "Termo de imagem enviado por WhatsApp.")
    return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=imagem")


def _aluno_termo(aluno):
    return aluno.cdTermoUso or models.TermoUso.objects.first()


@login_required
def aluno_termo_assinar_local(request, pk):
    aluno = get_object_or_404(models.Aluno, pk=pk)
    if not _aluno_termo(aluno):
        messages.warning(request, "Nenhum termo cadastrado para assinar.")
        return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=imagem")
    token = services.gerar_token_termo(aluno)
    return redirect("termo_assinar", token=token)


@login_required
@require_POST
def aluno_termo_enviar_assinatura_whatsapp(request, pk):
    aluno = get_object_or_404(models.Aluno, pk=pk)
    if not _aluno_termo(aluno):
        messages.warning(request, "Nenhum termo cadastrado para enviar.")
        return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=imagem")
    service = WhatsappService()
    telefone = service.get_aluno_phone(aluno)
    if not telefone:
        messages.warning(request, "Aluno sem telefone valido cadastrado.")
        return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=imagem")
    token = services.gerar_token_termo(aluno)
    link = request.build_absolute_uri(reverse("termo_assinar", args=[token]))
    mensagem = (
        f"Olá {aluno.dsNome}! Para autorizar o uso da sua imagem, assine nosso termo neste link: {link}"
    )
    resp = service.send(aluno, telefone, mensagem, WhatsappMessageType.MANUAL)
    if resp.get("error"):
        messages.warning(request, "Nao foi possivel enviar o link de assinatura via WhatsApp.")
    else:
        messages.success(request, "Link de assinatura do termo enviado por WhatsApp.")
    return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=imagem")


def termo_assinar(request, token):
    try:
        aluno_id = services.validar_token_termo(token)
    except Exception:
        return render(request, "alunos/termo_assinatura.html", {"token_invalido": True})
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    termo = _aluno_termo(aluno)
    if not termo:
        return render(request, "alunos/termo_assinatura.html", {"token_invalido": True})
    if request.method == "POST":
        if aluno.termo_assinado_em:
            return render(request, "alunos/termo_assinatura_sucesso.html", {"aluno": aluno, "ja_assinado": True})
        assinatura_nome = request.POST.get("assinatura_nome", "").strip()
        assinatura_documento = request.POST.get("assinatura_documento", "").strip()
        assinatura_data = request.POST.get("assinatura_data", "").strip()
        if assinatura_nome and assinatura_data.startswith("data:image/"):
            try:
                from django.core.files.base import ContentFile
                import base64

                _header, data = assinatura_data.split(",", 1)
                aluno.termo_assinatura_imagem = ContentFile(base64.b64decode(data), name=f"termo_{aluno.id}.png")
            except Exception:
                pass
        aluno.cdTermoUso = termo
        aluno.termo_assinatura_nome = assinatura_nome
        aluno.termo_assinatura_documento = assinatura_documento
        aluno.termo_assinatura_ip = request.META.get("REMOTE_ADDR")
        aluno.autoriza_imagem = True
        aluno.termo_aceite_em = timezone.now()
        aluno.termo_assinado_em = timezone.now()
        aluno.save()
        # Confirmacao por WhatsApp (resposta assinada)
        try:
            service = WhatsappService()
            telefone = service.get_aluno_phone(aluno)
            if telefone:
                service.send(
                    aluno,
                    telefone,
                    f"Olá {aluno.dsNome}! Recebemos sua autorização de uso de imagem assinada. Obrigada! 🌿",
                    WhatsappMessageType.MANUAL,
                )
        except Exception:
            logger.exception("Erro ao enviar confirmacao do termo assinado")
        return render(request, "alunos/termo_assinatura_sucesso.html", {"aluno": aluno})
    conteudo = services.render_termo_html(termo, aluno)
    return render(request, "alunos/termo_assinatura.html", {"aluno": aluno, "termo_html": conteudo, "token": token})


@login_required
def aluno_whatsapp_message(request, pk):
    aluno = get_object_or_404(models.Aluno, pk=pk)
    if request.method != "POST":
        return redirect("alunos_detail", pk=pk)
    next_url = request.POST.get("next")
    form = forms.WhatsappMessageForm(request.POST)
    if not form.is_valid():
        messages.error(request, "Informe uma mensagem valida para enviar.")
        return redirect(next_url or "alunos_detail", pk=aluno.pk) if not next_url else redirect(next_url)
    service = WhatsappService()
    telefone = service.get_aluno_phone(aluno)
    if not telefone:
        messages.warning(request, "Aluno sem telefone valido cadastrado.")
        return redirect(next_url or "alunos_detail", pk=aluno.pk) if not next_url else redirect(next_url)
    resp = service.send(aluno, telefone, form.cleaned_data["mensagem"], WhatsappMessageType.MANUAL)
    if resp.get("error"):
        messages.warning(request, "Mensagem registrada, mas nao foi possivel enviar via WhatsApp.")
    else:
        messages.success(request, "Mensagem enviada e registrada.")
    return redirect(next_url or "alunos_detail", pk=aluno.pk) if not next_url else redirect(next_url)


@login_required
@require_POST
def aluno_documento_create(request, pk):
    aluno = get_object_or_404(models.Aluno, pk=pk)
    data = request.POST.copy()
    data = _inject_cd_value(models.AlunoDocumento, data)
    form = forms.AlunoDocumentoForm(data, request.FILES)
    if form.is_valid():
        documento = form.save(commit=False)
        documento.aluno = aluno
        documento.origem = "UPLOAD"
        documento.save()
        messages.success(request, "Documento adicionado com sucesso.")
    else:
        messages.error(request, "Verifique os campos do documento.")
    return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=documentos")


@login_required
@require_POST
def aluno_documento_whatsapp(request, pk):
    aluno = get_object_or_404(models.Aluno, pk=pk)
    documento_id = request.POST.get("documento_id")
    documento = get_object_or_404(models.AlunoDocumento, pk=documento_id, aluno=aluno)
    service = WhatsappService()
    telefone = service.get_aluno_phone(aluno)
    if not telefone:
        messages.warning(request, "Aluno sem telefone valido cadastrado.")
        return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=documentos")
    if not documento.arquivo:
        messages.warning(request, "Documento sem arquivo disponivel.")
        return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=documentos")
    media_url = request.build_absolute_uri(documento.arquivo.url)
    resp = service.send_document(
        aluno,
        telefone,
        media_url,
        filename=Path(documento.arquivo.name).name,
        caption=documento.titulo,
        contrato=documento.contrato,
        tipo=WhatsappMessageType.STUDENT_DOCUMENT,
    )
    if resp.get("error"):
        messages.warning(request, "Nao foi possivel enviar o documento por WhatsApp.")
    else:
        messages.success(request, "Documento enviado por WhatsApp.")
    return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=documentos")


@login_required
@require_POST
def aluno_documento_delete(request, pk):
    documento = get_object_or_404(models.AlunoDocumento, pk=pk)
    aluno_id = documento.aluno_id
    if documento.arquivo:
        documento.arquivo.delete(save=False)
    documento.delete()
    messages.success(request, "Documento removido.")
    return redirect(f"{reverse('alunos_detail', args=[aluno_id])}?tab=documentos")


@login_required
@require_POST
def aluno_aula_avulsa_create(request, aluno_id):
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    data = request.POST.copy()
    data = _inject_cd_value(models.AulaAvulsa, data)
    data["aluno"] = aluno.pk
    plano_id = data.get("plano")
    plano = models.Plano.objects.filter(pk=plano_id, is_avulso=True).first() if plano_id else None
    if not plano:
        messages.error(request, "Selecione um plano de aula avulsa valido.")
        return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=agenda")
    quantidade_raw = data.get("quantidade") or 1
    valor_aula_input = data.get("valor_aula")
    recorrencia, valor_aula, valor_total, quantidade = _aula_avulsa_precificacao(plano, quantidade_raw, valor_aula_input)
    dt_inicio_raw = data.get("dtInicio") or timezone.now().date().strftime("%Y-%m-%d")
    try:
        dt_inicio = datetime.strptime(dt_inicio_raw, "%Y-%m-%d").date()
    except ValueError:
        dt_inicio = timezone.now().date()
    if recorrencia == "MENSAL":
        dt_fim = _add_months(dt_inicio, quantidade)
    else:
        dt_fim = dt_inicio + timedelta(days=7 * quantidade)
    data.update(
        {
            "recorrencia": recorrencia,
            "quantidade": quantidade,
            "valor_aula": valor_aula,
            "valor_total": valor_total,
            "dtInicio": dt_inicio,
            "dtFim": dt_fim,
        }
    )
    form = forms.AulaAvulsaForm(data)
    if not form.is_valid():
        messages.error(request, "Nao foi possivel criar a aula avulsa. Verifique os campos.")
        return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=agenda")
    pacote = form.save()
    return redirect("alunos_aula_avulsa_agenda", aluno_id=aluno.pk, pk=pacote.pk)


@login_required
def aluno_aula_avulsa_agenda(request, aluno_id, pk):
    pacote = get_object_or_404(
        models.AulaAvulsa.objects.select_related("aluno", "plano", "unidade", "profissional"),
        pk=pk,
        aluno_id=aluno_id,
    )
    aluno = pacote.aluno
    plano = pacote.plano
    quantidade = max(pacote.quantidade or 1, 1)
    agenda_fim = pacote.dtFim
    aulas = models.AulaSessao.objects.filter(
        unidade=pacote.unidade,
        data__range=(pacote.dtInicio, agenda_fim),
    ).select_related("unidade", "tipoServico").order_by("data", "horaInicio", "tipoServico_id")

    profissionais = list(models.Profissional.objects.all())
    duracao = pacote.unidade.duracao_aula_minutos or 50
    funcionamento, horarios_configurados = _load_horarios_ativos(pacote.unidade_id)

    slots = {}
    aulas_by_key = {(aula.data, aula.horaInicio, aula.horaFim, aula.profissional_id, aula.tipoServico_id): aula for aula in aulas}
    capacidade_padrao = pacote.unidade.capacidade or 0

    dates_by_weekday = {i: [] for i in range(7)}
    current = pacote.dtInicio
    while current <= agenda_fim:
        dates_by_weekday[current.weekday()].append(current)
        current += timedelta(days=1)

    blocks_by_prof = {}
    for prof in profissionais:
        blocks_by_prof[prof.id] = _load_bloqueios(
            pacote.unidade_id,
            None,
            prof.id,
            pacote.dtInicio,
            agenda_fim,
        )

    for weekday, windows in funcionamento.items():
        if not dates_by_weekday.get(weekday):
            continue
        sample_date = dates_by_weekday[weekday][0]
        time_slots = _generate_slots_for_date(sample_date, windows, duracao)
        for inicio, fim in time_slots:
            allowed_profs = []
            for prof in profissionais:
                prof_ok = True
                blocks = blocks_by_prof.get(prof.id, [])
                for day in dates_by_weekday[weekday]:
                    if _is_slot_blocked(blocks, day, inicio, fim):
                        prof_ok = False
                        break
                    aula = aulas_by_key.get((day, inicio, fim, prof.id, plano.cdTipoServico_id))
                    if aula:
                        reservadas = models.Reserva.objects.filter(aulaSessao=aula, status="RESERVADA").count()
                        cap = aula.capacidade_efetiva()
                    else:
                        reservadas = 0
                        cap = capacidade_padrao
                    if reservadas >= (cap or 0):
                        prof_ok = False
                        break
                if prof_ok:
                    allowed_profs.append(prof.id)
            if allowed_profs:
                slots[(weekday, inicio, fim)] = {
                    "weekday": weekday,
                    "inicio": inicio,
                    "fim": fim,
                    "allowed_profs": allowed_profs,
                }

    if not slots and not horarios_configurados:
        slots = {}
        for aula in aulas:
            reservadas = models.Reserva.objects.filter(aulaSessao=aula, status="RESERVADA").count()
            if reservadas >= aula.capacidade_efetiva():
                continue
            key = (aula.data.weekday(), aula.horaInicio, aula.horaFim)
            slots[key] = {"weekday": key[0], "inicio": key[1], "fim": key[2], "allowed_profs": []}

    if request.method == "POST":
        slot_value = request.POST.get("slot_1") or ""
        dia_raw = request.POST.get("slot_day_1") or ""
        prof_id = request.POST.get("prof_for_1") or ""
        if not slot_value:
            messages.error(request, "Selecione um horario para a aula avulsa.")
            return redirect("alunos_aula_avulsa_agenda", aluno_id=aluno.pk, pk=pacote.pk)
        try:
            weekday_raw, inicio, fim = slot_value.split("|")
            weekday = int(weekday_raw)
            inicio_time = datetime.strptime(inicio, "%H:%M").time()
            fim_time = datetime.strptime(fim, "%H:%M").time()
        except ValueError:
            messages.error(request, "Horario invalido na selecao.")
            return redirect("alunos_aula_avulsa_agenda", aluno_id=aluno.pk, pk=pacote.pk)
        if dia_raw:
            try:
                dia_selected = int(dia_raw)
            except ValueError:
                messages.error(request, "Dia da semana invalido.")
                return redirect("alunos_aula_avulsa_agenda", aluno_id=aluno.pk, pk=pacote.pk)
            if dia_selected != weekday:
                messages.error(request, "Dia e horario nao conferem.")
                return redirect("alunos_aula_avulsa_agenda", aluno_id=aluno.pk, pk=pacote.pk)
        try:
            prof_id = int(prof_id)
        except ValueError:
            prof_id = None
        if not prof_id:
            messages.error(request, "Selecione o professor.")
            return redirect("alunos_aula_avulsa_agenda", aluno_id=aluno.pk, pk=pacote.pk)
        horario_key = (weekday, inicio_time, fim_time)
        if horario_key not in slots:
            messages.error(request, "Horario invalido para o dia selecionado.")
            return redirect("alunos_aula_avulsa_agenda", aluno_id=aluno.pk, pk=pacote.pk)
        slot_payload = slots.get(horario_key, {})
        allowed = set(slot_payload.get("allowed_profs") or [])
        if allowed and prof_id not in allowed:
            messages.error(request, "Professor invalido para o horario selecionado.")
            return redirect("alunos_aula_avulsa_agenda", aluno_id=aluno.pk, pk=pacote.pk)

        conflitos = []
        primeira_data = _align_date_to_weekday(pacote.dtInicio, weekday)
        for idx in range(quantidade):
            target_date = _advance_recorrencia_date(primeira_data, pacote.recorrencia, idx)
            if target_date > agenda_fim:
                conflitos.append(f"Fora do periodo em {target_date} {inicio}")
                continue
            blocks = blocks_by_prof.get(prof_id, [])
            if _is_slot_blocked(blocks, target_date, inicio_time, fim_time):
                conflitos.append(f"Bloqueio em {target_date} {inicio}")
                continue
            servicos_ids = _horario_servicos_ids(
                pacote.unidade_id,
                target_date,
                inicio_time,
                fim_time,
                fallback=[plano.cdTipoServico_id] if plano.cdTipoServico_id else [],
            )
            if not servicos_ids:
                servicos_ids = [plano.cdTipoServico_id] if plano.cdTipoServico_id else []
            for servico_id in servicos_ids:
                try:
                    aula = models.AulaSessao.objects.filter(
                        unidade=pacote.unidade,
                        tipoServico_id=servico_id,
                        profissional_id=prof_id,
                        data=target_date,
                        horaInicio=inicio_time,
                        horaFim=fim_time,
                    ).first()
                    if not aula:
                        aula = models.AulaSessao.objects.create(
                            unidade=pacote.unidade,
                            tipoServico_id=servico_id,
                            profissional_id=prof_id,
                            data=target_date,
                            horaInicio=inicio_time,
                            horaFim=fim_time,
                        )
                    else:
                        aula.profissional_id = prof_id
                        aula.save(update_fields=["profissional"])
                    if models.Reserva.objects.filter(aluno=pacote.aluno, aulaSessao=aula).exists():
                        continue
                    reserva = services.create_reserva(pacote.aluno, aula, status="RESERVADA", pacote_avulso=pacote)
                    if not models.ContasReceber.objects.filter(reserva=reserva).exists():
                        models.ContasReceber.objects.create(
                            contrato=None,
                            reserva=reserva,
                            status="ABERTO",
                            valor=pacote.valor_aula or 0,
                            dtVencimento=target_date,
                            competencia=_competencia_avulsa(target_date, pacote.recorrencia),
                        )
                except Exception:
                    conflitos.append(f"Sem vaga em {target_date} {inicio}")
        if conflitos:
            messages.warning(request, "Conflitos ao reservar: " + "; ".join(conflitos[:5]))
        else:
            messages.success(request, "Aulas avulsas agendadas com sucesso.")
            return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=agenda")

    weekday_labels = {
        0: "Segunda",
        1: "Terca",
        2: "Quarta",
        3: "Quinta",
        4: "Sexta",
        5: "Sabado",
        6: "Domingo",
    }
    slots_by_day = {key: [] for key in weekday_labels}
    seen_slots = set()
    for item in sorted(slots.values(), key=lambda x: (x["weekday"], x["inicio"])):
        key = (item["weekday"], item["inicio"], item["fim"])
        if key in seen_slots:
            continue
        seen_slots.add(key)
        slots_by_day[item["weekday"]].append(
            {
                "label": f'{item["inicio"].strftime("%H:%M")} - {item["fim"].strftime("%H:%M")}',
                "value": f'{item["weekday"]}|{item["inicio"].strftime("%H:%M")}|{item["fim"].strftime("%H:%M")}',
                "allowed_profs": ",".join([str(pid) for pid in (item.get("allowed_profs") or [])]),
            }
        )

    slot_options = [
        {
            "label": f'{weekday_labels[item["weekday"]]} {item["inicio"].strftime("%H:%M")} - {item["fim"].strftime("%H:%M")}',
            "value": f'{item["weekday"]}|{item["inicio"].strftime("%H:%M")}|{item["fim"].strftime("%H:%M")}',
            "allowed_profs": item.get("allowed_profs") or [],
        }
        for item in sorted(slots.values(), key=lambda x: (x["weekday"], x["inicio"]))
    ]

    context = {
        "pacote": pacote,
        "aluno": pacote.aluno,
        "aulas_por_semana": 1,
        "agenda_semanas": quantidade if pacote.recorrencia == "SEMANAL" else None,
        "slot_indices": [1],
        "slots_by_day": slots_by_day,
        "weekday_labels": weekday_labels,
        "slot_options": slot_options,
        "profissionais": profissionais,
        "horarios_configurados": horarios_configurados,
        "breadcrumbs": [("Home", reverse("dashboard")), ("Alunos", reverse("alunos_list")), ("Aula Avulsa", "#")],
        "active_menu": "cadastros",
    }
    return render(request, "alunos/aula_avulsa_agenda.html", context)


@login_required
@require_POST
def aluno_aula_avaliativa_create(request, aluno_id):
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    redir = f"{reverse('alunos_detail', args=[aluno.pk])}?tab=agenda"

    unidade = models.Unidade.objects.filter(pk=request.POST.get("unidade")).first()
    profissional = models.Profissional.objects.filter(pk=request.POST.get("profissional")).first()
    tipo = models.TipoServico.objects.filter(pk=request.POST.get("tipoServico")).first()
    data_raw = request.POST.get("data") or ""
    hora_raw = request.POST.get("horaInicio") or ""
    if not (unidade and profissional and tipo and data_raw and hora_raw):
        messages.error(request, "Preencha unidade, profissional, tipo de servico, data e horario.")
        return redirect(redir)
    try:
        data_val = datetime.strptime(data_raw, "%Y-%m-%d").date()
        hora_inicio = datetime.strptime(hora_raw, "%H:%M").time()
    except ValueError:
        messages.error(request, "Data ou horario invalido.")
        return redirect(redir)

    duracao = unidade.duracao_aula_minutos or 50
    hora_fim = (datetime.combine(data_val, hora_inicio) + timedelta(minutes=duracao)).time()

    ultimo = models.AulaAvaliativa.objects.order_by("-cdAulaAvaliativa").first()
    cd = (ultimo.cdAulaAvaliativa if ultimo else 0) + 1
    avaliativa = models.AulaAvaliativa.objects.create(
        cdAulaAvaliativa=cd,
        aluno=aluno,
        unidade=unidade,
        profissional=profissional,
        tipoServico=tipo,
        data=data_val,
        horaInicio=hora_inicio,
        horaFim=hora_fim,
    )

    # Agenda: cria/encontra a AulaSessao e reserva o aluno (aula gratuita, sem cobranca)
    aula = models.AulaSessao.objects.filter(
        unidade=unidade,
        tipoServico=tipo,
        profissional=profissional,
        data=data_val,
        horaInicio=hora_inicio,
        horaFim=hora_fim,
    ).first()
    if not aula:
        aula = models.AulaSessao.objects.create(
            unidade=unidade,
            tipoServico=tipo,
            profissional=profissional,
            data=data_val,
            horaInicio=hora_inicio,
            horaFim=hora_fim,
        )
    if models.Reserva.objects.filter(aluno=aluno, aulaSessao=aula).exists():
        messages.success(request, "Aula avaliativa registrada (aluno ja tinha reserva nesse horario).")
        return redirect(redir)
    try:
        reserva = services.create_reserva(aluno, aula, status="RESERVADA")
        reserva.aula_avaliativa = avaliativa
        reserva.save(update_fields=["aula_avaliativa"])
        messages.success(request, "Aula avaliativa agendada com sucesso.")
    except Exception:
        messages.warning(request, "Aula avaliativa criada, mas o horario esta sem vaga na agenda.")
    return redirect(redir)


@login_required
@require_POST
def aluno_aula_avaliativa_ficha(request, aluno_id, pk):
    avaliativa = get_object_or_404(models.AulaAvaliativa, pk=pk, aluno_id=aluno_id)
    avaliativa.queixa_principal = request.POST.get("queixa_principal", "").strip()
    avaliativa.objetivos = request.POST.get("objetivos", "").strip()
    avaliativa.historico_saude = request.POST.get("historico_saude", "").strip()
    avaliativa.observacoes = request.POST.get("observacoes", "").strip()
    status = request.POST.get("status", "").strip()
    if status in dict(models.AulaAvaliativa.STATUS_CHOICES):
        avaliativa.status = status
    avaliativa.save()
    messages.success(request, "Ficha da aula avaliativa salva.")
    return redirect(f"{reverse('alunos_detail', args=[aluno_id])}?tab=agenda")


@login_required
@require_POST
def aluno_aula_avaliativa_converter(request, aluno_id, pk):
    avaliativa = get_object_or_404(models.AulaAvaliativa, pk=pk, aluno_id=aluno_id)
    avaliativa.status = "CONVERTIDA"
    avaliativa.save(update_fields=["status"])
    messages.success(
        request,
        "Aula avaliativa marcada como convertida. Crie o contrato do aluno na aba Contratos.",
    )
    return redirect(f"{reverse('alunos_detail', args=[aluno_id])}?tab=contratos")


@login_required
@require_POST
def aluno_evolucao_create(request, aluno_id):
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    reserva_id = request.POST.get("reserva_id")
    profissional_id = request.POST.get("profissional_id")
    texto = (request.POST.get("texto") or "").strip()
    if not texto or not reserva_id or not profissional_id:
        messages.error(request, "Preencha todos os campos da evolucao.")
        return redirect("alunos_detail", pk=aluno.pk)
    reserva = get_object_or_404(models.Reserva, pk=reserva_id, aluno=aluno)
    profissional = get_object_or_404(models.Profissional, pk=profissional_id)
    models.EvolucaoAluno.objects.create(reserva=reserva, profissional=profissional, texto=texto)
    messages.success(request, "Evolucao registrada.")
    return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=evolucao")


@login_required
@require_POST
def aluno_evolucao_update(request, evolucao_id):
    evolucao = get_object_or_404(models.EvolucaoAluno, pk=evolucao_id)
    texto = (request.POST.get("texto") or "").strip()
    if not texto:
        messages.error(request, "Informe o texto da evolucao.")
        return redirect(f"{reverse('alunos_detail', args=[evolucao.reserva.aluno_id])}?tab=evolucao")
    evolucao.texto = texto
    evolucao.save(update_fields=["texto"])
    messages.success(request, "Evolucao atualizada.")
    return redirect(f"{reverse('alunos_detail', args=[evolucao.reserva.aluno_id])}?tab=evolucao")


@login_required
@require_POST
def aluno_evolucao_delete(request, evolucao_id):
    evolucao = get_object_or_404(models.EvolucaoAluno, pk=evolucao_id)
    aluno_id = evolucao.reserva.aluno_id
    evolucao.delete()
    messages.success(request, "Evolucao removida.")
    return redirect(f"{reverse('alunos_detail', args=[aluno_id])}?tab=evolucao")


@login_required
@require_POST
def aluno_avaliacao_create(request, aluno_id):
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    reserva_id = request.POST.get("reserva_id")
    profissional_id = request.POST.get("profissional_id")
    texto = (request.POST.get("texto") or "").strip()
    if not texto or not reserva_id or not profissional_id:
        messages.error(request, "Preencha todos os campos da avaliacao.")
        return redirect("alunos_detail", pk=aluno.pk)
    reserva = get_object_or_404(models.Reserva, pk=reserva_id, aluno=aluno)
    profissional = get_object_or_404(models.Profissional, pk=profissional_id)
    models.AvaliacaoAluno.objects.create(reserva=reserva, profissional=profissional, texto=texto)
    messages.success(request, "Avaliacao registrada.")
    return redirect(f"{reverse('alunos_detail', args=[aluno.pk])}?tab=avaliacao")


@login_required
@require_POST
def aluno_avaliacao_update(request, avaliacao_id):
    avaliacao = get_object_or_404(models.AvaliacaoAluno, pk=avaliacao_id)
    texto = (request.POST.get("texto") or "").strip()
    if not texto:
        messages.error(request, "Informe o texto da avaliacao.")
        return redirect(f"{reverse('alunos_detail', args=[avaliacao.reserva.aluno_id])}?tab=avaliacao")
    avaliacao.texto = texto
    avaliacao.save(update_fields=["texto"])
    messages.success(request, "Avaliacao atualizada.")
    return redirect(f"{reverse('alunos_detail', args=[avaliacao.reserva.aluno_id])}?tab=avaliacao")


@login_required
@require_POST
def aluno_avaliacao_delete(request, avaliacao_id):
    avaliacao = get_object_or_404(models.AvaliacaoAluno, pk=avaliacao_id)
    aluno_id = avaliacao.reserva.aluno_id
    avaliacao.delete()
    messages.success(request, "Avaliacao removida.")
    return redirect(f"{reverse('alunos_detail', args=[aluno_id])}?tab=avaliacao")


def _sync_user_for_profissional(profissional, raw_password=None, old_cd=None):
    User = get_user_model()
    base_username = (profissional.email or "").strip().lower()
    if not base_username:
        base_username = slugify(profissional.profissional) or f"user-{profissional.id}"
    username = base_username
    counter = 1
    while User.objects.filter(username=username).exclude(pk=getattr(profissional.user, "pk", None)).exists():
        counter += 1
        username = f"{base_username}-{counter}"

    if profissional.user_id:
        user = profissional.user
        user.username = username
        user.first_name = profissional.profissional
        if profissional.email:
            user.email = profissional.email
    else:
        user = User(username=username, first_name=profissional.profissional, email=(profissional.email or ""))

    if raw_password:
        user.set_password(raw_password)
    elif not user.has_usable_password():
        user.set_unusable_password()
    user.save()

    if not profissional.user_id:
        profissional.user = user
        profissional.save(update_fields=["user"])


def _inject_cd_value(model, data):
    cd_field = next((f.name for f in model._meta.fields if f.name.startswith("cd")), None)
    if not cd_field:
        return data
    if data.get(cd_field):
        return data
    max_val = model.objects.order_by(f"-{cd_field}").values_list(cd_field, flat=True).first() or 0
    data[cd_field] = max_val + 1
    return data


def _sync_aluno_address(aluno, data):
    fields = {
        "dsLogradouro": data.get("dsLogradouro", "").strip(),
        "dsNumero": data.get("dsNumero", "").strip(),
        "dsCEP": data.get("dsCEP", "").strip(),
        "dsCidade": data.get("dsCidade", "").strip(),
        "dsBairro": data.get("dsBairro", "").strip(),
    }
    if not any(fields.values()):
        return
    endereco = aluno.cdEndereco
    if not endereco:
        max_cd = models.EnderecoAluno.objects.order_by("-cdEndereco").values_list("cdEndereco", flat=True).first() or 0
        endereco = models.EnderecoAluno(cdEndereco=max_cd + 1, cdAluno=aluno)
    for key, value in fields.items():
        setattr(endereco, key, value)
    endereco.save()
    if aluno.cdEndereco_id != endereco.id:
        aluno.cdEndereco = endereco
        aluno.save(update_fields=["cdEndereco"])


def _sync_aluno_phones(aluno, data):
    phones = []
    for key, value in data.items():
        if key.startswith("telefone_"):
            raw = value.strip()
            if raw:
                phones.append(raw)
    if not phones:
        return
    models.TelefoneAluno.objects.filter(cdAluno=aluno).delete()
    max_cd = models.TelefoneAluno.objects.order_by("-cdTelefone").values_list("cdTelefone", flat=True).first() or 0
    for idx, numero in enumerate(phones, start=1):
        models.TelefoneAluno.objects.create(cdTelefone=max_cd + idx, cdAluno=aluno, dsTelefone=numero)


def _to_time(raw_value):
    if not raw_value:
        return None
    if isinstance(raw_value, str):
        try:
            return datetime.strptime(raw_value.strip(), "%H:%M").time()
        except ValueError:
            return None
    return raw_value


def _add_months(base_date, months):
    year = base_date.year + (base_date.month - 1 + months) // 12
    month = (base_date.month - 1 + months) % 12 + 1
    day = min(base_date.day, 28)
    return date(year, month, day)


def _add_years(base_date, years):
    return _add_months(base_date, years * 12)


def _first_last_day_month(ref_date):
    first = ref_date.replace(day=1)
    last_day = calendar.monthrange(ref_date.year, ref_date.month)[1]
    last = ref_date.replace(day=last_day)
    return first, last


@login_required
def gerar_horarios_studio(request):
    if request.method != "POST":
        return redirect("horarios_studio_list")
    unidade_id = request.POST.get("unidade") or ""
    tipo_id = request.POST.get("tipoServico") or ""
    profissional_id = request.POST.get("profissional") or ""
    dias = request.POST.getlist("dias")
    inicio = _to_time(request.POST.get("horaInicio"))
    fim = _to_time(request.POST.get("horaFim"))
    intervalo = int(request.POST.get("intervalo") or 0)
    capacidade = request.POST.get("capacidade") or ""
    try:
        capacidade = int(capacidade) if capacidade else None
    except ValueError:
        capacidade = None
    if not (unidade_id and tipo_id and dias and inicio and fim and intervalo):
        messages.error(request, "Preencha unidade, tipo de servico, dias e horarios.")
        return redirect("horarios_studio_list")
    try:
        unidade_id = int(unidade_id)
        tipo_id = int(tipo_id)
    except ValueError:
        messages.error(request, "Unidade ou tipo de servico invalido.")
        return redirect("horarios_studio_list")
    prof_id = None
    if profissional_id:
        try:
            prof_id = int(profissional_id)
        except ValueError:
            prof_id = None
    start_minutes = inicio.hour * 60 + inicio.minute
    end_minutes = fim.hour * 60 + fim.minute
    if end_minutes <= start_minutes:
        messages.error(request, "Horario final deve ser maior que o inicial.")
        return redirect("horarios_studio_list")
    if intervalo <= 0:
        messages.error(request, "Intervalo invalido.")
        return redirect("horarios_studio_list")
    max_cd = models.HorarioStudio.objects.order_by("-cdHorario").values_list("cdHorario", flat=True).first() or 0
    created = 0
    for dia in dias:
        try:
            dia_int = int(dia)
        except ValueError:
            continue
        current = start_minutes
        while current + intervalo <= end_minutes:
            hora_inicio = datetime.strptime(f"{current // 60:02d}:{current % 60:02d}", "%H:%M").time()
            next_min = current + intervalo
            hora_fim = datetime.strptime(f"{next_min // 60:02d}:{next_min % 60:02d}", "%H:%M").time()
            exists = models.HorarioStudio.objects.filter(
                unidade_id=unidade_id,
                tipoServico_id=tipo_id,
                profissional_id=prof_id,
                diaSemana=dia_int,
                horaInicio=hora_inicio,
                horaFim=hora_fim,
            ).exists()
            if not exists:
                max_cd += 1
                models.HorarioStudio.objects.create(
                    cdHorario=max_cd,
                    unidade_id=unidade_id,
                    tipoServico_id=tipo_id,
                    profissional_id=prof_id,
                    diaSemana=dia_int,
                    horaInicio=hora_inicio,
                    horaFim=hora_fim,
                    capacidade=capacidade,
                )
                created += 1
            current = next_min
    if created:
        messages.success(request, f"{created} horarios criados.")
    else:
        messages.info(request, "Nenhum horario novo criado.")
    return redirect("horarios_studio_list")


@login_required
def gerar_horarios_funcionamento(request):
    if request.method != "POST":
        return redirect("horarios_funcionamento_list")
    unidade_id = request.POST.get("unidade") or ""
    tipos_ids = _parse_horario_servicos(request.POST)
    dias = request.POST.getlist("dias")
    inicio = _to_time(request.POST.get("horaInicio"))
    fim = _to_time(request.POST.get("horaFim"))
    ativo = request.POST.get("ativo") in ("1", "on", "true", "True")
    if not (unidade_id and dias and inicio and fim):
        messages.error(request, "Preencha unidade, dias e horarios.")
        return redirect("horarios_funcionamento_list")
    try:
        unidade_id = int(unidade_id)
    except ValueError:
        messages.error(request, "Unidade invalida.")
        return redirect("horarios_funcionamento_list")
    if fim <= inicio:
        messages.error(request, "Horario final deve ser maior que o inicial.")
        return redirect("horarios_funcionamento_list")
    created = 0
    for dia in dias:
        try:
            dia_int = int(dia)
        except ValueError:
            continue
        qs = models.HorarioFuncionamento.objects.filter(
            unidade_id=unidade_id,
            diaSemana=dia_int,
            horaInicio=inicio,
            horaFim=fim,
        )
        if tipos_ids:
            qs = qs.filter(
                Q(tipoServico_id__in=tipos_ids)
                | Q(tipos_servico__in=tipos_ids)
                | Q(tipoServico__isnull=True, tipos_servico__isnull=True)
            )
        else:
            qs = qs.filter(tipoServico__isnull=True, tipos_servico__isnull=True)
        exists = qs.distinct().exists()
        if exists:
            continue
        obj = models.HorarioFuncionamento.objects.create(
            unidade_id=unidade_id,
            tipoServico_id=tipos_ids[0] if tipos_ids else None,
            diaSemana=dia_int,
            horaInicio=inicio,
            horaFim=fim,
            ativo=ativo,
        )
        if tipos_ids:
            obj.tipos_servico.set(tipos_ids)
        created += 1
    if created:
        messages.success(request, f"{created} horarios criados.")
    else:
        messages.info(request, "Nenhum horario novo criado.")
    return redirect("horarios_funcionamento_list")


@login_required
def atualizar_horario_funcionamento(request, pk):
    obj = get_object_or_404(models.HorarioFuncionamento, pk=pk)
    if request.method != "POST":
        return redirect("horarios_funcionamento_list")
    unidade_id = request.POST.get("unidade") or obj.unidade_id
    tipos_ids = _parse_horario_servicos(request.POST)
    dias = request.POST.getlist("dias")
    inicio = _to_time(request.POST.get("horaInicio")) or obj.horaInicio
    fim = _to_time(request.POST.get("horaFim")) or obj.horaFim
    ativo = request.POST.get("ativo") in ("1", "on", "true", "True")
    try:
        unidade_id = int(unidade_id)
    except ValueError:
        messages.error(request, "Unidade invalida.")
        return redirect("horarios_funcionamento_list")
    if fim <= inicio:
        messages.error(request, "Horario final deve ser maior que o inicial.")
        return redirect("horarios_funcionamento_list")
    if not dias:
        dias = [str(obj.diaSemana)]
    dias_int = []
    for dia in dias:
        try:
            dias_int.append(int(dia))
        except ValueError:
            continue
    if not dias_int:
        dias_int = [obj.diaSemana]
    primary_day = dias_int[0]
    obj.unidade_id = unidade_id
    obj.tipoServico_id = tipos_ids[0] if tipos_ids else None
    obj.diaSemana = primary_day
    obj.horaInicio = inicio
    obj.horaFim = fim
    obj.ativo = ativo
    obj.save()
    if tipos_ids:
        obj.tipos_servico.set(tipos_ids)
    else:
        obj.tipos_servico.clear()

    created = 0
    for dia in dias_int[1:]:
        qs = models.HorarioFuncionamento.objects.filter(
            unidade_id=unidade_id,
            diaSemana=dia,
            horaInicio=inicio,
            horaFim=fim,
        )
        if tipos_ids:
            qs = qs.filter(
                Q(tipoServico_id__in=tipos_ids)
                | Q(tipos_servico__in=tipos_ids)
                | Q(tipoServico__isnull=True, tipos_servico__isnull=True)
            )
        else:
            qs = qs.filter(tipoServico__isnull=True, tipos_servico__isnull=True)
        exists = qs.distinct().exists()
        if exists:
            continue
        extra = models.HorarioFuncionamento.objects.create(
            unidade_id=unidade_id,
            tipoServico_id=tipos_ids[0] if tipos_ids else None,
            diaSemana=dia,
            horaInicio=inicio,
            horaFim=fim,
            ativo=ativo,
        )
        if tipos_ids:
            extra.tipos_servico.set(tipos_ids)
        created += 1
    if created:
        messages.success(request, f"{created} horario(s) extras criados.")
    else:
        messages.success(request, "Horario atualizado.")
    return redirect("horarios_funcionamento_list")


@login_required
def criar_bloqueio_agenda(request):
    if request.method != "POST":
        return redirect("bloqueios_list")
    unidade_id = request.POST.get("unidade") or ""
    tipo_id = request.POST.get("tipoServico") or ""
    profissional_id = request.POST.get("profissional") or ""
    recorrente = request.POST.get("recorrente") in ("1", "on", "true", "True")
    dias = request.POST.getlist("dias")
    data_inicio = request.POST.get("dataInicio") or ""
    data_fim = request.POST.get("dataFim") or ""
    inicio = _to_time(request.POST.get("horaInicio"))
    fim = _to_time(request.POST.get("horaFim"))
    motivo = request.POST.get("motivo", "").strip()
    ativo = request.POST.get("ativo") in ("1", "on", "true", "True")

    try:
        unidade_id = int(unidade_id)
    except ValueError:
        messages.error(request, "Unidade invalida.")
        return redirect("bloqueios_list")
    tipo_servico_id = None
    if tipo_id:
        try:
            tipo_servico_id = int(tipo_id)
        except ValueError:
            tipo_servico_id = None
    prof_id = None
    if profissional_id:
        try:
            prof_id = int(profissional_id)
        except ValueError:
            prof_id = None
    if not data_inicio:
        messages.error(request, "Informe a data de inicio.")
        return redirect("bloqueios_list")
    try:
        data_inicio_dt = datetime.strptime(data_inicio, "%Y-%m-%d").date()
    except ValueError:
        messages.error(request, "Data de inicio invalida.")
        return redirect("bloqueios_list")
    data_fim_dt = None
    if data_fim:
        try:
            data_fim_dt = datetime.strptime(data_fim, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "Data final invalida.")
            return redirect("bloqueios_list")
    if not inicio or not fim or fim <= inicio:
        messages.error(request, "Horario final deve ser maior que o inicial.")
        return redirect("bloqueios_list")

    created = 0
    if recorrente:
        if not dias:
            messages.error(request, "Selecione os dias da semana.")
            return redirect("bloqueios_list")
        for dia in dias:
            try:
                dia_int = int(dia)
            except ValueError:
                continue
            models.BloqueioAgenda.objects.create(
                unidade_id=unidade_id,
                tipoServico_id=tipo_servico_id,
                profissional_id=prof_id,
                recorrente=True,
                diaSemana=dia_int,
                dataInicio=data_inicio_dt,
                dataFim=data_fim_dt,
                horaInicio=inicio,
                horaFim=fim,
                motivo=motivo,
                ativo=ativo,
            )
            created += 1
    else:
        models.BloqueioAgenda.objects.create(
            unidade_id=unidade_id,
            tipoServico_id=tipo_servico_id,
            profissional_id=prof_id,
            recorrente=False,
            diaSemana=None,
            dataInicio=data_inicio_dt,
            dataFim=data_fim_dt,
            horaInicio=inicio,
            horaFim=fim,
            motivo=motivo,
            ativo=ativo,
        )
        created = 1
    messages.success(request, f"{created} bloqueio(s) criado(s).")
    return redirect("bloqueios_list")


@login_required
def atualizar_bloqueio_agenda(request, pk):
    obj = get_object_or_404(models.BloqueioAgenda, pk=pk)
    if request.method != "POST":
        return redirect("bloqueios_list")
    unidade_id = request.POST.get("unidade") or obj.unidade_id
    tipo_id = request.POST.get("tipoServico") or ""
    profissional_id = request.POST.get("profissional") or ""
    recorrente = request.POST.get("recorrente") in ("1", "on", "true", "True")
    dias = request.POST.getlist("dias")
    data_inicio = request.POST.get("dataInicio") or obj.dataInicio.strftime("%Y-%m-%d")
    data_fim = request.POST.get("dataFim") or ""
    inicio = _to_time(request.POST.get("horaInicio")) or obj.horaInicio
    fim = _to_time(request.POST.get("horaFim")) or obj.horaFim
    motivo = request.POST.get("motivo", "").strip()
    ativo = request.POST.get("ativo") in ("1", "on", "true", "True")

    try:
        unidade_id = int(unidade_id)
    except ValueError:
        messages.error(request, "Unidade invalida.")
        return redirect("bloqueios_list")
    tipo_servico_id = None
    if tipo_id:
        try:
            tipo_servico_id = int(tipo_id)
        except ValueError:
            tipo_servico_id = None
    prof_id = None
    if profissional_id:
        try:
            prof_id = int(profissional_id)
        except ValueError:
            prof_id = None
    try:
        data_inicio_dt = datetime.strptime(data_inicio, "%Y-%m-%d").date()
    except ValueError:
        messages.error(request, "Data de inicio invalida.")
        return redirect("bloqueios_list")
    data_fim_dt = None
    if data_fim:
        try:
            data_fim_dt = datetime.strptime(data_fim, "%Y-%m-%d").date()
        except ValueError:
            messages.error(request, "Data final invalida.")
            return redirect("bloqueios_list")
    if not inicio or not fim or fim <= inicio:
        messages.error(request, "Horario final deve ser maior que o inicial.")
        return redirect("bloqueios_list")

    if recorrente and not dias:
        dias = [str(obj.diaSemana)] if obj.diaSemana is not None else []

    if recorrente and dias:
        dias_int = []
        for dia in dias:
            try:
                dias_int.append(int(dia))
            except ValueError:
                continue
        if not dias_int:
            messages.error(request, "Selecione os dias da semana.")
            return redirect("bloqueios_list")
        primary_day = dias_int[0]
        obj.unidade_id = unidade_id
        obj.tipoServico_id = tipo_servico_id
        obj.profissional_id = prof_id
        obj.recorrente = True
        obj.diaSemana = primary_day
        obj.dataInicio = data_inicio_dt
        obj.dataFim = data_fim_dt
        obj.horaInicio = inicio
        obj.horaFim = fim
        obj.motivo = motivo
        obj.ativo = ativo
        obj.save()
        created = 0
        for dia in dias_int[1:]:
            models.BloqueioAgenda.objects.create(
                unidade_id=unidade_id,
                tipoServico_id=tipo_servico_id,
                profissional_id=prof_id,
                recorrente=True,
                diaSemana=dia,
                dataInicio=data_inicio_dt,
                dataFim=data_fim_dt,
                horaInicio=inicio,
                horaFim=fim,
                motivo=motivo,
                ativo=ativo,
            )
            created += 1
        if created:
            messages.success(request, f"{created} bloqueio(s) extras criados.")
        else:
            messages.success(request, "Bloqueio atualizado.")
        return redirect("bloqueios_list")

    obj.unidade_id = unidade_id
    obj.tipoServico_id = tipo_servico_id
    obj.profissional_id = prof_id
    obj.recorrente = False
    obj.diaSemana = None
    obj.dataInicio = data_inicio_dt
    obj.dataFim = data_fim_dt
    obj.horaInicio = inicio
    obj.horaFim = fim
    obj.motivo = motivo
    obj.ativo = ativo
    obj.save()
    messages.success(request, "Bloqueio atualizado.")
    return redirect("bloqueios_list")


def list_view(request, model, form_class, title, allow_modal=True, extra_context=None):
    if model in _admin_only_models() and _is_professor_user(request.user):
        messages.error(request, "Sem permissao para acessar esta area.")
        return redirect("dashboard")
    query = request.GET.get("q", "").strip()
    order = request.GET.get("order", "id")
    show_inativos = request.GET.get("inativos") in {"1", "true", "on", "sim"}
    page_size_key = f"list_page_size_{model._meta.model_name}"
    page_size_raw = request.GET.get("page_size")
    if page_size_raw is None:
        page_size = request.session.get(page_size_key, 100)
    else:
        try:
            page_size = int(page_size_raw)
        except (TypeError, ValueError):
            page_size = 100
        if page_size not in (10, 25, 50, 100):
            page_size = 100
        request.session[page_size_key] = page_size
    if page_size not in (10, 25, 50, 100):
        page_size = 100
    if model is models.Plano and not request.GET.get("order"):
        order = "dsPlano"
    if model is models.Aluno and not request.GET.get("order"):
        order = "dsNome"
    qs = model.objects.all()
    if model is models.ContasReceber:
        qs = qs.select_related(
            "contrato",
            "contrato__cdAluno",
            "contrato__cdPlano",
            "contrato__cdPlano__subcategoria_receita",
            "reserva",
            "reserva__aluno",
            "reserva__pacote_avulso",
            "reserva__pacote_avulso__plano",
        )
    if model is models.Contrato:
        qs = qs.select_related("cdAluno", "cdPlano", "cdUnidade")
    if model is models.HorarioFuncionamento:
        qs = qs.select_related("unidade", "tipoServico").prefetch_related("tipos_servico")
    if model is models.Aluno:
        qs = qs.annotate(
            aulas_reservadas_count=Count(
                "reserva",
                filter=Q(reserva__status="RESERVADA"),
                distinct=True,
            )
        )
        if not show_inativos:
            qs = qs.filter(status="ATIVO")
    if query:
        if model is models.Aluno:
            qs = qs.filter(
                Q(dsNome__icontains=query)
                | Q(dsCPF__icontains=query)
                | Q(dsRg__icontains=query)
                | Q(telefones__dsTelefone__icontains=query)
                | Q(id__icontains=query)
            )
            qs = qs.distinct()
        else:
            field_name = model._meta.fields[1].name
            qs = qs.filter(Q(**{f"{field_name}__icontains": query}) | Q(id__icontains=query))
    if order:
        qs = qs.order_by(order)
    paginator = Paginator(qs, page_size)
    page = paginator.get_page(request.GET.get("page"))
    page_range = list(paginator.get_elided_page_range(page.number, on_each_side=1, on_ends=1))
    query_params = request.GET.copy()
    query_params.pop("page", None)
    pagination_query = query_params.urlencode()
    display_fields = [
        {"name": field.name, "label": str(field.verbose_name)}
        for field in model._meta.fields
        if not field.primary_key and not field.name.startswith("cd")
    ][:3]
    if model is models.Aluno:
        display_fields = [
            {"name": "dsNome", "label": "Nome"},
            {"name": "dsCPF", "label": "CPF"},
            {"name": "dsRg", "label": "RG"},
            {"name": "telefone", "label": "Telefone"},
            {"name": "status", "label": "Status"},
        ]
    if model is models.ContasReceber:
        display_fields = [
            {"name": "competencia", "label": "Competencia"},
            {"name": "dtVencimento", "label": "Vencimento"},
            {"name": "dtPagamento", "label": "Pagamento"},
            {"name": "status", "label": "Status"},
            {"name": "valor", "label": "Valor"},
        ]
    if model is models.Plano:
        display_fields = [
            {"name": "dsPlano", "label": "Plano"},
            {"name": "aulas_por_semana", "label": "Aulas por semana"},
            {"name": "duracao_meses", "label": "Duracao (meses)"},
            {"name": "recorrencia", "label": "Recorrencia"},
        ]
    if model is models.ModeloContrato:
        display_fields = [
            {"name": "dsNome", "label": "Nome"},
            {"name": "ativo", "label": "Ativo"},
        ]
    if model is models.TermoUso:
        display_fields = [
            {"name": "nome", "label": "Nome"},
        ]
    if model is models.Contrato:
        display_fields = [
            {"name": "cdPlano", "label": "Plano"},
            {"name": "recorrencia", "label": "Recorrencia"},
            {"name": "valor_parcela", "label": "Valor (parcela)"},
            {"name": "valor_total", "label": "Valor total"},
            {"name": "status", "label": "Status"},
        ]
    if model is models.HorarioFuncionamento:
        display_fields = [
            {"name": "unidade", "label": "Unidade"},
            {"name": "servicos_resumo", "label": "Servicos"},
            {"name": "dia_semana_label", "label": "Dia"},
            {"name": "horaInicio", "label": "Inicio"},
            {"name": "horaFim", "label": "Fim"},
            {"name": "ativo", "label": "Ativo"},
        ]
    edit_forms = {}
    if allow_modal:
        for obj in page:
            edit_forms[obj.id] = form_class(instance=obj)

    context = {
        "title": title,
        "page": page,
        "form": form_class(),
        "model_name": model._meta.model_name,
        "breadcrumbs": [("Home", reverse("dashboard")), (title, "#")],
        "active_menu": _active_menu(request.path),
        "allow_modal": allow_modal,
        "display_fields": display_fields,
        "edit_forms": edit_forms,
        "page_size": page_size,
        "page_range": page_range,
        "pagination_query": pagination_query,
    }
    if model is models.Aluno:
        address_map = {}
        phones_map = {}
        for obj in page:
            address_map[obj.id] = obj.cdEndereco
            phones = list(obj.telefones.values_list("dsTelefone", flat=True))
            phones_map[obj.id] = ", ".join(phones) if phones else "-"
        context.update({"address_map": address_map, "phones_map": phones_map, "show_inativos": show_inativos})
    if model is models.Contrato:
        context.update(
            {
                "alunos": models.Aluno.objects.all(),
                "planos": models.Plano.objects.filter(is_avulso=False),
                "unidades": models.Unidade.objects.all(),
                "profissionais": models.Profissional.objects.all(),
            }
        )
    if extra_context:
        context.update(extra_context)
    return render(request, "generic/list.html", context)


@login_required
def modelos_contrato_list(request):
    return list_view(
        request,
        models.ModeloContrato,
        forms.ModeloContratoForm,
        "Modelos de Contrato",
        allow_modal=False,
        extra_context={
            "variables": CONTRACT_TEMPLATE_VARIABLES,
            "variable_target": "conteudo_html",
        },
    )


@login_required
def modelos_contrato_create(request):
    return _modelos_contrato_editor(request)


@login_required
def modelos_contrato_edit(request, pk):
    return _modelos_contrato_editor(request, pk=pk)


def _modelos_contrato_editor(request, pk=None):
    obj = get_object_or_404(models.ModeloContrato, pk=pk) if pk else None
    form = forms.ModeloContratoForm(request.POST or None, instance=obj)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Modelo de contrato salvo com sucesso.")
            return redirect("modelos_contrato_list")
        messages.error(request, "Revise os campos antes de salvar.")
    context = {
        "title": "Novo modelo de contrato" if obj is None else "Editar modelo de contrato",
        "subtitle": "Monte o contrato com variaveis dinamicas e salve para uso imediato.",
        "form": form,
        "object": obj,
        "variables": CONTRACT_TEMPLATE_VARIABLES,
        "variable_target": "conteudo_html",
        "breadcrumbs": [("Home", reverse("dashboard")), ("Modelos de Contrato", reverse("modelos_contrato_list")), ("Editor", "#")],
        "active_menu": "contratos",
    }
    return render(request, "contratos/modelo_contrato_editor.html", context)


@login_required
def termos_list(request):
    return list_view(
        request,
        models.TermoUso,
        forms.TermoUsoForm,
        "Termo de Uso",
        allow_modal=False,
        extra_context={
            "variables": CONTRACT_TEMPLATE_VARIABLES,
            "variable_target": "dsTermoUso",
        },
    )


@login_required
def termos_create(request):
    return _termos_editor(request)


@login_required
def termos_edit(request, pk):
    return _termos_editor(request, pk=pk)


def _termos_editor(request, pk=None):
    obj = get_object_or_404(models.TermoUso, pk=pk) if pk else None
    form = forms.TermoUsoForm(request.POST or None, instance=obj)
    if request.method == "POST":
        if form.is_valid():
            form.save()
            messages.success(request, "Termo de uso salvo com sucesso.")
            return redirect("termos_list")
        messages.error(request, "Revise os campos antes de salvar.")
    context = {
        "title": "Novo termo de uso" if obj is None else "Editar termo de uso",
        "subtitle": "Monte o termo com variaveis dinamicas e salve para uso imediato.",
        "form": form,
        "object": obj,
        "variables": CONTRACT_TEMPLATE_VARIABLES,
        "variable_target": "dsTermoUso",
        "breadcrumbs": [("Home", reverse("dashboard")), ("Termo de Uso", reverse("termos_list")), ("Editor", "#")],
        "active_menu": "cadastros",
    }
    return render(request, "cadastros/termo_uso_editor.html", context)


@login_required
def horarios_studio_list(request):
    extra_context = {
        "unidades": models.Unidade.objects.all(),
        "tipos_servico": models.TipoServico.objects.all(),
        "profissionais": models.Profissional.objects.all(),
        "dias_semana": models.HorarioStudio.DIAS_SEMANA,
    }
    return list_view(
        request,
        models.HorarioStudio,
        forms.HorarioStudioForm,
        "Horario do Studio",
        extra_context=extra_context,
    )


@login_required
def horarios_funcionamento_list(request):
    extra_context = {
        "unidades": models.Unidade.objects.all(),
        "tipos_servico": models.TipoServico.objects.all(),
        "dias_semana": models.HorarioStudio.DIAS_SEMANA,
    }
    return list_view(
        request,
        models.HorarioFuncionamento,
        forms.HorarioFuncionamentoForm,
        "Horario de Funcionamento",
        extra_context=extra_context,
    )


@login_required
def bloqueios_list(request):
    extra_context = {
        "unidades": models.Unidade.objects.all(),
        "tipos_servico": models.TipoServico.objects.all(),
        "profissionais": models.Profissional.objects.all(),
        "dias_semana": models.HorarioStudio.DIAS_SEMANA,
    }
    return list_view(
        request,
        models.BloqueioAgenda,
        forms.BloqueioAgendaForm,
        "Bloqueios de Agenda",
        extra_context=extra_context,
    )


@login_required
def contas_pagar_list(request):
    bloqueio = _professor_block(request)
    if bloqueio:
        return bloqueio
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    status = request.GET.get("status", "").strip()
    fornecedor_id = request.GET.get("fornecedor", "").strip()
    categoria_id = request.GET.get("categoria", "").strip()
    subcategoria_id = request.GET.get("subcategoria", "").strip()
    if not inicio or not fim:
        first, last = _first_last_day_month(today)
        if not inicio:
            inicio = first.strftime("%Y-%m-%d")
        if not fim:
            fim = last.strftime("%Y-%m-%d")

    qs = models.ContasPagar.objects.select_related("cdFornecedor", "cdCategoria", "cdSubcategoria").all()
    if fornecedor_id:
        qs = qs.filter(cdFornecedor_id=fornecedor_id)
    if categoria_id:
        qs = qs.filter(cdCategoria_id=categoria_id)
    if subcategoria_id:
        qs = qs.filter(cdSubcategoria_id=subcategoria_id)
    if inicio:
        try:
            inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
            qs = qs.filter(dtVencimento__gte=inicio_dt)
        except ValueError:
            pass
    if fim:
        try:
            fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
            qs = qs.filter(dtVencimento__lte=fim_dt)
        except ValueError:
            pass
    if status:
        if status == "ATRASADO":
            qs = qs.filter(dtVencimento__lt=today).exclude(status__in=["PAGO", "CANCELADO"])
        else:
            qs = qs.filter(status=status)

    qs = qs.order_by("dtVencimento", "id")
    paginator = Paginator(qs, 10)
    page = paginator.get_page(request.GET.get("page"))
    edit_forms = {obj.id: forms.ContasPagarForm(instance=obj) for obj in page}
    for obj in page:
        if obj.status == "PAGO":
            display_status = "PAGO"
        elif obj.status == "CANCELADO":
            display_status = "CANCELADO"
        elif obj.dtVencimento and obj.dtVencimento < today:
            display_status = "ATRASADO"
        else:
            display_status = "AGENDADO"
        obj.display_status = display_status

    query_params = request.GET.copy()
    if query_params.get("page"):
        query_params.pop("page")
    context = {
        "title": "Contas a Pagar",
        "page": page,
        "form": forms.ContasPagarForm(),
        "edit_forms": edit_forms,
        "fornecedores": models.Fornecedor.objects.all(),
        "categorias": models.Categoria.objects.all(),
        "subcategorias": models.Subcategoria.objects.all(),
        "filtros": {
            "inicio": inicio,
            "fim": fim,
            "status": status,
            "fornecedor": fornecedor_id,
            "categoria": categoria_id,
            "subcategoria": subcategoria_id,
        },
        "today": today.strftime("%Y-%m-%d"),
        "pagination_query": query_params.urlencode(),
        "breadcrumbs": [("Home", reverse("dashboard")), ("Contas a Pagar", "#")],
        "active_menu": "financeiro",
    }
    return render(request, "financeiro/contas_pagar_list.html", context)


def _filtrar_contas_pagar(request):
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    status = request.GET.get("status", "").strip()
    fornecedor_id = request.GET.get("fornecedor", "").strip()
    categoria_id = request.GET.get("categoria", "").strip()
    subcategoria_id = request.GET.get("subcategoria", "").strip()

    qs = models.ContasPagar.objects.select_related("cdFornecedor", "cdCategoria", "cdSubcategoria").all()
    if fornecedor_id:
        qs = qs.filter(cdFornecedor_id=fornecedor_id)
    if categoria_id:
        qs = qs.filter(cdCategoria_id=categoria_id)
    if subcategoria_id:
        qs = qs.filter(cdSubcategoria_id=subcategoria_id)
    if inicio:
        try:
            inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
            qs = qs.filter(dtVencimento__gte=inicio_dt)
        except ValueError:
            pass
    if fim:
        try:
            fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
            qs = qs.filter(dtVencimento__lte=fim_dt)
        except ValueError:
            pass
    if status:
        if status == "ATRASADO":
            qs = qs.filter(dtVencimento__lt=today).exclude(status__in=["PAGO", "CANCELADO"])
        else:
            qs = qs.filter(status=status)
    return qs.order_by("dtVencimento", "id")


def _parse_fluxo_caixa_periodo(request):
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    if not inicio or not fim:
        first, last = _first_last_day_month(today)
        if not inicio:
            inicio = first.strftime("%Y-%m-%d")
        if not fim:
            fim = last.strftime("%Y-%m-%d")
    try:
        inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
    except ValueError:
        inicio_dt = today.replace(day=1)
        inicio = inicio_dt.strftime("%Y-%m-%d")
    try:
        fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
    except ValueError:
        fim_dt = today
        fim = fim_dt.strftime("%Y-%m-%d")
    if fim_dt < inicio_dt:
        fim_dt = inicio_dt
        fim = fim_dt.strftime("%Y-%m-%d")
    return inicio, fim, inicio_dt, fim_dt


def _shift_month_clamped(base_date, months):
    year = base_date.year + (base_date.month - 1 + months) // 12
    month = (base_date.month - 1 + months) % 12 + 1
    day = min(base_date.day, calendar.monthrange(year, month)[1])
    return date(year, month, day)


def _build_fluxo_caixa_data_for_period(inicio_dt, fim_dt, conta_filtro="all"):
    contas_ativas = models.ContaBancaria.objects.filter(ativo=True).order_by("banco")
    conta_selecionada = None
    conta_filtro = (str(conta_filtro or "all")).strip() or "all"
    if conta_filtro != "all":
        conta_selecionada = contas_ativas.filter(pk=conta_filtro).first()
        if not conta_selecionada:
            conta_filtro = "all"

    if conta_filtro == "all":
        contas_base = contas_ativas
        movimentos_anteriores = models.MovimentoConta.objects.filter(conta__ativo=True, data__lt=inicio_dt)
    else:
        contas_base = contas_ativas.filter(pk=conta_selecionada.pk)
        movimentos_anteriores = models.MovimentoConta.objects.filter(conta=conta_selecionada, data__lt=inicio_dt)

    receitas_qs = (
        models.ContasReceber.objects.select_related("contrato", "contrato__cdAluno", "contrato__cdPlano", "contrato__cdUnidade")
        .exclude(status="CANCELADO")
        .filter(dtVencimento__range=(inicio_dt, fim_dt))
        .order_by("dtVencimento", "id")
    )
    despesas_qs = (
        models.ContasPagar.objects.select_related("cdFornecedor", "cdCategoria", "cdSubcategoria")
        .exclude(status="CANCELADO")
        .filter(dtVencimento__range=(inicio_dt, fim_dt))
        .order_by("dtVencimento", "id")
    )

    saldo_inicial_total = contas_base.aggregate(total=Sum("saldo_inicial"))["total"] or Decimal("0")
    entradas_anteriores = movimentos_anteriores.filter(tipo="ENTRADA").aggregate(total=Sum("valor"))["total"] or Decimal("0")
    saidas_anteriores = movimentos_anteriores.filter(tipo="SAIDA").aggregate(total=Sum("valor"))["total"] or Decimal("0")
    saldo_bancario_base = saldo_inicial_total + entradas_anteriores - saidas_anteriores

    receita_por_dia = defaultdict(Decimal)
    despesa_por_dia = defaultdict(Decimal)
    lancamentos = []

    for item in receitas_qs:
        valor = item.valor or Decimal("0")
        receita_por_dia[item.dtVencimento] += valor
        contrato = item.contrato if item.contrato_id else None
        reserva = item.reserva if item.reserva_id else None
        if contrato:
            cliente = contrato.cdAluno.dsNome if contrato.cdAluno_id else "Sem aluno"
            plano = contrato.cdPlano.dsPlano if contrato.cdPlano_id else "Sem plano"
            descricao = f"Contrato #{contrato.cdContrato} - {cliente}"
        elif reserva:
            cliente = reserva.aluno.dsNome if reserva.aluno_id else "Sem aluno"
            aula = reserva.aulaSessao
            horario = aula.horaInicio.strftime("%H:%M") if aula and aula.horaInicio else ""
            servico = aula.tipoServico.dsTipoServico if aula and aula.tipoServico_id else "Aula"
            plano = servico
            descricao = f"Aula avulsa - {cliente}" + (f" ({horario})" if horario else "")
        else:
            plano = "Lancamento manual"
            descricao = f"Receita #{item.id}"
        lancamentos.append(
            {
                "data": item.dtVencimento,
                "tipo": "Receita",
                "origem": "Contas a Receber",
                "descricao": descricao,
                "detalhe": plano,
                "status": item.status,
                "valor": valor,
                "sort_tipo": 0,
                "sort_id": item.id,
            }
        )

    for item in despesas_qs:
        valor = item.valor or Decimal("0")
        despesa_por_dia[item.dtVencimento] += valor
        fornecedor = str(item.cdFornecedor) if item.cdFornecedor_id else "Sem fornecedor"
        categoria = str(item.cdCategoria) if item.cdCategoria_id else "Sem categoria"
        lancamentos.append(
            {
                "data": item.dtVencimento,
                "tipo": "Despesa",
                "origem": "Contas a Pagar",
                "descricao": fornecedor,
                "detalhe": categoria if not item.cdSubcategoria_id else f"{categoria} / {item.cdSubcategoria}",
                "status": item.status,
                "valor": valor,
                "sort_tipo": 1,
                "sort_id": item.id,
            }
        )

    lancamentos.sort(key=lambda row: (row["data"], row["sort_tipo"], row["sort_id"]))

    saldo_diario = []
    saldo_acumulado = saldo_bancario_base
    total_receitas = Decimal("0")
    total_despesas = Decimal("0")
    cursor = inicio_dt
    while cursor <= fim_dt:
        receita_dia = receita_por_dia.get(cursor, Decimal("0"))
        despesa_dia = despesa_por_dia.get(cursor, Decimal("0"))
        total_receitas += receita_dia
        total_despesas += despesa_dia
        saldo_acumulado += receita_dia - despesa_dia
        saldo_diario.append(
            {
                "data": cursor,
                "receitas": receita_dia,
                "despesas": despesa_dia,
                "saldo": saldo_acumulado,
            }
        )
        cursor += timedelta(days=1)

    saldo_projetado_final = saldo_bancario_base + total_receitas - total_despesas

    saldo_semanal = []
    semana_inicio = inicio_dt
    saldo_semana_acumulado = saldo_bancario_base
    while semana_inicio <= fim_dt:
        semana_fim = min(semana_inicio + timedelta(days=6), fim_dt)
        receitas_semana = sum(
            item["receitas"] for item in saldo_diario if semana_inicio <= item["data"] <= semana_fim
        )
        despesas_semana = sum(
            item["despesas"] for item in saldo_diario if semana_inicio <= item["data"] <= semana_fim
        )
        saldo_semana_acumulado += receitas_semana - despesas_semana
        saldo_semanal.append(
            {
                "inicio": semana_inicio,
                "fim": semana_fim,
                "receitas": receitas_semana,
                "despesas": despesas_semana,
                "saldo": saldo_semana_acumulado,
            }
        )
        semana_inicio = semana_fim + timedelta(days=1)

    return {
        "conta_filtro": conta_filtro,
        "conta_selecionada": conta_selecionada,
        "contas_ativas": contas_ativas,
        "saldo_bancario_base": saldo_bancario_base,
        "saldo_projetado_final": saldo_projetado_final,
        "total_receitas": total_receitas,
        "total_despesas": total_despesas,
        "lancamentos": lancamentos,
        "saldo_diario": saldo_diario,
        "saldo_semanal": saldo_semanal,
    }


def _build_fluxo_caixa_data(request):
    inicio, fim, inicio_dt, fim_dt = _parse_fluxo_caixa_periodo(request)
    conta_filtro = request.GET.get("conta", "all").strip() or "all"
    today = timezone.now().date()
    periodo_atual = _build_fluxo_caixa_data_for_period(inicio_dt, fim_dt, conta_filtro=conta_filtro)
    prev_inicio_dt = _shift_month_clamped(inicio_dt, -1)
    prev_fim_dt = _shift_month_clamped(fim_dt, -1)
    periodo_anterior = _build_fluxo_caixa_data_for_period(prev_inicio_dt, prev_fim_dt, conta_filtro=conta_filtro)
    if conta_filtro == "all":
        conta_label = "Todas as contas"
    else:
        conta_label = periodo_atual["conta_selecionada"].__str__() if periodo_atual["conta_selecionada"] else "Todas as contas"

    comparativo_mes_anterior = {
        "periodo": f"{prev_inicio_dt.strftime('%d/%m/%Y')} a {prev_fim_dt.strftime('%d/%m/%Y')}",
        "receitas": periodo_anterior["total_receitas"],
        "despesas": periodo_anterior["total_despesas"],
        "saldo": periodo_anterior["saldo_projetado_final"],
        "delta_receitas": periodo_atual["total_receitas"] - periodo_anterior["total_receitas"],
        "delta_despesas": periodo_atual["total_despesas"] - periodo_anterior["total_despesas"],
        "delta_saldo": periodo_atual["saldo_projetado_final"] - periodo_anterior["saldo_projetado_final"],
    }

    return {
        "inicio": inicio,
        "fim": fim,
        "inicio_dt": inicio_dt,
        "fim_dt": fim_dt,
        "today": today,
        "conta_filtro": conta_filtro,
        "conta_label": conta_label,
        "conta_selecionada": periodo_atual["conta_selecionada"],
        "contas_ativas": periodo_atual["contas_ativas"],
        "saldo_bancario_base": periodo_atual["saldo_bancario_base"],
        "saldo_projetado_final": periodo_atual["saldo_projetado_final"],
        "total_receitas": periodo_atual["total_receitas"],
        "total_despesas": periodo_atual["total_despesas"],
        "lancamentos": periodo_atual["lancamentos"],
        "saldo_diario": periodo_atual["saldo_diario"],
        "saldo_semanal": periodo_atual["saldo_semanal"],
        "comparativo_mes_anterior": comparativo_mes_anterior,
    }


@login_required
def conta_bancaria_view(request):
    bloqueio = _professor_block(request)
    if bloqueio:
        return bloqueio
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    conta_id = request.GET.get("conta", "").strip()
    tipo = request.GET.get("tipo", "").strip()
    if not inicio or not fim:
        first, last = _first_last_day_month(today)
        if not inicio:
            inicio = first.strftime("%Y-%m-%d")
        if not fim:
            fim = last.strftime("%Y-%m-%d")

    contas = models.ContaBancaria.objects.filter(ativo=True).order_by("banco")
    conta_selecionada = contas.filter(pk=conta_id).first() if conta_id else contas.first()
    movimentos_qs = models.MovimentoConta.objects.select_related("conta").all()
    if conta_selecionada:
        movimentos_qs = movimentos_qs.filter(conta=conta_selecionada)
    if tipo:
        movimentos_qs = movimentos_qs.filter(tipo=tipo)
    if inicio:
        try:
            inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
            movimentos_qs = movimentos_qs.filter(data__gte=inicio_dt)
        except ValueError:
            inicio_dt = None
    else:
        inicio_dt = None
    if fim:
        try:
            fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
            movimentos_qs = movimentos_qs.filter(data__lte=fim_dt)
        except ValueError:
            fim_dt = None
    else:
        fim_dt = None
    movimentos = movimentos_qs.order_by("-data", "-id")
    total_entrada = movimentos_qs.filter(tipo="ENTRADA").aggregate(total=Sum("valor"))["total"] or 0
    total_saida = movimentos_qs.filter(tipo="SAIDA").aggregate(total=Sum("valor"))["total"] or 0
    saldo_inicial = conta_selecionada.saldo_inicial if conta_selecionada else 0
    saldo_atual = saldo_inicial + total_entrada - total_saida

    context = {
        "title": "Conta Bancaria",
        "contas": contas,
        "conta_selecionada": conta_selecionada,
        "movimentos": movimentos,
        "form_conta": forms.ContaBancariaForm(),
        "form_movimento": forms.MovimentoContaForm(),
        "filtros": {"inicio": inicio, "fim": fim, "conta": conta_id, "tipo": tipo},
        "total_entrada": total_entrada,
        "total_saida": total_saida,
        "saldo_inicial": saldo_inicial,
        "saldo_atual": saldo_atual,
        "today": today.strftime("%Y-%m-%d"),
        "breadcrumbs": [("Home", reverse("dashboard")), ("Conta Bancaria", "#")],
        "active_menu": "financeiro",
    }
    return render(request, "financeiro/conta_bancaria.html", context)


@login_required
def fluxo_caixa_view(request):
    bloqueio = _professor_block(request)
    if bloqueio:
        return bloqueio
    data = _build_fluxo_caixa_data(request)
    chart_labels = [item["data"].strftime("%d/%m") for item in data["saldo_diario"]]
    chart_receitas = [float(item["receitas"]) for item in data["saldo_diario"]]
    chart_despesas = [float(item["despesas"]) for item in data["saldo_diario"]]
    chart_saldo = [float(item["saldo"]) for item in data["saldo_diario"]]
    weekly_labels = [f"{item['inicio'].strftime('%d/%m')} - {item['fim'].strftime('%d/%m')}" for item in data["saldo_semanal"]]
    weekly_receitas = [float(item["receitas"]) for item in data["saldo_semanal"]]
    weekly_despesas = [float(item["despesas"]) for item in data["saldo_semanal"]]
    weekly_saldo = [float(item["saldo"]) for item in data["saldo_semanal"]]

    context = {
        "title": "Fluxo de Caixa",
        "inicio": data["inicio"],
        "fim": data["fim"],
        "conta_filtro": data["conta_filtro"],
        "conta_label": data["conta_label"],
        "saldo_bancario_base": data["saldo_bancario_base"],
        "saldo_projetado_final": data["saldo_projetado_final"],
        "total_receitas": data["total_receitas"],
        "total_despesas": data["total_despesas"],
        "lancamentos": data["lancamentos"],
        "saldo_semanal": data["saldo_semanal"],
        "comparativo_mes_anterior": data["comparativo_mes_anterior"],
        "chart_labels": json.dumps(chart_labels),
        "chart_receitas": json.dumps(chart_receitas),
        "chart_despesas": json.dumps(chart_despesas),
        "chart_saldo": json.dumps(chart_saldo),
        "weekly_labels": json.dumps(weekly_labels),
        "weekly_receitas": json.dumps(weekly_receitas),
        "weekly_despesas": json.dumps(weekly_despesas),
        "weekly_saldo": json.dumps(weekly_saldo),
        "contas_ativas": data["contas_ativas"],
        "breadcrumbs": [("Home", reverse("dashboard")), ("Financeiro", "#"), ("Fluxo de Caixa", "#")],
        "active_menu": "financeiro",
    }
    return render(request, "financeiro/fluxo_caixa.html", context)


@login_required
def exportar_fluxo_caixa_excel(request):
    data = _build_fluxo_caixa_data(request)
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Periodo inicio", data["inicio"]])
    ws.append(["Periodo fim", data["fim"]])
    ws.append(["Conta", data["conta_label"]])
    ws.append(["Saldo bancario base", float(data["saldo_bancario_base"])])
    ws.append(["Receitas previstas", float(data["total_receitas"])])
    ws.append(["Despesas previstas", float(data["total_despesas"])])
    ws.append(["Saldo projetado final", float(data["saldo_projetado_final"])])

    ws_comp = wb.create_sheet("Comparativo")
    comp = data["comparativo_mes_anterior"]
    ws_comp.append(["Periodo anterior", comp["periodo"]])
    ws_comp.append(["Receitas anteriores", float(comp["receitas"])])
    ws_comp.append(["Despesas anteriores", float(comp["despesas"])])
    ws_comp.append(["Saldo anterior", float(comp["saldo"])])
    ws_comp.append(["Delta receitas", float(comp["delta_receitas"])])
    ws_comp.append(["Delta despesas", float(comp["delta_despesas"])])
    ws_comp.append(["Delta saldo", float(comp["delta_saldo"])])

    ws_lanc = wb.create_sheet("Lancamentos")
    ws_lanc.append(["Data", "Tipo", "Origem", "Descricao", "Detalhe", "Status", "Valor"])
    for item in data["lancamentos"]:
        ws_lanc.append([
            item["data"].strftime("%d/%m/%Y"),
            item["tipo"],
            item["origem"],
            item["descricao"],
            item["detalhe"],
            item["status"],
            float(item["valor"]),
        ])

    ws_sem = wb.create_sheet("Saldo semanal")
    ws_sem.append(["Semana", "Receitas", "Despesas", "Saldo acumulado"])
    for item in data["saldo_semanal"]:
        ws_sem.append([
            f"{item['inicio'].strftime('%d/%m/%Y')} - {item['fim'].strftime('%d/%m/%Y')}",
            float(item["receitas"]),
            float(item["despesas"]),
            float(item["saldo"]),
        ])

    ws_diario = wb.create_sheet("Diario")
    ws_diario.append(["Data", "Receitas", "Despesas", "Saldo acumulado"])
    for item in data["saldo_diario"]:
        ws_diario.append([
            item["data"].strftime("%d/%m/%Y"),
            float(item["receitas"]),
            float(item["despesas"]),
            float(item["saldo"]),
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="fluxo-de-caixa.xlsx"'
    return response


@login_required
def exportar_fluxo_caixa_pdf(request):
    data = _build_fluxo_caixa_data(request)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.enums import TA_LEFT
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4), title="Fluxo de Caixa")
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="SmallLeft", parent=styles["BodyText"], alignment=TA_LEFT, fontSize=8, leading=10))

    story = [
        Paragraph("Fluxo de Caixa", styles["Title"]),
        Paragraph(f"Periodo {datetime.strptime(data['inicio'], '%Y-%m-%d').strftime('%d/%m/%Y')} a {datetime.strptime(data['fim'], '%Y-%m-%d').strftime('%d/%m/%Y')}", styles["Normal"]),
        Paragraph(f"Conta: {data['conta_label']}", styles["Normal"]),
        Spacer(1, 10),
    ]

    resumo = Table(
        [
            ["Saldo bancario base", f"R$ {data['saldo_bancario_base']}"],
            ["Receitas previstas", f"R$ {data['total_receitas']}"],
            ["Despesas previstas", f"R$ {data['total_despesas']}"],
            ["Saldo projetado final", f"R$ {data['saldo_projetado_final']}"],
        ],
        colWidths=[180, 120],
    )
    resumo.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5f5f5")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(resumo)
    story.append(Spacer(1, 12))

    linhas = [["Data", "Tipo", "Origem", "Descricao", "Status", "Valor"]]
    for item in data["lancamentos"][:120]:
        linhas.append([
            item["data"].strftime("%d/%m/%Y"),
            item["tipo"],
            item["origem"],
            item["descricao"][:40],
            item["status"],
            f"R$ {item['valor']}",
        ])
    tabela = Table(linhas, repeatRows=1, colWidths=[60, 55, 95, 220, 70, 70])
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(Paragraph("Lancamentos", styles["Heading2"]))
    story.append(tabela)
    story.append(Spacer(1, 12))

    comp = data["comparativo_mes_anterior"]
    resumo_comp = Table(
        [
            ["Periodo anterior", comp["periodo"]],
            ["Receitas anteriores", f"R$ {comp['receitas']}"],
            ["Despesas anteriores", f"R$ {comp['despesas']}"],
            ["Saldo anterior", f"R$ {comp['saldo']}"],
            ["Delta receitas", f"R$ {comp['delta_receitas']}"],
            ["Delta despesas", f"R$ {comp['delta_despesas']}"],
            ["Delta saldo", f"R$ {comp['delta_saldo']}"],
        ],
        colWidths=[180, 120],
    )
    resumo_comp.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f5f5f5")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
                ("FONTSIZE", (0, 0), (-1, -1), 9),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]
        )
    )
    story.append(Paragraph("Comparativo com o mes anterior", styles["Heading2"]))
    story.append(resumo_comp)
    story.append(Spacer(1, 12))

    linhas_sem = [["Semana", "Receitas", "Despesas", "Saldo"]]
    for item in data["saldo_semanal"]:
        linhas_sem.append([
            f"{item['inicio'].strftime('%d/%m/%Y')} - {item['fim'].strftime('%d/%m/%Y')}",
            f"R$ {item['receitas']}",
            f"R$ {item['despesas']}",
            f"R$ {item['saldo']}",
        ])
    tabela_sem = Table(linhas_sem, repeatRows=1, colWidths=[160, 90, 90, 90])
    tabela_sem.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f0f0f0")),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(Paragraph("Saldo semanal", styles["Heading2"]))
    story.append(tabela_sem)

    doc.build(story)
    buffer.seek(0)
    response = HttpResponse(buffer.read(), content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="fluxo-de-caixa.pdf"'
    return response


@login_required
def criar_conta_bancaria(request):
    if request.method != "POST":
        return redirect("conta_bancaria")
    data = request.POST.copy()
    data = _inject_cd_value(models.ContaBancaria, data)
    form = forms.ContaBancariaForm(data)
    if form.is_valid():
        form.save()
        messages.success(request, "Conta bancaria criada.")
    else:
        messages.error(request, "Verifique os erros.")
    return redirect("conta_bancaria")


@login_required
def criar_movimento_conta(request):
    if request.method != "POST":
        return redirect("conta_bancaria")
    data = request.POST.copy()
    form = forms.MovimentoContaForm(data, files=request.FILES or None)
    if form.is_valid():
        form.save()
        messages.success(request, "Lancamento registrado.")
    else:
        messages.error(request, "Verifique os erros.")
    return redirect("conta_bancaria")


@login_required
def dre_view(request):
    bloqueio = _professor_block(request)
    if bloqueio:
        return bloqueio
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    if not inicio or not fim:
        first, last = _first_last_day_month(today)
        if not inicio:
            inicio = first.strftime("%Y-%m-%d")
        if not fim:
            fim = last.strftime("%Y-%m-%d")
    try:
        inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
    except ValueError:
        inicio_dt = today.replace(day=1)
    try:
        fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
    except ValueError:
        fim_dt = today

    receitas = models.ContasReceber.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))
    despesas = models.ContasPagar.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))
    total_receitas = receitas.aggregate(total=Sum("valor"))["total"] or 0
    total_despesas = despesas.aggregate(total=Sum("valor"))["total"] or 0
    resultado = total_receitas - total_despesas
    status_label = "Lucro" if resultado >= 0 else "Prejuizo"

    context = {
        "title": "DRE",
        "inicio": inicio,
        "fim": fim,
        "total_receitas": total_receitas,
        "total_despesas": total_despesas,
        "resultado": resultado,
        "status_label": status_label,
        "receitas": receitas.order_by("-dtPagamento")[:10],
        "despesas": despesas.order_by("-dtPagamento")[:10],
        "breadcrumbs": [("Home", reverse("dashboard")), ("DRE", "#")],
        "active_menu": "financeiro",
    }
    return render(request, "financeiro/dre.html", context)


@login_required
def dre_relatorio(request):
    bloqueio = _professor_block(request)
    if bloqueio:
        return bloqueio
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    if not inicio or not fim:
        first, last = _first_last_day_month(today)
        if not inicio:
            inicio = first.strftime("%Y-%m-%d")
        if not fim:
            fim = last.strftime("%Y-%m-%d")
    try:
        inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
    except ValueError:
        inicio_dt = today.replace(day=1)
    try:
        fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
    except ValueError:
        fim_dt = today

    prev_month_last = inicio_dt.replace(day=1) - timedelta(days=1)
    movimentos_prev = models.MovimentoConta.objects.filter(data__lte=prev_month_last)
    saldo_inicial_total = models.ContaBancaria.objects.filter(ativo=True).aggregate(total=Sum("saldo_inicial"))["total"] or 0
    entradas_prev = movimentos_prev.filter(tipo="ENTRADA").aggregate(total=Sum("valor"))["total"] or 0
    saidas_prev = movimentos_prev.filter(tipo="SAIDA").aggregate(total=Sum("valor"))["total"] or 0
    saldo_bancario_anterior = saldo_inicial_total + entradas_prev - saidas_prev

    receitas_qs = models.ContasReceber.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))
    despesas_qs = models.ContasPagar.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))

    receita_bruta = receitas_qs.aggregate(total=Sum("valor"))["total"] or 0
    deducoes = 0
    receita_liquida = receita_bruta - deducoes
    custo_direto = 0
    lucro_bruto = receita_liquida - custo_direto
    despesas_operacionais = despesas_qs.aggregate(total=Sum("valor"))["total"] or 0
    resultado_final = lucro_bruto - despesas_operacionais
    saldo_bancario_final = saldo_bancario_anterior + receita_bruta - despesas_operacionais

    receitas_por_plano = (
        receitas_qs.values("contrato__cdPlano__dsPlano")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    receitas_por_categoria = (
        receitas_qs.values("contrato__cdPlano__categoria_receita__dsCategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    receitas_por_subcategoria = (
        receitas_qs.values("contrato__cdPlano__subcategoria_receita__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_categoria = (
        despesas_qs.values("cdCategoria__dsCategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_subcategoria = (
        despesas_qs.values("cdSubcategoria__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_subcategoria = (
        despesas_qs.values("cdSubcategoria__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_subcategoria = (
        despesas_qs.values("cdSubcategoria__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_subcategoria = (
        despesas_qs.values("cdSubcategoria__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_subcategoria = (
        despesas_qs.values("cdSubcategoria__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_subcategoria = (
        despesas_qs.values("cdSubcategoria__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    receitas_labels = [item["contrato__cdPlano__dsPlano"] or "Sem plano" for item in receitas_por_plano]
    receitas_values = [float(item["total"] or 0) for item in receitas_por_plano]
    receitas_cat_labels = [item["contrato__cdPlano__categoria_receita__dsCategoria"] or "Sem categoria" for item in receitas_por_categoria]
    receitas_cat_values = [float(item["total"] or 0) for item in receitas_por_categoria]
    receitas_sub_labels = [item["contrato__cdPlano__subcategoria_receita__dsSubcategoria"] or "Sem subcategoria" for item in receitas_por_subcategoria]
    receitas_sub_values = [float(item["total"] or 0) for item in receitas_por_subcategoria]
    despesas_labels = [item["cdCategoria__dsCategoria"] or "Sem categoria" for item in despesas_por_categoria]
    despesas_values = [float(item["total"] or 0) for item in despesas_por_categoria]
    despesas_sub_labels = [item["cdSubcategoria__dsSubcategoria"] or "Sem subcategoria" for item in despesas_por_subcategoria]
    despesas_sub_values = [float(item["total"] or 0) for item in despesas_por_subcategoria]

    context = {
        "inicio": inicio,
        "fim": fim,
        "receita_bruta": receita_bruta,
        "deducoes": deducoes,
        "receita_liquida": receita_liquida,
        "custo_direto": custo_direto,
        "lucro_bruto": lucro_bruto,
        "despesas_operacionais": despesas_operacionais,
        "resultado_final": resultado_final,
        "resultado_label": "Lucro" if resultado_final >= 0 else "Prejuizo",
        "saldo_bancario_anterior": saldo_bancario_anterior,
        "saldo_bancario_final": saldo_bancario_final,
        "receitas_por_plano": receitas_por_plano,
        "receitas_por_categoria": receitas_por_categoria,
        "receitas_por_subcategoria": receitas_por_subcategoria,
        "despesas_por_categoria": despesas_por_categoria,
        "despesas_por_subcategoria": despesas_por_subcategoria,
        "chart_receitas_labels": json.dumps(receitas_labels),
        "chart_receitas_values": json.dumps(receitas_values),
        "chart_receitas_cat_labels": json.dumps(receitas_cat_labels),
        "chart_receitas_cat_values": json.dumps(receitas_cat_values),
        "chart_receitas_sub_labels": json.dumps(receitas_sub_labels),
        "chart_receitas_sub_values": json.dumps(receitas_sub_values),
        "chart_despesas_labels": json.dumps(despesas_labels),
        "chart_despesas_values": json.dumps(despesas_values),
        "chart_despesas_sub_labels": json.dumps(despesas_sub_labels),
        "chart_despesas_sub_values": json.dumps(despesas_sub_values),
        "breadcrumbs": [("Home", reverse("dashboard")), ("DRE", reverse("dre_view")), ("Relatorio", "#")],
        "active_menu": "financeiro",
    }
    return render(request, "financeiro/dre_relatorio.html", context)


@login_required
def exportar_dre_excel(request):
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    if not inicio or not fim:
        first, last = _first_last_day_month(today)
        if not inicio:
            inicio = first.strftime("%Y-%m-%d")
        if not fim:
            fim = last.strftime("%Y-%m-%d")
    try:
        inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
    except ValueError:
        inicio_dt = today.replace(day=1)
    try:
        fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
    except ValueError:
        fim_dt = today

    receitas_qs = models.ContasReceber.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))
    despesas_qs = models.ContasPagar.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))
    receita_bruta = receitas_qs.aggregate(total=Sum("valor"))["total"] or 0
    deducoes = 0
    receita_liquida = receita_bruta - deducoes
    custo_direto = 0
    lucro_bruto = receita_liquida - custo_direto
    despesas_operacionais = despesas_qs.aggregate(total=Sum("valor"))["total"] or 0
    resultado_final = lucro_bruto - despesas_operacionais

    receitas_por_plano = (
        receitas_qs.values("contrato__cdPlano__dsPlano")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    receitas_por_categoria = (
        receitas_qs.values("contrato__cdPlano__categoria_receita__dsCategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_categoria = (
        despesas_qs.values("cdCategoria__dsCategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )

    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumo"
    ws.append(["Periodo inicio", inicio])
    ws.append(["Periodo fim", fim])
    ws.append(["Receita bruta", float(receita_bruta)])
    ws.append(["Deducoes", float(deducoes)])
    ws.append(["Receita liquida", float(receita_liquida)])
    ws.append(["Custo direto", float(custo_direto)])
    ws.append(["Lucro bruto", float(lucro_bruto)])
    ws.append(["Despesas operacionais", float(despesas_operacionais)])
    ws.append(["Resultado final", float(resultado_final)])

    ws_receitas = wb.create_sheet("Receitas por plano")
    ws_receitas.append(["Plano", "Total"])
    for item in receitas_por_plano:
        ws_receitas.append([item["contrato__cdPlano__dsPlano"] or "Sem plano", float(item["total"] or 0)])

    ws_receitas_cat = wb.create_sheet("Receitas por categoria")
    ws_receitas_cat.append(["Categoria", "Total"])
    for item in receitas_por_categoria:
        ws_receitas_cat.append([item["contrato__cdPlano__categoria_receita__dsCategoria"] or "Sem categoria", float(item["total"] or 0)])

    ws_receitas_sub = wb.create_sheet("Receitas por subcategoria")
    ws_receitas_sub.append(["Subcategoria", "Total"])
    for item in receitas_por_subcategoria:
        ws_receitas_sub.append([item["contrato__cdPlano__subcategoria_receita__dsSubcategoria"] or "Sem subcategoria", float(item["total"] or 0)])

    ws_despesas = wb.create_sheet("Despesas por categoria")
    ws_despesas.append(["Categoria", "Total"])
    for item in despesas_por_categoria:
        ws_despesas.append([item["cdCategoria__dsCategoria"] or "Sem categoria", float(item["total"] or 0)])

    ws_despesas_sub = wb.create_sheet("Despesas por subcategoria")
    ws_despesas_sub.append(["Subcategoria", "Total"])
    for item in despesas_por_subcategoria:
        ws_despesas_sub.append([item["cdSubcategoria__dsSubcategoria"] or "Sem subcategoria", float(item["total"] or 0)])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="dre-completo.xlsx"'
    return response


@login_required
def exportar_dre_pdf(request):
    today = timezone.now().date()
    inicio = request.GET.get("inicio", "").strip()
    fim = request.GET.get("fim", "").strip()
    if not inicio or not fim:
        first, last = _first_last_day_month(today)
        if not inicio:
            inicio = first.strftime("%Y-%m-%d")
        if not fim:
            fim = last.strftime("%Y-%m-%d")
    try:
        inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
    except ValueError:
        inicio_dt = today.replace(day=1)
    try:
        fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
    except ValueError:
        fim_dt = today

    receitas_qs = models.ContasReceber.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))
    despesas_qs = models.ContasPagar.objects.filter(status="PAGO", dtPagamento__range=(inicio_dt, fim_dt))
    receita_bruta = receitas_qs.aggregate(total=Sum("valor"))["total"] or 0
    deducoes = 0
    receita_liquida = receita_bruta - deducoes
    custo_direto = 0
    lucro_bruto = receita_liquida - custo_direto
    despesas_operacionais = despesas_qs.aggregate(total=Sum("valor"))["total"] or 0
    resultado_final = lucro_bruto - despesas_operacionais

    receitas_por_plano = (
        receitas_qs.values("contrato__cdPlano__dsPlano")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    receitas_por_categoria = (
        receitas_qs.values("contrato__cdPlano__categoria_receita__dsCategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_categoria = (
        despesas_qs.values("cdCategoria__dsCategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )
    despesas_por_subcategoria = (
        despesas_qs.values("cdSubcategoria__dsSubcategoria")
        .annotate(total=Sum("valor"))
        .order_by("-total")
    )

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="DRE Completo")
    styles = getSampleStyleSheet()
    title = Paragraph("DRE Completo", styles["Title"])
    subtitle = Paragraph(f"Periodo {inicio_dt.strftime('%d/%m/%Y')} a {fim_dt.strftime('%d/%m/%Y')}", styles["Normal"])
    resumo = [
        ["Receita bruta", f"R$ {receita_bruta}"],
        ["Deducoes", f"R$ {deducoes}"],
        ["Receita liquida", f"R$ {receita_liquida}"],
        ["Custo direto", f"R$ {custo_direto}"],
        ["Lucro bruto", f"R$ {lucro_bruto}"],
        ["Despesas operacionais", f"R$ {despesas_operacionais}"],
        ["Resultado final", f"R$ {resultado_final}"],
    ]
    resumo_table = Table(resumo, colWidths=[220, 160])
    resumo_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4d4d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )

    receitas_table = [["Plano", "Total"]]
    for item in receitas_por_plano:
        receitas_table.append([item["contrato__cdPlano__dsPlano"] or "Sem plano", f"R$ {item['total']}"])
    receitas_table = Table(receitas_table, colWidths=[260, 120])
    receitas_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#e0f2fe")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )

    receitas_cat_table = [["Categoria", "Total"]]
    for item in receitas_por_categoria:
        receitas_cat_table.append([item["contrato__cdPlano__categoria_receita__dsCategoria"] or "Sem categoria", f"R$ {item['total']}"])
    receitas_cat_table = Table(receitas_cat_table, colWidths=[260, 120])
    receitas_cat_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#dcfce7")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )

    despesas_table = [["Categoria", "Total"]]
    for item in despesas_por_categoria:
        despesas_table.append([item["cdCategoria__dsCategoria"] or "Sem categoria", f"R$ {item['total']}"])
    despesas_table = Table(despesas_table, colWidths=[260, 120])
    despesas_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#fee2e2")),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ]
        )
    )

    elements = [
        title,
        subtitle,
        Spacer(1, 12),
        resumo_table,
        Spacer(1, 16),
        Paragraph("Receitas por plano", styles["Heading3"]),
        receitas_table,
        Spacer(1, 16),
        Paragraph("Receitas por categoria", styles["Heading3"]),
        receitas_cat_table,
        Spacer(1, 16),
        Paragraph("Receitas por subcategoria", styles["Heading3"]),
        Table(
            [["Subcategoria", "Total"]] + [
                [item["contrato__cdPlano__subcategoria_receita__dsSubcategoria"] or "Sem subcategoria", f"R$ {item['total']}"]
                for item in receitas_por_subcategoria
            ],
            colWidths=[260, 120],
        ),
        Spacer(1, 16),
        Paragraph("Despesas por categoria", styles["Heading3"]),
        despesas_table,
        Spacer(1, 16),
        Paragraph("Despesas por subcategoria", styles["Heading3"]),
        Table(
            [["Subcategoria", "Total"]] + [
                [item["cdSubcategoria__dsSubcategoria"] or "Sem subcategoria", f"R$ {item['total']}"]
                for item in despesas_por_subcategoria
            ],
            colWidths=[260, 120],
        ),
    ]
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="dre-completo.pdf"'
    return response

@login_required
def exportar_contas_pagar_excel(request):
    qs = _filtrar_contas_pagar(request)
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Contas a Pagar"
    ws.append(["Fornecedor", "Categoria", "Subcategoria", "Vencimento", "Pagamento", "Valor", "Status"])
    for f in qs:
        if f.status == "PAGO":
            display_status = "PAGO"
        elif f.status == "CANCELADO":
            display_status = "CANCELADO"
        elif f.dtVencimento and f.dtVencimento < timezone.now().date():
            display_status = "ATRASADO"
        else:
            display_status = "AGENDADO"
        ws.append(
            [
                str(f.cdFornecedor),
                str(f.cdCategoria),
                str(f.cdSubcategoria),
                f.dtVencimento.strftime("%d/%m/%Y") if f.dtVencimento else "",
                f.dtPagamento.strftime("%d/%m/%Y") if f.dtPagamento else "",
                float(f.valor),
                display_status,
            ]
        )
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="contas-a-pagar.xlsx"'
    return response


@login_required
def exportar_contas_pagar_pdf(request):
    qs = _filtrar_contas_pagar(request)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Contas a Pagar")
    styles = getSampleStyleSheet()
    title = Paragraph("Contas a Pagar", styles["Title"])
    data = [["Fornecedor", "Categoria", "Subcategoria", "Vencimento", "Pagamento", "Valor", "Status"]]
    for f in qs:
        if f.status == "PAGO":
            display_status = "PAGO"
        elif f.status == "CANCELADO":
            display_status = "CANCELADO"
        elif f.dtVencimento and f.dtVencimento < timezone.now().date():
            display_status = "ATRASADO"
        else:
            display_status = "AGENDADO"
        data.append(
            [
                str(f.cdFornecedor),
                str(f.cdCategoria),
                str(f.cdSubcategoria),
                f.dtVencimento.strftime("%d/%m/%Y") if f.dtVencimento else "-",
                f.dtPagamento.strftime("%d/%m/%Y") if f.dtPagamento else "-",
                f"R$ {f.valor}",
                display_status,
            ]
        )
    table = Table(data, repeatRows=1, colWidths=[90, 80, 90, 70, 70, 60, 70])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4d4d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f7fb")]),
            ]
        )
    )
    elements = [title, Spacer(1, 12), table]
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = 'attachment; filename="contas-a-pagar.pdf"'
    return response


@login_required
def _cor_status_reserva(r):
    if r.status == "RESERVADA" and r.confirmada_em:
        return ("#3b9ad9", "Confirmada")
    if r.status == "RESERVADA":
        return ("#f59e0b", "A confirmar")
    return {
        "CONCLUIDA": ("#16a34a", "Concluída"),
        "FALTOU_AVISOU": ("#eab308", "Falta avisada"),
        "FALTOU_SEM_AVISAR": ("#ef4444", "Falta"),
        "CANCELADA": ("#9ca3af", "Cancelada"),
        "PENDENTE": ("#a78bfa", "Pendente"),
    }.get(r.status, ("#94a3b8", r.status))


@login_required
def aulas_semana(request):
    services.concluir_aulas_passadas()
    week_str = request.GET.get("week", "").strip()
    profissional_id = request.GET.get("profissional", "").strip()
    profissional_id = _aplica_filtro_professor(request, profissional_id)
    try:
        ref = datetime.strptime(week_str, "%Y-%m-%d").date() if week_str else date.today()
    except ValueError:
        ref = date.today()
    # Sem semana escolhida e hoje e sabado/domingo -> mostra a proxima semana
    if not week_str and ref.weekday() >= 5:
        ref = ref + timedelta(days=7)
    week_start = ref - timedelta(days=ref.weekday())
    days = [week_start + timedelta(days=i) for i in range(6)]  # Seg a Sab
    week_end = days[-1]

    qs = models.AulaSessao.objects.select_related("tipoServico", "profissional", "unidade").filter(
        data__range=(week_start, week_end)
    )
    if profissional_id:
        qs = qs.filter(profissional_id=profissional_id)
    aulas = list(qs)
    reservas = models.Reserva.objects.filter(aulaSessao__in=aulas).select_related("aluno")
    res_by_aula = {}
    for r in reservas:
        res_by_aula.setdefault(r.aulaSessao_id, []).append(r)

    GRID_START, GRID_END, PPM = 6 * 60, 22 * 60, 0.8

    def _min(t):
        return t.hour * 60 + t.minute

    weekday_labels = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab"]
    grid_days = []
    for dia in days:
        raw = []
        for aula in [a for a in aulas if a.data == dia]:
            ini = _min(aula.horaInicio)
            fim = _min(aula.horaFim) if aula.horaFim else ini + 50
            tipo = aula.tipoServico.dsTipoServico if aula.tipoServico_id else "Aula"
            prof = aula.profissional.profissional if aula.profissional_id else ""
            hora = aula.horaInicio.strftime("%H:%M")
            alunos = res_by_aula.get(aula.id, [])
            if alunos:
                for r in alunos:
                    cor, label = _cor_status_reserva(r)
                    raw.append({"ini": ini, "fim": fim, "cor": cor, "label": label,
                                "aluno": r.aluno.dsNome, "tipo": tipo, "hora": hora, "prof": prof})
            else:
                raw.append({"ini": ini, "fim": fim, "cor": "#cbd5e1", "label": "Sem alunos",
                            "aluno": "", "tipo": tipo, "hora": hora, "prof": prof})
        raw.sort(key=lambda b: (b["ini"], b["fim"]))
        # clusters de blocos que se sobrepoem -> colunas lado a lado
        clusters, cur, cur_end = [], [], -1
        for b in raw:
            if cur and b["ini"] >= cur_end:
                clusters.append(cur)
                cur, cur_end = [], -1
            cur.append(b)
            cur_end = max(cur_end, b["fim"])
        if cur:
            clusters.append(cur)
        blocks = []
        for cl in clusters:
            cols_end = []
            for b in cl:
                placed = False
                for ci, end in enumerate(cols_end):
                    if b["ini"] >= end:
                        cols_end[ci] = b["fim"]
                        b["col"] = ci
                        placed = True
                        break
                if not placed:
                    b["col"] = len(cols_end)
                    cols_end.append(b["fim"])
            ncols = len(cols_end)
            wpct = 100.0 / ncols
            for b in cl:
                blocks.append({
                    **b,
                    "top": round((b["ini"] - GRID_START) * PPM, 1),
                    "height": round(max((b["fim"] - b["ini"]) * PPM, 22), 1),
                    "leftpct": round(b["col"] * wpct, 2),
                    "widthpct": round(wpct - 1.5, 2),
                })
        grid_days.append({"date": dia, "label": weekday_labels[dia.weekday()], "number": dia.day,
                          "is_today": dia == date.today(), "blocks": blocks})

    hours = [{"label": f"{h:02d}:00", "top": round((h * 60 - GRID_START) * PPM, 1)} for h in range(6, 23)]
    meses = ["Janeiro", "Fevereiro", "Marco", "Abril", "Maio", "Junho", "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]
    context = {
        "title": "Agenda",
        "grid_days": grid_days,
        "hours": hours,
        "grid_height": round((GRID_END - GRID_START) * PPM, 1),
        "week_start": week_start,
        "week_end": week_end,
        "prev_week": (week_start - timedelta(days=7)).isoformat(),
        "next_week": (week_start + timedelta(days=7)).isoformat(),
        "hoje": date.today().isoformat(),
        "month_label": f"{meses[week_start.month - 1]} {week_start.year}",
        "profissionais": models.Profissional.objects.all(),
        "profissional_id": profissional_id,
        "breadcrumbs": [("Home", reverse("dashboard")), ("Agenda", "#")],
        "active_menu": "agenda",
    }
    return render(request, "agenda/semana.html", context)


def aulas_list(request):
    services.concluir_aulas_passadas()
    qs = models.AulaSessao.objects.select_related("unidade", "tipoServico", "profissional")
    week_str = request.GET.get("week", "").strip()
    profissional_id = request.GET.get("profissional", "").strip()
    profissional_id = _aplica_filtro_professor(request, profissional_id)
    view_mode = request.GET.get("view", "week").strip().lower()
    if view_mode == "list":
        view_mode = "week"
    if view_mode not in {"week", "day", "month"}:
        view_mode = "week"
    try:
        ref_date = datetime.strptime(week_str, "%Y-%m-%d").date() if week_str else date.today()
    except ValueError:
        ref_date = date.today()

    if view_mode == "day":
        week_start = ref_date
        week_end = ref_date
        days = [ref_date]
        columns = 1
    elif view_mode == "month":
        first = ref_date.replace(day=1)
        if first.month == 12:
            last = date(first.year + 1, 1, 1) - timedelta(days=1)
        else:
            last = date(first.year, first.month + 1, 1) - timedelta(days=1)
        week_start = first
        week_end = last
        leading = first.weekday()
        days = [None] * leading + [first + timedelta(days=i) for i in range((last - first).days + 1)]
        columns = 7
    else:
        week_start = ref_date - timedelta(days=ref_date.weekday())
        week_end = week_start + timedelta(days=5)
        days = [week_start + timedelta(days=i) for i in range(6)]
        columns = 6

    qs = qs.filter(data__range=(week_start, week_end))
    if profissional_id:
        qs = qs.filter(profissional_id=profissional_id)

    aulas = list(qs.order_by("data", "horaInicio"))
    aula_ids = [aula.id for aula in aulas]
    reservas = (
        models.Reserva.objects.filter(aulaSessao_id__in=aula_ids)
        .select_related("aluno", "aulaSessao", "aulaSessao__profissional")
        .order_by("aulaSessao__data", "aulaSessao__horaInicio")
    )
    reservas_by_aula = {}
    for reserva in reservas:
        reservas_by_aula.setdefault(reserva.aulaSessao_id, []).append(reserva)

    aulas_by_day = {day: [] for day in days if day}
    for aula in aulas:
        aulas_by_day.setdefault(aula.data, []).append(aula)
    for aulas in aulas_by_day.values():
        aulas.sort(key=lambda x: x.horaInicio)

    weekday_labels = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sab", "Dom"]
    week_cards = []
    for dia in days:
        if not dia:
            week_cards.append({"is_placeholder": True})
            continue
        week_cards.append(
            {
                "date": dia,
                "label": weekday_labels[dia.weekday()],
                "number": dia.day,
                "is_today": dia == date.today(),
                "aulas": aulas_by_day.get(dia, []),
            }
        )

    prof_chips = []
    for prof in models.Profissional.objects.all():
        parts = [p for p in prof.profissional.split() if p]
        initials = "".join([p[0] for p in parts[:2]]).upper()
        prof_chips.append({"name": prof.profissional, "initials": initials})

    meses = [
        "Janeiro",
        "Fevereiro",
        "Marco",
        "Abril",
        "Maio",
        "Junho",
        "Julho",
        "Agosto",
        "Setembro",
        "Outubro",
        "Novembro",
        "Dezembro",
    ]
    month_label = f"{meses[week_start.month - 1]} {week_start.year}"

    context = {
        "title": "Aulas",
        "week_cards": week_cards,
        "week_start": week_start,
        "week_end": week_end,
        "prev_week": week_start - timedelta(days=7),
        "next_week": week_start + timedelta(days=7),
        "month_label": month_label,
        "view_mode": view_mode,
        "columns": columns,
        "form": forms.AulaSessaoForm(),
        "profissionais": models.Profissional.objects.all(),
        "unidades": models.Unidade.objects.all(),
        "prof_chips": prof_chips[:5],
        "reservas_by_aula": reservas_by_aula,
        "modelos_evolucao": models.ModeloEvolucao.objects.filter(ativo=True).order_by("titulo"),
        "is_professor": _is_professor_user(request.user),
        "breadcrumbs": [("Home", reverse("dashboard")), ("Aulas", "#")],
        "active_menu": "agenda",
    }
    return render(request, "agenda/aulas_list.html", context)


def _parse_date(value: str | None) -> date | None:
    if not value:
        return None
    try:
        return datetime.strptime(value, "%Y-%m-%d").date()
    except ValueError:
        return None


def _build_period_range(target: date | None, periodo: str | None) -> tuple[date, date]:
    base = target or date.today()
    if periodo == "amanha":
        start = base + timedelta(days=1)
        return start, start
    if periodo == "semana":
        start = base - timedelta(days=base.weekday())
        if base.weekday() >= 5:  # sabado/domingo -> proxima semana
            start = start + timedelta(days=7)
        end = start + timedelta(days=6)
        return start, end
    return base, base


def _map_status(reserva_status: str, inicio: datetime, fim: datetime) -> str:
    status_map = {
        "RESERVADA": "aguardando_chegar",
        "PENDENTE": "aguardando_chegar",
        "CONCLUIDA": "finalizada",
        "FALTOU_AVISOU": "faltou",
        "FALTOU_SEM_AVISAR": "faltou",
        "CANCELADA": "remarcada",
    }
    mapped = status_map.get(reserva_status, "aguardando_chegar")
    if mapped == "aguardando_chegar":
        now = timezone.localtime()
        if inicio <= now <= fim:
            return "em_aula"
    return mapped


def _horario_servicos_resumo(unidade_id, data_value, hora_inicio, hora_fim, fallback=None):
    if not unidade_id or not data_value or not hora_inicio or not hora_fim:
        return fallback or []
    horarios = (
        models.HorarioFuncionamento.objects.filter(
            unidade_id=unidade_id,
            diaSemana=data_value.weekday(),
            horaInicio__lte=hora_fim,
            horaFim__gte=hora_inicio,
            ativo=True,
        )
        .prefetch_related("tipos_servico")
        .select_related("tipoServico")
    )
    servicos = []
    seen = set()
    for horario in horarios:
        if horario.tipoServico_id and horario.tipoServico_id not in seen:
            seen.add(horario.tipoServico_id)
            servicos.append(horario.tipoServico.dsTipoServico if horario.tipoServico else str(horario.tipoServico_id))
        for servico in horario.tipos_servico.all():
            if servico.id in seen:
                continue
            seen.add(servico.id)
            servicos.append(servico.dsTipoServico)
    return servicos or (fallback or [])


def _horario_servicos_ids(unidade_id, data_value, hora_inicio, hora_fim, fallback=None):
    if not unidade_id or not data_value or not hora_inicio or not hora_fim:
        return fallback or []
    horarios = (
        models.HorarioFuncionamento.objects.filter(
            unidade_id=unidade_id,
            diaSemana=data_value.weekday(),
            horaInicio__lte=hora_fim,
            horaFim__gte=hora_inicio,
            ativo=True,
        )
        .prefetch_related("tipos_servico")
        .select_related("tipoServico")
    )
    servicos = []
    seen = set()
    for horario in horarios:
        if horario.tipoServico_id and horario.tipoServico_id not in seen:
            seen.add(horario.tipoServico_id)
            servicos.append(horario.tipoServico_id)
        for servico in horario.tipos_servico.all():
            if servico.id in seen:
                continue
            seen.add(servico.id)
            servicos.append(servico.id)
    return servicos or (fallback or [])


@login_required
def aulas_operacao_api(request):
    services.concluir_aulas_passadas()
    target = _parse_date(request.GET.get("data", "").strip())
    periodo = request.GET.get("periodo", "hoje").strip()
    start_date, end_date = _build_period_range(target, periodo)

    unidade_id = request.GET.get("unidade_id") or None
    profissional_id = request.GET.get("profissional_id") or None
    if _is_professor_user(request.user):
        prof = _get_profissional_for_user(request.user)
        profissional_id = str(prof.id) if prof else "0"
    status_filter = request.GET.get("status_aula") or None
    query = (request.GET.get("q") or "").strip()

    contrato_qs = (
        models.Contrato.objects.filter(
            cdAluno=OuterRef("aluno_id"),
            dtInicioContrato__lte=OuterRef("aulaSessao__data"),
            dtFimContrato__gte=OuterRef("aulaSessao__data"),
        )
        .order_by("-dtFimContrato", "-id")
    )

    telefone_qs = models.TelefoneAluno.objects.filter(cdAluno=OuterRef("aluno_id")).order_by("-dtCadastro", "-id")
    evolucao_qs = models.EvolucaoAluno.objects.filter(reserva_id=OuterRef("pk")).order_by("-dtEvolucao", "-id")
    cobranca_exists = Exists(
        models.ContasReceber.objects.filter(
            contrato__cdAluno=OuterRef("aluno_id"),
            status="ABERTO",
        )
    )

    reservas = (
        models.Reserva.objects.select_related(
            "aluno",
            "aulaSessao",
            "aulaSessao__unidade",
            "aulaSessao__profissional",
            "aulaSessao__tipoServico",
        )
        .annotate(
            plano_id=Subquery(contrato_qs.values("cdPlano_id")[:1]),
            plano_descricao=Subquery(contrato_qs.values("cdPlano__dsPlano")[:1]),
            aluno_telefone=Subquery(telefone_qs.values("dsTelefone")[:1]),
            ultima_evolucao=Subquery(evolucao_qs.values("texto")[:1]),
            ultima_evolucao_em=Subquery(evolucao_qs.values("dtEvolucao")[:1]),
            cobranca_pendente=cobranca_exists,
        )
        .filter(aulaSessao__data__range=(start_date, end_date))
    )

    if unidade_id:
        reservas = reservas.filter(aulaSessao__unidade_id=unidade_id)
    if profissional_id:
        reservas = reservas.filter(aulaSessao__profissional_id=profissional_id)
    if query:
        reservas = reservas.filter(
            Q(aluno__dsNome__icontains=query)
            | Q(aluno__dsCPF__icontains=query)
            | Q(aluno__telefones__dsTelefone__icontains=query)
        ).distinct()

    items = []
    for reserva in reservas.order_by("aulaSessao__data", "aulaSessao__horaInicio", "aluno__dsNome"):
        inicio = datetime.combine(reserva.aulaSessao.data, reserva.aulaSessao.horaInicio)
        fim = datetime.combine(reserva.aulaSessao.data, reserva.aulaSessao.horaFim)
        inicio = timezone.make_aware(inicio)
        fim = timezone.make_aware(fim)
        status_calc = _map_status(reserva.status, inicio, fim)
        servicos = _horario_servicos_resumo(
            reserva.aulaSessao.unidade_id if reserva.aulaSessao else None,
            reserva.aulaSessao.data if reserva.aulaSessao else None,
            reserva.aulaSessao.horaInicio if reserva.aulaSessao else None,
            reserva.aulaSessao.horaFim if reserva.aulaSessao else None,
            fallback=[reserva.aulaSessao.tipoServico.dsTipoServico] if reserva.aulaSessao and reserva.aulaSessao.tipoServico else [],
        )
        if status_filter and status_calc != status_filter:
            continue
        items.append(
            {
                "id": reserva.id,
                "aula_sessao_id": reserva.aulaSessao_id,
                "dt_inicio": inicio.isoformat(),
                "dt_fim": fim.isoformat(),
                "unidade_id": reserva.aulaSessao.unidade_id if reserva.aulaSessao else None,
                "unidade": reserva.aulaSessao.unidade.dsUnidade if reserva.aulaSessao and reserva.aulaSessao.unidade else None,
                "sala": None,
                "profissional": {
                    "id": reserva.aulaSessao.profissional_id if reserva.aulaSessao else None,
                    "nome": reserva.aulaSessao.profissional.profissional if reserva.aulaSessao and reserva.aulaSessao.profissional else None,
                },
                "aluno": {
                    "id": reserva.aluno_id,
                    "nome": reserva.aluno.dsNome,
                    "telefone": reserva.aluno_telefone,
                    "avatar_url": reserva.aluno.foto.url if reserva.aluno.foto else None,
                    "ficha_url": reverse("alunos_detail", args=[reserva.aluno_id]),
                },
                "plano": {"id": reserva.plano_id, "descricao": reserva.plano_descricao},
                "servicos": servicos,
                "servicos_resumo": ", ".join(servicos) if servicos else None,
                "status_aula": status_calc,
                "confirmacao": reserva.status != "PENDENTE",
                "flags": {
                    "tem_preliminares": bool(reserva.aluno.termo_aceite_em),
                    "cobranca_pendente": bool(reserva.cobranca_pendente),
                    "observacao_importante": False,
                },
                "ultima_evolucao": {
                    "texto": reserva.ultima_evolucao,
                    "dt_evolucao": reserva.ultima_evolucao_em.isoformat() if reserva.ultima_evolucao_em else None,
                },
            }
        )

    return JsonResponse(
        {
            "data_inicio": start_date.isoformat(),
            "data_fim": end_date.isoformat(),
            "total": len(items),
            "items": items,
        }
    )


def _parse_json_body(request):
    if request.content_type and "application/json" in request.content_type:
        try:
            return json.loads(request.body.decode("utf-8"))
        except json.JSONDecodeError:
            return {}
    return request.POST


def _normalize_cpf(value):
    digits = re.sub(r"\D", "", value or "")
    if len(digits) != 11:
        return digits, ""
    masked = f"{digits[:3]}.{digits[3:6]}.{digits[6:9]}-{digits[9:]}"
    return digits, masked


def _parse_totalpass_datetime(payload):
    slot = payload.get("slot") or {}
    date_raw = slot.get("date") or slot.get("start_at") or slot.get("startAt") or ""
    time_raw = slot.get("start_time") or slot.get("startTime") or slot.get("time") or ""

    if not date_raw:
        return None, "Data nao informada no payload."

    try:
        if "T" in date_raw:
            base = datetime.fromisoformat(date_raw.replace("Z", "+00:00"))
        else:
            base = datetime.strptime(date_raw, "%Y-%m-%d")
    except ValueError:
        return None, "Data invalida no payload."

    if time_raw:
        try:
            time_value = datetime.strptime(time_raw, "%H:%M").time()
            base = datetime.combine(base.date(), time_value)
        except ValueError:
            return None, "Hora invalida no payload."
    elif base.time() == datetime.min.time():
        return None, "Horario nao informado no payload."

    if timezone.is_naive(base):
        base = timezone.make_aware(base)
    return base, ""


def _get_totalpass_config(payload):
    place_id = ""
    place = payload.get("place") or {}
    for key in ["place", "id", "place_id", "uuid"]:
        if place.get(key):
            place_id = place.get(key)
            break
    if place_id:
        cfg = models.TotalpassConfiguracao.objects.filter(place_id=place_id).select_related("unidade").first()
        if cfg:
            return cfg
    unidade_id = getattr(settings, "TOTALPASS_UNIDADE_ID", "") or ""
    if unidade_id:
        return models.TotalpassConfiguracao.objects.filter(unidade_id=unidade_id).select_related("unidade").first()
    return models.TotalpassConfiguracao.objects.select_related("unidade").first()


def _resolve_totalpass_unidade(payload):
    cfg = _get_totalpass_config(payload)
    if cfg and cfg.unidade_id:
        return cfg.unidade
    unidade_id = getattr(settings, "TOTALPASS_UNIDADE_ID", "") or ""
    if unidade_id:
        return models.Unidade.objects.filter(pk=unidade_id).first()
    return models.Unidade.objects.filter(dsUnidade__iexact="Matriz").first() or models.Unidade.objects.filter(dsUnidade__icontains="Matriz").first()


def _resolve_totalpass_tipo_servico(title):
    if not title:
        return None
    return (
        models.TipoServico.objects.filter(dsTipoServico__iexact=title).first()
        or models.TipoServico.objects.filter(dsTipoServico__icontains=title).first()
    )


def _build_totalpass_payload_from_slot(slot):
    if not isinstance(slot, dict):
        return {}
    if slot.get("event") or slot.get("user"):
        if slot.get("slot"):
            return slot
        payload = dict(slot)
        payload["slot"] = slot.get("slot") or {}
        return payload
    event_title = slot.get("event_title") or slot.get("eventTitle") or slot.get("title") or ""
    user_name = slot.get("user_name") or slot.get("userName") or ""
    user_email = slot.get("user_email") or slot.get("userEmail") or ""
    user_doc = slot.get("document_number") or slot.get("documentNumber") or slot.get("userDocument") or ""
    user_phone = slot.get("phone") or slot.get("userPhone") or ""
    place_id = slot.get("place_id") or slot.get("placeId") or ""
    return {
        "event": {"id": slot.get("event_id") or slot.get("eventId") or "", "title": event_title},
        "place": {"place": place_id},
        "user": {
            "name": user_name,
            "email": user_email,
            "phone": user_phone,
            "document_number": user_doc,
            "document_type": "cpf",
        },
        "slot": {
            "id": slot.get("slot_id") or slot.get("id") or "",
            "status": slot.get("status") or slot.get("slot_status") or "",
            "date": slot.get("date") or slot.get("start_at") or slot.get("startAt") or "",
            "start_time": slot.get("start_time") or slot.get("startTime") or slot.get("time") or "",
        },
    }


def _process_totalpass_payload(payload, cfg, event_obj=None):
    event_id = (payload.get("event") or {}).get("id") or ""
    slot_id = (payload.get("slot") or {}).get("id") or ""
    user_document = (payload.get("user") or {}).get("document_number") or ""
    slot_status = (payload.get("slot") or {}).get("status") or ""

    if not event_obj:
        event_obj = models.TotalpassWebhookEvent.objects.create(
            event_id=event_id,
            slot_id=slot_id,
            user_document=user_document,
            status=slot_status,
            payload=payload or {},
        )
    else:
        event_obj.event_id = event_id
        event_obj.slot_id = slot_id
        event_obj.user_document = user_document
        event_obj.status = slot_status
        event_obj.payload = payload or {}
        event_obj.save(update_fields=["event_id", "slot_id", "user_document", "status", "payload"])

    dt_inicio, err = _parse_totalpass_datetime(payload)
    if err:
        event_obj.error = err
        event_obj.save(update_fields=["error"])
        return None, err

    if cfg and cfg.somente_dia:
        if dt_inicio.date() != timezone.localdate():
            event_obj.error = "Fora do dia atual."
            event_obj.save(update_fields=["error"])
            return None, "Fora do dia atual."

    unidade = cfg.unidade if cfg and cfg.unidade_id else _resolve_totalpass_unidade(payload)
    if not unidade:
        event_obj.error = "Unidade Matriz nao encontrada."
        event_obj.save(update_fields=["error"])
        return None, event_obj.error

    tipo_servico = _resolve_totalpass_tipo_servico((payload.get("event") or {}).get("title"))
    if not tipo_servico:
        event_obj.error = "Tipo de servico nao encontrado."
        event_obj.save(update_fields=["error"])
        return None, event_obj.error

    cpf_digits, cpf_mask = _normalize_cpf(user_document)
    aluno = (
        models.Aluno.objects.filter(dsCPF=cpf_mask).first()
        or models.Aluno.objects.filter(dsCPF=cpf_digits).first()
    )
    if not aluno and cfg and cfg.criar_aluno_automatico is False:
        event_obj.error = "Aluno nao encontrado pelo CPF."
        event_obj.save(update_fields=["error"])
        return None, event_obj.error
    if not aluno:
        nome = (payload.get("user") or {}).get("name") or "Aluno TotalPass"
        email = (payload.get("user") or {}).get("email") or ""
        telefone = (payload.get("user") or {}).get("phone") or ""
        max_cd = models.Aluno.objects.order_by("-cdAluno").values_list("cdAluno", flat=True).first() or 0
        aluno = models.Aluno.objects.create(
            cdAluno=max_cd + 1,
            dsNome=nome,
            dsCPF=cpf_mask or cpf_digits,
            dsEmail=email,
            cdUnidade=unidade,
        )
        if telefone:
            max_tel = models.TelefoneAluno.objects.order_by("-cdTelefone").values_list("cdTelefone", flat=True).first() or 0
            models.TelefoneAluno.objects.create(cdTelefone=max_tel + 1, cdAluno=aluno, dsTelefone=telefone)

    duracao_min = unidade.duracao_aula_minutos or 50
    hora_fim = (dt_inicio + timedelta(minutes=duracao_min)).time()

    aula, _ = models.AulaSessao.objects.get_or_create(
        unidade=unidade,
        tipoServico=tipo_servico,
        profissional=None,
        data=dt_inicio.date(),
        horaInicio=dt_inicio.time().replace(second=0, microsecond=0),
        horaFim=hora_fim,
        defaults={"capacidade": unidade.capacidade},
    )

    status_lower = str(slot_status).lower()
    if status_lower in ["canceled", "cancelled", "inactive"]:
        reserva = models.Reserva.objects.filter(aluno=aluno, aulaSessao=aula).first()
        if reserva:
            reserva.status = "CANCELADA"
            reserva.save(update_fields=["status"])
        event_obj.processed_at = timezone.now()
        event_obj.save(update_fields=["processed_at"])
        return reserva, ""

    capacidade = aula.capacidade_efetiva()
    total = models.Reserva.objects.filter(aulaSessao=aula, status="RESERVADA").count()
    if capacidade and total >= capacidade:
        event_obj.error = "Capacidade excedida."
        event_obj.save(update_fields=["error"])
        return None, event_obj.error

    reserva, created = models.Reserva.objects.get_or_create(
        aluno=aluno,
        aulaSessao=aula,
        defaults={"status": "RESERVADA"},
    )
    if not created and reserva.status != "RESERVADA":
        reserva.status = "RESERVADA"
        reserva.save(update_fields=["status"])

    event_obj.processed_at = timezone.now()
    event_obj.save(update_fields=["processed_at"])
    return reserva, ""


@login_required
def aula_evolucao_api(request, reserva_id):
    reserva = get_object_or_404(models.Reserva, pk=reserva_id)
    if request.method == "POST":
        payload = _parse_json_body(request)
        texto = (payload.get("texto") or "").strip()
        profissional_id = payload.get("profissional_id")
        finalizar = bool(payload.get("finalizar"))
        if not texto:
            return JsonResponse({"error": "Texto obrigatorio."}, status=400)

        profissional = None
        if profissional_id:
            profissional = models.Profissional.objects.filter(pk=profissional_id).first()
        if not profissional:
            profissional = reserva.aulaSessao.profissional if reserva.aulaSessao else None
        if not profissional:
            return JsonResponse({"error": "Profissional invalido."}, status=400)

        evolucao = models.EvolucaoAluno.objects.create(
            reserva=reserva,
            profissional=profissional,
            texto=texto,
            dtEvolucao=timezone.now(),
        )
        if finalizar:
            reserva.status = "CONCLUIDA"
            reserva.save(update_fields=["status"])
        return JsonResponse(
            {
                "id": evolucao.id,
                "reserva_id": reserva.id,
                "texto": evolucao.texto,
                "dt_evolucao": evolucao.dtEvolucao.isoformat(),
            }
        )

    evolucoes = (
        models.EvolucaoAluno.objects.filter(reserva__aluno=reserva.aluno)
        .select_related("profissional", "reserva", "reserva__aulaSessao")
        .order_by("-dtEvolucao")[:20]
    )
    return JsonResponse(
        {
            "items": [
                {
                    "id": e.id,
                    "texto": e.texto,
                    "dt_evolucao": e.dtEvolucao.isoformat(),
                    "profissional": e.profissional.profissional if e.profissional else None,
                    "reserva_id": e.reserva_id,
                    "data": e.reserva.aulaSessao.data.isoformat() if e.reserva and e.reserva.aulaSessao else None,
                }
                for e in evolucoes
            ]
        }
    )


@login_required
@require_POST
def aula_evolucao_enriquecer_api(request, reserva_id):
    get_object_or_404(models.Reserva, pk=reserva_id)
    return _enriquecer_evolucao_response(request)


@login_required
@require_POST
def evolucao_enriquecer_api(request):
    return _enriquecer_evolucao_response(request)


def _enriquecer_evolucao_response(request):
    payload = _parse_json_body(request)
    texto = (payload.get("texto") or "").strip()
    if not texto:
        return JsonResponse({"error": "Informe a evolucao antes de enriquecer."}, status=400)
    try:
        data = improve_evolution_text(texto)
    except GeminiError as exc:
        logger.warning("Falha ao enriquecer evolucao por IA: %s", exc)
        return JsonResponse({"error": str(exc)}, status=503)
    except Exception as exc:
        logger.exception("Falha ao enriquecer evolucao por IA.")
        return JsonResponse({"error": f"Nao foi possivel enriquecer a evolucao agora: {exc}"}, status=503)
    enriched = (data.get("texto") or "").strip()
    if not enriched:
        return JsonResponse({"error": "A IA nao retornou um texto valido."}, status=503)
    return JsonResponse({"texto": enriched})


@login_required
def aula_avaliacoes_api(request, reserva_id):
    reserva = get_object_or_404(models.Reserva, pk=reserva_id)
    if request.method == "POST":
        payload = _parse_json_body(request)
        texto = (payload.get("texto") or "").strip()
        profissional_id = payload.get("profissional_id")
        acao = (payload.get("acao") or "create").lower()
        avaliacao_id = payload.get("avaliacao_id")

        if acao in ["update", "delete"]:
            avaliacao = (
                models.AvaliacaoAluno.objects.filter(
                    pk=avaliacao_id,
                    reserva__aluno=reserva.aluno,
                )
                .select_related("profissional")
                .first()
            )
            if not avaliacao:
                return JsonResponse({"error": "Avaliacao nao encontrada."}, status=404)
            if acao == "delete":
                avaliacao.delete()
                return JsonResponse({"ok": True})
            if not texto:
                return JsonResponse({"error": "Texto obrigatorio."}, status=400)
            avaliacao.texto = texto
            avaliacao.save(update_fields=["texto"])
            return JsonResponse(
                {
                    "id": avaliacao.id,
                    "reserva_id": avaliacao.reserva_id,
                    "texto": avaliacao.texto,
                    "dt_avaliacao": avaliacao.dtAvaliacao.isoformat(),
                    "profissional": avaliacao.profissional.profissional if avaliacao.profissional else None,
                }
            )
        if not texto:
            return JsonResponse({"error": "Texto obrigatorio."}, status=400)

        profissional = None
        if profissional_id:
            profissional = models.Profissional.objects.filter(pk=profissional_id).first()
        if not profissional:
            profissional = reserva.aulaSessao.profissional if reserva.aulaSessao else None
        if not profissional:
            return JsonResponse({"error": "Profissional invalido."}, status=400)

        avaliacao = models.AvaliacaoAluno.objects.create(
            reserva=reserva,
            profissional=profissional,
            texto=texto,
            dtAvaliacao=timezone.now(),
        )
        return JsonResponse(
            {
                "id": avaliacao.id,
                "reserva_id": reserva.id,
                "texto": avaliacao.texto,
                "dt_avaliacao": avaliacao.dtAvaliacao.isoformat(),
            }
        )

    avaliacoes = (
        models.AvaliacaoAluno.objects.filter(reserva__aluno=reserva.aluno)
        .select_related("profissional", "reserva", "reserva__aulaSessao")
        .order_by("-dtAvaliacao")[:20]
    )
    return JsonResponse(
        {
            "items": [
                {
                    "id": a.id,
                    "texto": a.texto,
                    "dt_avaliacao": a.dtAvaliacao.isoformat(),
                    "profissional": a.profissional.profissional if a.profissional else None,
                    "reserva_id": a.reserva_id,
                    "data": a.reserva.aulaSessao.data.isoformat() if a.reserva and a.reserva.aulaSessao else None,
                }
                for a in avaliacoes
            ]
        }
    )


@login_required
def aula_cobranca_api(request, reserva_id):
    reserva = get_object_or_404(models.Reserva, pk=reserva_id)
    contas = (
        models.ContasReceber.objects.filter(contrato__cdAluno=reserva.aluno)
        .select_related("contrato")
        .order_by("-dtVencimento", "-id")
    )
    items = []
    for conta in contas:
        items.append(
            {
                "id": conta.id,
                "competencia": conta.competencia,
                "dt_vencimento": conta.dtVencimento.isoformat() if conta.dtVencimento else None,
                "dt_pagamento": conta.dtPagamento.isoformat() if conta.dtPagamento else None,
                "status": conta.status,
                "valor": float(conta.valor),
                "contrato": conta.contrato.cdContrato if conta.contrato else None,
                "baixar_url": reverse("contas_receber_baixar", args=[conta.id]),
                "recibo_url": reverse("contas_receber_recibo", args=[conta.id]),
                "excluir_url": reverse("contas_receber_excluir", args=[conta.id]),
            }
        )
    return JsonResponse({"items": items})


@login_required
def aula_historico_api(request, reserva_id):
    reserva = get_object_or_404(models.Reserva, pk=reserva_id)
    reservas = (
        models.Reserva.objects.filter(aluno=reserva.aluno)
        .select_related("aulaSessao", "aulaSessao__unidade", "aulaSessao__profissional", "aulaSessao__tipoServico")
        .order_by("-aulaSessao__data", "-aulaSessao__horaInicio")[:20]
    )
    items = []
    for item in reservas:
        aula = item.aulaSessao
        if not aula:
            continue
        inicio = timezone.make_aware(datetime.combine(aula.data, aula.horaInicio))
        fim = timezone.make_aware(datetime.combine(aula.data, aula.horaFim))
        status_calc = _map_status(item.status, inicio, fim)
        items.append(
            {
                "id": item.id,
                "data": aula.data.isoformat(),
                "hora_inicio": aula.horaInicio.strftime("%H:%M"),
                "hora_fim": aula.horaFim.strftime("%H:%M"),
                "unidade": aula.unidade.dsUnidade if aula.unidade else None,
                "profissional": aula.profissional.profissional if aula.profissional else None,
                "servico": aula.tipoServico.dsTipoServico if aula.tipoServico else None,
                "status": status_calc,
            }
        )
    return JsonResponse({"items": items})


@login_required
@require_POST
def aula_status_api(request, reserva_id):
    payload = _parse_json_body(request)
    acao = (payload.get("acao") or "").strip().lower()
    status_map = {
        "chegou": "RESERVADA",
        "iniciar": "RESERVADA",
        "finalizar": "CONCLUIDA",
        "faltou": "FALTOU_SEM_AVISAR",
        "remarcar": "CANCELADA",
    }
    # Desmarcar com reposicao: cancela a aula e libera uma reposicao (30 dias)
    # para o aluno reagendar, mesma regra do app do aluno.
    if acao == "desmarcar":
        if _is_professor_user(request.user):
            return JsonResponse({"error": "Sem permissao para desmarcar."}, status=403)
        reserva = get_object_or_404(models.Reserva.objects.select_related("aulaSessao"), pk=reserva_id)
        aula = reserva.aulaSessao
        reposicao_ate = (aula.data + timedelta(days=30)) if aula else None
        reserva.status = "CANCELADA"
        reserva.confirmada_em = None
        reserva.desmarcada_em = timezone.now()
        reserva.reposicao_ate = reposicao_ate
        reserva.save(update_fields=["status", "confirmada_em", "desmarcada_em", "reposicao_ate"])
        return JsonResponse(
            {
                "reserva_id": reserva.id,
                "status": reserva.status,
                "reposicao_ate": reposicao_ate.isoformat() if reposicao_ate else None,
            }
        )
    if acao not in status_map:
        return JsonResponse({"error": "Acao invalida."}, status=400)
    if acao == "remarcar" and _is_professor_user(request.user):
        return JsonResponse({"error": "Sem permissao para remarcar."}, status=403)
    reserva = get_object_or_404(models.Reserva, pk=reserva_id)
    reserva.status = status_map[acao]
    reserva.save(update_fields=["status"])
    return JsonResponse({"reserva_id": reserva.id, "status": reserva.status})


@login_required
@require_POST
def aula_reserva_excluir(request, reserva_id):
    """Exclui a aula do aluno (a reserva) direto pela agenda. Se o horario
    ficar sem nenhuma reserva, remove tambem a sessao."""
    if _is_professor_user(request.user):
        return JsonResponse({"error": "Sem permissao para excluir."}, status=403)
    reserva = get_object_or_404(models.Reserva.objects.select_related("aulaSessao"), pk=reserva_id)
    aula_id = reserva.aulaSessao_id
    reserva.delete()
    if aula_id and not models.Reserva.objects.filter(aulaSessao_id=aula_id).exists():
        models.AulaSessao.objects.filter(pk=aula_id).delete()
    return JsonResponse({"ok": True, "reserva_id": reserva_id})


@login_required
def aula_remarcar_api(request, reserva_id):
    if request.method != "POST":
        return JsonResponse({"error": "Metodo invalido."}, status=405)
    if _is_professor_user(request.user):
        return JsonResponse({"error": "Sem permissao para remarcar."}, status=403)
    payload = _parse_json_body(request)
    data_str = (payload.get("data") or "").strip()
    hora_str = (payload.get("hora_inicio") or "").strip()
    profissional_id = payload.get("profissional_id")
    if not data_str or not hora_str:
        return JsonResponse({"error": "Informe data e hora."}, status=400)
    try:
        nova_data = datetime.strptime(data_str, "%Y-%m-%d").date()
        nova_hora_inicio = datetime.strptime(hora_str, "%H:%M").time()
    except ValueError:
        return JsonResponse({"error": "Data ou hora invalida."}, status=400)

    reserva = get_object_or_404(models.Reserva.objects.select_related("aulaSessao", "aulaSessao__unidade", "aulaSessao__tipoServico", "aulaSessao__profissional"), pk=reserva_id)
    aula = reserva.aulaSessao
    if not aula:
        return JsonResponse({"error": "Aula nao encontrada."}, status=404)
    profissional = aula.profissional
    if profissional_id:
        profissional = models.Profissional.objects.filter(pk=profissional_id).first()
        if not profissional:
            return JsonResponse({"error": "Profissional invalido."}, status=400)

    duracao_min = aula.unidade.duracao_aula_minutos if aula.unidade else 50
    if aula.horaFim and aula.horaInicio:
        base_inicio = datetime.combine(date.today(), aula.horaInicio)
        base_fim = datetime.combine(date.today(), aula.horaFim)
        diff_min = int((base_fim - base_inicio).total_seconds() / 60)
        if diff_min > 0:
            duracao_min = diff_min
    hora_fim = (datetime.combine(nova_data, nova_hora_inicio) + timedelta(minutes=duracao_min)).time()

    nova_sessao, _ = models.AulaSessao.objects.get_or_create(
        unidade=aula.unidade,
        tipoServico=aula.tipoServico,
        profissional=profissional,
        data=nova_data,
        horaInicio=nova_hora_inicio,
        horaFim=hora_fim,
        defaults={"capacidade": aula.capacidade},
    )

    capacidade = nova_sessao.capacidade_efetiva()
    total = (
        models.Reserva.objects.filter(aulaSessao=nova_sessao, status="RESERVADA")
        .exclude(pk=reserva.pk)
        .count()
    )
    if capacidade and total >= capacidade:
        return JsonResponse({"error": "Sem capacidade para este horario."}, status=400)

    antiga_sessao = reserva.aulaSessao
    reserva.aulaSessao = nova_sessao
    reserva.status = "RESERVADA"
    reserva.save(update_fields=["aulaSessao", "status"])

    if antiga_sessao and not models.Reserva.objects.filter(aulaSessao=antiga_sessao).exists():
        antiga_sessao.delete()

    inicio_dt = timezone.make_aware(datetime.combine(nova_data, nova_sessao.horaInicio))
    fim_dt = timezone.make_aware(datetime.combine(nova_data, nova_sessao.horaFim))
    return JsonResponse(
          {
              "ok": True,
              "reserva_id": reserva.id,
              "dt_inicio": inicio_dt.isoformat(),
              "dt_fim": fim_dt.isoformat(),
          }
      )


@login_required
@require_POST
def aula_whatsapp_now_api(request, reserva_id):
    reserva = get_object_or_404(
        models.Reserva.objects.select_related(
            "aluno",
            "aulaSessao",
            "aulaSessao__unidade",
            "aulaSessao__profissional",
            "aulaSessao__tipoServico",
        ),
        pk=reserva_id,
    )
    service = WhatsappService()
    telefone = service.get_aluno_phone(reserva.aluno)
    if not telefone:
        return JsonResponse({"error": "Aluno sem telefone valido cadastrado."}, status=400)
    config = models.WhatsappConfiguracao.objects.filter(unidade=reserva.aulaSessao.unidade).first()
    template = (
        config.template_aviso_aluno
        if config and config.template_aviso_aluno
        else "Olá {aluno}, amanhã temos aula de {tipo_servico} às {horario}. Podemos confirmar?"
    )
    mensagem = _render_whatsapp_template(
        template,
        aluno=reserva.aluno.dsNome,
        unidade=reserva.aulaSessao.unidade.dsUnidade if reserva.aulaSessao and reserva.aulaSessao.unidade else "",
        data=reserva.aulaSessao.data.strftime("%d/%m/%Y") if reserva.aulaSessao and reserva.aulaSessao.data else "",
        horario=reserva.aulaSessao.horaInicio.strftime("%H:%M") if reserva.aulaSessao and reserva.aulaSessao.horaInicio else "",
        tipo_servico=reserva.aulaSessao.tipoServico.dsTipoServico if reserva.aulaSessao and reserva.aulaSessao.tipoServico else "Pilates",
        profissional=reserva.aulaSessao.profissional.profissional if reserva.aulaSessao and reserva.aulaSessao.profissional else "",
        aulas=f"{reserva.aulaSessao.horaInicio.strftime('%H:%M')} - {reserva.aulaSessao.tipoServico.dsTipoServico if reserva.aulaSessao and reserva.aulaSessao.tipoServico else 'Pilates'}",
    )
    resp = service.send(reserva.aluno, telefone, mensagem, WhatsappMessageType.MANUAL)
    if resp.get("error"):
        return JsonResponse({"error": "Nao foi possivel enviar a mensagem."}, status=400)
    return JsonResponse({"ok": True, "message": "Mensagem enviada agora."})


@csrf_exempt
@require_POST
def totalpass_webhook(request):
    payload = _parse_json_body(request)
    cfg = _get_totalpass_config(payload)
    token = ""
    if cfg and cfg.webhook_token:
        token = cfg.webhook_token
    if not token:
        token = getattr(settings, "TOTALPASS_WEBHOOK_TOKEN", "") or ""
    if token:
        header_token = request.headers.get("X-Totalpass-Token") or request.headers.get("x-totalpass-token") or ""
        if header_token != token:
            return JsonResponse({"error": "Token invalido."}, status=403)
    if cfg and cfg.ativo is False:
        return JsonResponse({"ok": True, "warning": "Integracao desativada."}, status=202)
    reserva, err = _process_totalpass_payload(payload, cfg)
    if err:
        return JsonResponse({"error": err}, status=400)
    return JsonResponse({"ok": True, "reserva_id": reserva.id if reserva else None})


def create_view(request, model, form_class, redirect_name):
    if model in _admin_only_models() and _is_professor_user(request.user):
        messages.error(request, "Sem permissao para acessar esta area.")
        return redirect("dashboard")
    if request.method == "POST":
        data = request.POST.copy()
        data = _inject_cd_value(model, data)
        if model is models.AulaSessao and not data.get("horaFim"):
            unidade_id = data.get("unidade")
            hora_inicio = data.get("horaInicio")
            if unidade_id and hora_inicio:
                unidade = models.Unidade.objects.filter(pk=unidade_id).first()
                if unidade and unidade.duracao_aula_minutos:
                    inicio_time = datetime.strptime(hora_inicio, "%H:%M").time()
                    base = datetime.combine(date.today(), inicio_time)
                    fim = base + timedelta(minutes=unidade.duracao_aula_minutos)
                    data["horaFim"] = fim.time().strftime("%H:%M")
        if model is models.Profissional and not data.get("cdPerfilAcesso"):
            perfil, _ = models.PerfilAcesso.objects.get_or_create(
                cdPerfilAcesso=1, defaults={"dsPerfilAcesso": "Padrao"}
            )
            data["cdPerfilAcesso"] = perfil.id
        if model is models.Contrato:
            plano_id = data.get("cdPlano")
            plano = models.Plano.objects.filter(pk=plano_id).first() if plano_id else None
            if plano:
                valor_aula_input = data.get("valor_aula")
                recorrencia, valor_aula, valor_parcela, valor_total = _contrato_precificacao(plano, valor_aula_input)
                data["recorrencia"] = recorrencia
                data["valor_aula"] = valor_aula if valor_aula is not None else ""
                data["valor_parcela"] = valor_parcela
                data["valor_total"] = valor_total
        form = form_class(data, files=request.FILES or None)
        if form.is_valid():
            if model is models.Contrato:
                cleaned = form.cleaned_data
                plano = cleaned.get("cdPlano")
                recorrencia, valor_aula, valor_parcela, valor_total = _contrato_precificacao(plano, cleaned.get("valor_aula"))
                is_avulso = bool(getattr(plano, "is_avulso", False))
                contrato_data = {
                    "cdContrato": cleaned["cdContrato"],
                    "cdAluno": cleaned["cdAluno"],
                    "cdPlano": cleaned["cdPlano"],
                    "recorrencia": recorrencia,
                    "valor_aula": valor_aula,
                    "cdUnidade": cleaned["cdUnidade"],
                    "cdProfissional": cleaned["cdProfissional"],
                    "valor_parcela": valor_parcela,
                    "valor_total": valor_total,
                    "dtInicioContrato": cleaned["dtInicioContrato"],
                    "dtFimContrato": cleaned["dtFimContrato"],
                    "modo_pagamento": cleaned.get("modo_pagamento"),
                }
                obj = services.criar_contrato_e_contas(contrato_data, valor_parcela, recorrencia=recorrencia)
                _salvar_documento_contrato(obj)
                if services.enviar_contrato_para_assinatura(obj, request.build_absolute_uri("/")):
                    messages.success(request, "Contrato criado e enviado por email. Agende as aulas.")
                else:
                    messages.warning(request, "Contrato criado, mas aluno sem email para assinatura.")
                    messages.success(request, "Contrato criado. Agende as aulas.")
                _enviar_contrato_whatsapp(request, obj, is_new=True)
                if is_avulso:
                    messages.info(request, "Contrato avulso criado. Sem agenda automatica.")
                    return redirect("alunos_detail", pk=obj.cdAluno_id)
                return redirect("contratos_agenda", pk=obj.id)
            obj = form.save()
            if model is models.ContasPagar:
                recorrencia = obj.recorrencia
                quantidade = obj.recorrencia_quantidade or 0
                if recorrencia and quantidade < 1:
                    quantidade = 1
                    obj.recorrencia_quantidade = quantidade
                    obj.save(update_fields=["recorrencia_quantidade"])
                if recorrencia and quantidade > 1:
                    max_cd = models.ContasPagar.objects.order_by("-cdContasPagar").values_list("cdContasPagar", flat=True).first() or 0
                    for idx in range(1, quantidade):
                        if recorrencia == "SEMANAL":
                            vencimento = obj.dtVencimento + timedelta(days=7 * idx)
                        elif recorrencia == "ANUAL":
                            vencimento = _add_years(obj.dtVencimento, idx)
                        else:
                            vencimento = _add_months(obj.dtVencimento, idx)
                        max_cd += 1
                        models.ContasPagar.objects.create(
                            cdContasPagar=max_cd,
                            cdFornecedor=obj.cdFornecedor,
                            cdCategoria=obj.cdCategoria,
                            cdSubcategoria=obj.cdSubcategoria,
                            dtVencimento=vencimento,
                            valor=obj.valor,
                            recorrencia=obj.recorrencia,
                            recorrencia_quantidade=obj.recorrencia_quantidade,
                        )
            if model is models.Profissional:
                _sync_user_for_profissional(obj, raw_password=form.cleaned_data.get("password"))
            if model is models.Aluno:
                _sync_aluno_address(obj, request.POST)
                _sync_aluno_phones(obj, request.POST)
            messages.success(request, "Salvo com sucesso")
        else:
            messages.error(request, "Verifique os erros do formulario.")
            return render(
                request,
                "generic/form.html",
                {
                    "form": form,
                    "title": "Novo",
                    "model_name": model._meta.model_name,
                    "active_menu": _active_menu(request.path),
                },
            )
        return redirect(redirect_name)
    return render(
        request,
        "generic/form.html",
        {"form": form_class(), "title": "Novo", "model_name": model._meta.model_name, "active_menu": _active_menu(request.path)},
    )


def edit_view(request, model, form_class, redirect_name, pk):
    if model in _admin_only_models() and _is_professor_user(request.user):
        messages.error(request, "Sem permissao para acessar esta area.")
        return redirect("dashboard")
    obj = get_object_or_404(model, pk=pk)
    if request.method == "POST":
        next_url = request.POST.get("next", "").strip()
        if next_url and not next_url.startswith("/"):
            next_url = ""
        data = request.POST.copy()
        for field in model._meta.fields:
            if field.name.startswith("cd") and not data.get(field.name):
                data[field.name] = getattr(obj, field.name)
        if model is models.AulaSessao and not data.get("horaFim"):
            unidade_id = data.get("unidade")
            hora_inicio = data.get("horaInicio")
            if unidade_id and hora_inicio:
                unidade = models.Unidade.objects.filter(pk=unidade_id).first()
                if unidade and unidade.duracao_aula_minutos:
                    inicio_time = datetime.strptime(hora_inicio, "%H:%M").time()
                    base = datetime.combine(date.today(), inicio_time)
                    fim = base + timedelta(minutes=unidade.duracao_aula_minutos)
                    data["horaFim"] = fim.time().strftime("%H:%M")
        if model is models.Contrato:
            plano_id = data.get("cdPlano")
            plano = models.Plano.objects.filter(pk=plano_id).first() if plano_id else None
            if plano:
                valor_aula_input = data.get("valor_aula") or getattr(obj, "valor_aula", None)
                recorrencia, valor_aula, valor_parcela, valor_total = _contrato_precificacao(plano, valor_aula_input)
                data["recorrencia"] = recorrencia
                data["valor_aula"] = valor_aula if valor_aula is not None else ""
                data["valor_parcela"] = valor_parcela
                data["valor_total"] = valor_total
        form = form_class(data, files=request.FILES or None, instance=obj)
        if form.is_valid():
            obj = form.save()
            if model is models.Contrato:
                recorrencia, valor_aula, valor_parcela, valor_total = _contrato_precificacao(obj.cdPlano, obj.valor_aula)
                updates = []
                if obj.recorrencia != recorrencia:
                    obj.recorrencia = recorrencia
                    updates.append("recorrencia")
                if obj.valor_aula != valor_aula:
                    obj.valor_aula = valor_aula
                    updates.append("valor_aula")
                if obj.valor_parcela != valor_parcela:
                    obj.valor_parcela = valor_parcela
                    updates.append("valor_parcela")
                if obj.valor_total != valor_total:
                    obj.valor_total = valor_total
                    updates.append("valor_total")
                if updates:
                    obj.save(update_fields=updates)
            if model is models.Profissional:
                _sync_user_for_profissional(obj, raw_password=form.cleaned_data.get("password"))
            if model is models.Aluno:
                _sync_aluno_address(obj, request.POST)
                _sync_aluno_phones(obj, request.POST)
            if model is models.Contrato and obj.status == "NAO_ASSINADO":
                if not services.enviar_contrato_para_assinatura(obj, request.build_absolute_uri("/")):
                    messages.warning(request, "Contrato atualizado, mas aluno sem email para assinatura.")
            messages.success(request, "Atualizado com sucesso")
            return redirect(next_url or redirect_name)
        messages.error(request, "Verifique os erros do formulario.")
        return render(
            request,
            "generic/form.html",
            {
                "form": form,
                "title": "Editar",
                "model_name": model._meta.model_name,
                "active_menu": _active_menu(request.path),
                "next_url": next_url or None,
            },
        )
    return render(
        request,
        "generic/form.html",
        {
            "form": form_class(instance=obj),
            "title": "Editar",
            "model_name": model._meta.model_name,
            "active_menu": _active_menu(request.path),
            "next_url": request.GET.get("next", ""),
        },
    )


def delete_view(request, model, redirect_name, pk):
    if model in _admin_only_models() and _is_professor_user(request.user):
        messages.error(request, "Sem permissao para acessar esta area.")
        return redirect("dashboard")
    obj = get_object_or_404(model, pk=pk)
    delete_context = {}
    if model is models.Contrato:
        reservas_qs = models.Reserva.objects.filter(
            aluno=obj.cdAluno,
            aulaSessao__data__range=(obj.dtInicioContrato, obj.dtFimContrato),
            aulaSessao__unidade=obj.cdUnidade,
            aulaSessao__tipoServico=obj.cdPlano.cdTipoServico,
        )
        aula_ids = list(reservas_qs.values_list("aulaSessao_id", flat=True))
        delete_context = {
            "requires_confirm": True,
            "delete_summary": {
                "contas_receber": models.ContasReceber.objects.filter(contrato=obj).count(),
                "reservas": reservas_qs.count(),
                "aulas": len(set(aula_ids)),
            },
        }
    if request.method == "POST":
        next_url = request.POST.get("next", "").strip()
        if next_url and not next_url.startswith("/"):
            next_url = ""
        if model is models.Contrato:
            if request.POST.get("confirm_delete") != "1":
                messages.error(request, "Confirme a exclusao do contrato.")
                return redirect(next_url or redirect_name)
            # Remove contas a receber e reservas do aluno dentro do periodo do contrato.
            models.ContasReceber.objects.filter(contrato=obj).delete()
            reservas_qs = models.Reserva.objects.filter(
                aluno=obj.cdAluno,
                aulaSessao__data__range=(obj.dtInicioContrato, obj.dtFimContrato),
                aulaSessao__unidade=obj.cdUnidade,
                aulaSessao__tipoServico=obj.cdPlano.cdTipoServico,
            )
            aula_ids = list(reservas_qs.values_list("aulaSessao_id", flat=True))
            reservas_qs.delete()
            if aula_ids:
                models.AulaSessao.objects.filter(id__in=aula_ids).annotate(total=Count("reserva")).filter(total=0).delete()
        obj.delete()
        messages.success(request, "Removido")
        return redirect(next_url or redirect_name)
    context = {"object": obj, "title": "Excluir", "active_menu": _active_menu(request.path)}
    context.update(delete_context)
    return render(request, "generic/confirm_delete.html", context)


@login_required
def aluno_reservas_delete_bulk(request, aluno_id):
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    if request.method != "POST":
        return redirect(f"{reverse('alunos_detail', args=[aluno.id])}?tab=agenda")
    next_url = request.POST.get("next", "").strip()
    if next_url and not next_url.startswith("/"):
        next_url = ""
    reserva_ids = [value for value in request.POST.getlist("reserva_ids") if value]
    if not reserva_ids:
        messages.error(request, "Selecione pelo menos uma aula.")
        return redirect(next_url or f"{reverse('alunos_detail', args=[aluno.id])}?tab=agenda")
    reservas_qs = models.Reserva.objects.filter(id__in=reserva_ids, aluno=aluno).select_related("aulaSessao")
    aula_ids = list(reservas_qs.values_list("aulaSessao_id", flat=True))
    if not aula_ids:
        messages.error(request, "Nenhuma aula valida foi selecionada.")
        return redirect(next_url or f"{reverse('alunos_detail', args=[aluno.id])}?tab=agenda")
    reservas_qs.delete()
    if aula_ids:
        models.AulaSessao.objects.filter(id__in=aula_ids).annotate(total=Count("reserva")).filter(total=0).delete()
    messages.success(request, "Aulas excluidas com sucesso.")
    return redirect(next_url or f"{reverse('alunos_detail', args=[aluno.id])}?tab=agenda")


@login_required
def baixar_conta_receber(request, pk):
    conta = get_object_or_404(
        models.ContasReceber.objects.select_related(
            "contrato",
            "contrato__cdAluno",
            "reserva",
            "reserva__aluno",
            "reserva__pacote_avulso",
        ),
        pk=pk,
    )
    if request.method not in {"GET", "POST"}:
        return redirect("contas_receber_list")
    next_url = request.POST.get("next", "").strip() if request.method == "POST" else request.GET.get("next", "").strip()
    if next_url and not next_url.startswith("/"):
        next_url = ""
    data_pagamento = (
        request.POST.get("dtPagamento")
        if request.method == "POST"
        else request.GET.get("dtPagamento", "")
    ) or ""
    try:
        pago_em = datetime.strptime(data_pagamento, "%Y-%m-%d").date() if data_pagamento else timezone.now().date()
    except ValueError:
        pago_em = timezone.now().date()
    conta.status = "PAGO"
    conta.dtPagamento = pago_em
    conta.save(update_fields=["status", "dtPagamento"])
    messages.success(request, "Lancamento baixado com sucesso.")
    aluno = _conta_receber_aluno(conta)
    fallback = f"{reverse('alunos_detail', args=[aluno.pk])}?tab=financeiro" if aluno else "contas_receber_list"
    return redirect(next_url or fallback)


@login_required
def excluir_conta_receber(request, pk):
    conta = get_object_or_404(
        models.ContasReceber.objects.select_related(
            "contrato",
            "contrato__cdAluno",
            "reserva",
            "reserva__aluno",
            "reserva__pacote_avulso",
        ),
        pk=pk,
    )
    if request.method not in {"GET", "POST"}:
        return redirect("contas_receber_list")
    next_url = request.POST.get("next", "").strip() if request.method == "POST" else request.GET.get("next", "").strip()
    if next_url and not next_url.startswith("/"):
        next_url = ""
    aluno = _conta_receber_aluno(conta)
    conta.delete()
    messages.success(request, "Lancamento excluido com sucesso.")
    fallback = f"{reverse('alunos_detail', args=[aluno.pk])}?tab=financeiro" if aluno else "contas_receber_list"
    return redirect(next_url or fallback)


@login_required
@require_POST
def cobrar_conta_receber_whatsapp(request, pk):
    conta = get_object_or_404(
        models.ContasReceber.objects.select_related(
            "contrato", "contrato__cdAluno", "contrato__cdUnidade",
            "reserva", "reserva__aluno", "aluno",
        ),
        pk=pk,
    )
    next_url = request.POST.get("next", "").strip()
    if next_url and not next_url.startswith("/"):
        next_url = ""
    aluno = _conta_receber_aluno(conta)
    fallback = f"{reverse('alunos_detail', args=[aluno.pk])}?tab=financeiro" if aluno else "contas_receber_list"
    destino = next_url or fallback

    if conta.status not in ("ABERTO", "ATRASADO"):
        messages.warning(request, "Esta fatura nao esta em aberto. Cobranca nao enviada.")
        return redirect(destino)
    if not aluno:
        messages.error(request, "Nao foi possivel identificar o aluno desta fatura.")
        return redirect(destino)

    service = WhatsappService()
    telefone = service.get_aluno_phone(aluno)
    if not telefone:
        messages.error(request, "Aluno sem telefone valido cadastrado.")
        return redirect(destino)

    valor = f"{float(conta.valor):.2f}".replace(".", ",") if conta.valor is not None else "0,00"
    vencimento = conta.dtVencimento.strftime("%d/%m/%Y") if conta.dtVencimento else ""
    mensagem = (
        f"Ola {aluno.dsNome}! Identificamos uma fatura em aberto no valor de R$ {valor}"
        + (f", com vencimento em {vencimento}" if vencimento else "")
        + ". Caso ja tenha efetuado o pagamento, por favor desconsidere esta mensagem. "
        "Qualquer duvida, estamos a disposicao!"
    )
    contrato = conta.contrato if conta.contrato_id else None
    resp = service.send(aluno, telefone, mensagem, models.WhatsappMessageType.MANUAL, contrato=contrato)
    if resp.get("error"):
        messages.error(request, "Nao foi possivel enviar a cobranca por WhatsApp.")
    else:
        messages.success(request, f"Cobranca enviada por WhatsApp para {aluno.dsNome}.")
    return redirect(destino)


@login_required
def efetuar_pagamento_conta_pagar(request, pk):
    conta = get_object_or_404(models.ContasPagar, pk=pk)
    if request.method != "POST":
        return redirect("contas_pagar_list")
    next_url = request.POST.get("next", "").strip()
    if next_url and not next_url.startswith("/"):
        next_url = ""
    data_pagamento = request.POST.get("dtPagamento") or ""
    try:
        pago_em = datetime.strptime(data_pagamento, "%Y-%m-%d").date() if data_pagamento else timezone.now().date()
    except ValueError:
        pago_em = timezone.now().date()
    comprovante = request.FILES.get("comprovante")
    conta.status = "PAGO"
    conta.dtPagamento = pago_em
    if comprovante:
        conta.comprovante = comprovante
    conta.save(update_fields=["status", "dtPagamento", "comprovante"])
    messages.success(request, "Pagamento registrado.")
    return redirect(next_url or "contas_pagar_list")


@login_required
def cancelar_conta_pagar(request, pk):
    conta = get_object_or_404(models.ContasPagar, pk=pk)
    if request.method != "POST":
        return redirect("contas_pagar_list")
    next_url = request.POST.get("next", "").strip()
    if next_url and not next_url.startswith("/"):
        next_url = ""
    motivo = request.POST.get("motivo_cancelamento", "").strip()
    conta.status = "CANCELADO"
    conta.motivo_cancelamento = motivo
    conta.save(update_fields=["status", "motivo_cancelamento"])
    messages.success(request, "Lancamento cancelado.")
    return redirect(next_url or "contas_pagar_list")


@login_required
def evoluir_reserva(request, pk):
    reserva = get_object_or_404(models.Reserva, pk=pk)
    if request.method != "POST":
        return redirect("aulas_list")
    next_url = request.POST.get("next", "").strip()
    if next_url and not next_url.startswith("/"):
        next_url = ""
    status = request.POST.get("status", "").strip()
    texto = request.POST.get("texto", "").strip()
    if status not in {"CONCLUIDA", "FALTOU_AVISOU", "FALTOU_SEM_AVISAR"}:
        messages.error(request, "Status invalido.")
        return redirect(next_url or "aulas_list")
    if not texto:
        messages.error(request, "Preencha a evolucao.")
        return redirect(next_url or "aulas_list")
    if not reserva.aulaSessao.profissional_id:
        messages.error(request, "Defina o profissional da aula antes de registrar a evolucao.")
        return redirect(next_url or "aulas_list")
    models.EvolucaoAluno.objects.create(
        reserva=reserva,
        profissional=reserva.aulaSessao.profissional,
        texto=texto,
    )
    reserva.status = status
    reserva.save(update_fields=["status"])
    messages.success(request, "Evolucao registrada.")
    return redirect(next_url or "aulas_list")


def _get_filtros_financeiro(request):
    return {
        "status": request.GET.get("status", "").strip(),
        "inicio": request.GET.get("inicio", "").strip(),
        "fim": request.GET.get("fim", "").strip(),
    }


def _filtrar_contas_receber(qs, request):
    filtros = _get_filtros_financeiro(request)
    if filtros["status"]:
        qs = qs.filter(status=filtros["status"])
    inicio = filtros["inicio"]
    fim = filtros["fim"]
    if inicio:
        try:
            inicio_dt = datetime.strptime(inicio, "%Y-%m-%d").date()
            qs = qs.filter(dtVencimento__gte=inicio_dt)
        except ValueError:
            pass
    if fim:
        try:
            fim_dt = datetime.strptime(fim, "%Y-%m-%d").date()
            qs = qs.filter(dtVencimento__lte=fim_dt)
        except ValueError:
            pass
    return qs.order_by("dtVencimento", "id")


def _conta_receber_aluno(conta):
    if getattr(conta, "contrato_id", None) and getattr(conta.contrato, "cdAluno_id", None):
        return conta.contrato.cdAluno
    if getattr(conta, "reserva_id", None) and getattr(conta.reserva, "aluno_id", None):
        return conta.reserva.aluno
    if getattr(conta, "aluno_id", None):
        return conta.aluno
    return None


def _conta_receber_origem(conta):
    if getattr(conta, "contrato_id", None):
        return f"Contrato #{conta.contrato.cdContrato}"
    if getattr(conta, "reserva_id", None):
        pacote = getattr(conta.reserva, "pacote_avulso", None)
        if pacote:
            return f"Aula avulsa #{pacote.cdAulaAvulsa}"
        return "Aula avulsa"
    if getattr(conta, "aluno_id", None):
        return conta.descricao or "Lancamento avulso"
    return "-"


@login_required
@require_POST
def lancar_conta_receber_avulsa(request, aluno_id):
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    destino = f"{reverse('alunos_detail', args=[aluno.pk])}?tab=financeiro"
    form = forms.LancamentoAvulsoForm(request.POST)
    if form.is_valid():
        conta = form.save(commit=False)
        conta.aluno = aluno
        conta.contrato = None
        conta.reserva = None
        conta.save()
        messages.success(request, "Lancamento avulso criado com sucesso.")
    else:
        erros = "; ".join(f"{campo}: {', '.join(msgs)}" for campo, msgs in form.errors.items())
        messages.error(request, f"Nao foi possivel lancar o valor. {erros}")
    return redirect(destino)


@login_required
def exportar_contas_receber_excel(request, aluno_id):
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    qs = _filtrar_contas_receber(
        models.ContasReceber.objects.filter(Q(contrato__cdAluno=aluno) | Q(reserva__aluno=aluno) | Q(aluno=aluno)).select_related(
            "contrato",
            "contrato__cdPlano",
            "contrato__cdPlano__subcategoria_receita",
            "reserva",
            "reserva__pacote_avulso",
            "reserva__pacote_avulso__plano",
        ),
        request,
    )
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Faturas"
    ws.append(["Competencia", "Vencimento", "Pagamento", "Origem", "Subcategoria", "Status", "Valor"])
    for f in qs:
        subcategoria = ""
        if f.contrato_id and f.contrato.cdPlano and f.contrato.cdPlano.subcategoria_receita:
            subcategoria = f.contrato.cdPlano.subcategoria_receita.dsSubcategoria
        elif f.reserva_id and f.reserva.pacote_avulso and f.reserva.pacote_avulso.plano and f.reserva.pacote_avulso.plano.subcategoria_receita:
            subcategoria = f.reserva.pacote_avulso.plano.subcategoria_receita.dsSubcategoria
        ws.append(
            [
                f.competencia or "",
                f.dtVencimento.strftime("%d/%m/%Y") if f.dtVencimento else "",
                f.dtPagamento.strftime("%d/%m/%Y") if f.dtPagamento else "",
                _conta_receber_origem(f),
                subcategoria,
                f.status,
                float(f.valor),
            ]
        )
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="faturas-aluno-{aluno.id}.xlsx"'
    return response


@login_required
def exportar_contas_receber_pdf(request, aluno_id):
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    qs = _filtrar_contas_receber(
        models.ContasReceber.objects.filter(Q(contrato__cdAluno=aluno) | Q(reserva__aluno=aluno) | Q(aluno=aluno)).select_related(
            "contrato",
            "contrato__cdPlano",
            "contrato__cdPlano__subcategoria_receita",
            "reserva",
            "reserva__pacote_avulso",
            "reserva__pacote_avulso__plano",
        ),
        request,
    )
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Faturas do Aluno")
    styles = getSampleStyleSheet()
    title = Paragraph(f"Faturas do aluno: {aluno.dsNome}", styles["Title"])
    subtitle = Paragraph("Resumo financeiro", styles["Normal"])
    data = [["Competencia", "Vencimento", "Pagamento", "Origem", "Subcategoria", "Status", "Valor"]]
    for f in qs:
        subcategoria = ""
        if f.contrato_id and f.contrato.cdPlano and f.contrato.cdPlano.subcategoria_receita:
            subcategoria = f.contrato.cdPlano.subcategoria_receita.dsSubcategoria
        elif f.reserva_id and f.reserva.pacote_avulso and f.reserva.pacote_avulso.plano and f.reserva.pacote_avulso.plano.subcategoria_receita:
            subcategoria = f.reserva.pacote_avulso.plano.subcategoria_receita.dsSubcategoria
        data.append(
            [
                f.competencia or "-",
                f.dtVencimento.strftime("%d/%m/%Y") if f.dtVencimento else "-",
                f.dtPagamento.strftime("%d/%m/%Y") if f.dtPagamento else "-",
                _conta_receber_origem(f),
                subcategoria or "-",
                f.status,
                f"R$ {f.valor}",
            ]
        )
    table = Table(data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4d4d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f7fb")]),
                ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
            ]
        )
    )
    elements = [title, subtitle, Spacer(1, 12), table]
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="faturas-aluno-{aluno.id}.pdf"'
    return response


@login_required
def recibo_conta_receber_pdf(request, pk):
    conta = get_object_or_404(
        models.ContasReceber.objects.select_related(
            "contrato",
            "contrato__cdAluno",
            "reserva",
            "reserva__aluno",
            "reserva__pacote_avulso",
        ),
        pk=pk,
    )
    aluno = _conta_receber_aluno(conta)
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Recibo")
    styles = getSampleStyleSheet()
    title = Paragraph("Recibo de Pagamento", styles["Title"])
    aluno_nome = aluno.dsNome if aluno else "-"
    pago_em = conta.dtPagamento.strftime("%d/%m/%Y") if conta.dtPagamento else "-"
    data = [
        ["Aluno", aluno_nome],
        ["Origem", _conta_receber_origem(conta)],
        ["Competencia", conta.competencia or "-"],
        ["Vencimento", conta.dtVencimento.strftime("%d/%m/%Y") if conta.dtVencimento else "-"],
        ["Pagamento", pago_em],
        ["Valor", f"R$ {conta.valor}"],
        ["Status", conta.status],
    ]
    table = Table(data, colWidths=[120, 360])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4d4d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f7fb")]),
            ]
        )
    )
    elements = [title, Spacer(1, 8), table, Spacer(1, 18), Paragraph("Obrigado!", styles["Heading3"])]
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="recibo-{conta.id}.pdf"'
    return response


@login_required
def exportar_evolucoes_excel(request, aluno_id):
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    qs = (
        models.EvolucaoAluno.objects.filter(reserva__aluno=aluno)
        .select_related("profissional", "reserva")
        .order_by("-dtEvolucao")
    )
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Evolucoes"
    ws.append(["Data", "Profissional", "Evolucao"])
    for e in qs:
        ws.append(
            [
                e.dtEvolucao.strftime("%d/%m/%Y %H:%M") if e.dtEvolucao else "",
                str(e.profissional),
                e.texto,
            ]
        )
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="evolucoes-aluno-{aluno.id}.xlsx"'
    return response


@login_required
def exportar_evolucoes_pdf(request, aluno_id):
    aluno = get_object_or_404(models.Aluno, pk=aluno_id)
    qs = (
        models.EvolucaoAluno.objects.filter(reserva__aluno=aluno)
        .select_related("profissional", "reserva")
        .order_by("-dtEvolucao")
    )
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, title="Evolucoes do Aluno")
    styles = getSampleStyleSheet()
    title = Paragraph(f"Evolucoes do aluno: {aluno.dsNome}", styles["Title"])
    data = [["Data", "Profissional", "Evolucao"]]
    for e in qs:
        data.append(
            [
                e.dtEvolucao.strftime("%d/%m/%Y %H:%M") if e.dtEvolucao else "-",
                str(e.profissional),
                e.texto,
            ]
        )
    table = Table(data, repeatRows=1, colWidths=[110, 140, 290])
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1d4d4d")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#dbe4ea")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f7fb")]),
            ]
        )
    )
    elements = [title, Spacer(1, 12), table]
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="evolucoes-aluno-{aluno.id}.pdf"'
    return response


@login_required
def wizard_step1(request):
    if request.method == "POST":
        if request.FILES.get("documento"):
            data = extract_student_from_document(request.FILES["documento"].read(), request.FILES["documento"].name)
            request.session["wizard_student"] = data
            return render(
                request,
                "wizard/step1_documento.html",
                {
                    "student": data,
                    "unidades": models.Unidade.objects.all(),
                    "breadcrumbs": [("Home", reverse("dashboard")), ("Wizard", "#")],
                    "active_menu": "cadastros",
                },
            )
        if request.POST.get("confirm") == "1":
            cpf = request.POST.get("cpf", "")
            if models.Aluno.objects.filter(dsCPF=cpf).exists():
                messages.error(request, "CPF duplicado")
                return redirect("wizard_step1")
            max_cd = models.Aluno.objects.order_by("-cdAluno").values_list("cdAluno", flat=True).first() or 0
            aluno = models.Aluno.objects.create(
                cdAluno=max_cd + 1,
                dsNome=request.POST.get("nome", ""),
                dsCPF=cpf,
                dsRg=request.POST.get("rg", ""),
                cdUnidade_id=int(request.POST.get("cdUnidade")),
            )
            request.session["wizard_aluno_id"] = aluno.id
            return redirect("wizard_step2")
    return render(
        request,
        "wizard/step1_documento.html",
        {"unidades": models.Unidade.objects.all(), "breadcrumbs": [("Home", reverse("dashboard")), ("Wizard", "#")], "active_menu": "cadastros"},
    )


@login_required
def wizard_step2(request):
    if request.method == "POST":
        if request.FILES.get("comprovante"):
            data = extract_address_from_proof(request.FILES["comprovante"].read(), request.FILES["comprovante"].name)
            request.session["wizard_address"] = data
            return render(
                request,
                "wizard/step2_endereco.html",
                {"address": data, "breadcrumbs": [("Home", reverse("dashboard")), ("Wizard", "#")], "active_menu": "cadastros"},
            )
        if request.POST.get("confirm") == "1":
            aluno_id = request.session.get("wizard_aluno_id")
            if not aluno_id:
                messages.error(request, "Aluno nao encontrado")
                return redirect("wizard_step1")
            max_cd = models.EnderecoAluno.objects.order_by("-cdEndereco").values_list("cdEndereco", flat=True).first() or 0
            endereco = models.EnderecoAluno.objects.create(
                cdEndereco=max_cd + 1,
                cdAluno_id=aluno_id,
                dsLogradouro=request.POST.get("logradouro", ""),
                dsNumero=request.POST.get("numero", ""),
                dsCEP=request.POST.get("cep", ""),
                dsCidade=request.POST.get("cidade", ""),
                dsBairro=request.POST.get("bairro", ""),
            )
            models.Aluno.objects.filter(pk=aluno_id).update(cdEndereco=endereco)
            return redirect("wizard_step3")
    return render(
        request,
        "wizard/step2_endereco.html",
        {"breadcrumbs": [("Home", reverse("dashboard")), ("Wizard", "#")], "active_menu": "cadastros"},
    )


@login_required
def wizard_step3(request):
    termos = models.TermoUso.objects.all()
    if request.method == "POST":
        termo_id = request.POST.get("termo")
        aluno_id = request.session.get("wizard_aluno_id")
        if aluno_id and termo_id:
            services.registrar_aceite_termo(models.Aluno.objects.get(pk=aluno_id), models.TermoUso.objects.get(pk=int(termo_id)))
        return redirect("wizard_step4")
    return render(
        request,
        "wizard/step3_termo.html",
        {"termos": termos, "breadcrumbs": [("Home", reverse("dashboard")), ("Wizard", "#")], "active_menu": "cadastros"},
    )


@login_required
def wizard_step4(request):
    if request.method == "POST":
        request.session["wizard_contrato"] = request.POST.dict()
        return redirect("wizard_step5")
    context = {
        "alunos": models.Aluno.objects.all(),
        "aluno_id": request.session.get("wizard_aluno_id"),
        "planos": models.Plano.objects.all(),
        "unidades": models.Unidade.objects.all(),
        "profissionais": models.Profissional.objects.all(),
        "breadcrumbs": [("Home", reverse("dashboard")), ("Wizard", "#")],
        "active_menu": "cadastros",
    }
    return render(request, "wizard/step4_contrato.html", context)


@login_required
def wizard_step5(request):
    if request.method == "POST":
        data = request.session.get("wizard_contrato", {})
        aluno = models.Aluno.objects.get(pk=int(data.get("cdAluno")))
        plano = models.Plano.objects.get(pk=int(data.get("cdPlano")))
        recorrencia, valor_aula, valor_parcela, valor_total = _contrato_precificacao(plano, data.get("valor_aula"))
        contrato_data = {
            "cdContrato": int(data.get("cdContrato")),
            "cdAluno": aluno,
            "cdPlano": plano,
            "recorrencia": recorrencia,
            "valor_aula": valor_aula,
            "cdUnidade": models.Unidade.objects.get(pk=int(data.get("cdUnidade"))),
            "cdProfissional": models.Profissional.objects.get(pk=int(data.get("cdProfissional"))),
            "valor_parcela": valor_parcela,
            "valor_total": valor_total,
            "dtInicioContrato": data.get("dtInicioContrato"),
            "dtFimContrato": data.get("dtFimContrato"),
        }
        contrato = services.criar_contrato_e_contas(contrato_data, valor_parcela, recorrencia=recorrencia)
        if services.enviar_contrato_para_assinatura(contrato, request.build_absolute_uri("/")):
            messages.success(request, "Contrato criado e enviado por email. Agende as aulas.")
        else:
            messages.warning(request, "Contrato criado, mas aluno sem email para assinatura.")
            messages.success(request, "Contrato criado. Agende as aulas.")
        _enviar_contrato_whatsapp(request, contrato, is_new=True)
        return redirect("contratos_agenda", pk=contrato.id)
    return render(
        request,
        "wizard/step5_reservas.html",
        {"breadcrumbs": [("Home", reverse("dashboard")), ("Wizard", "#")], "active_menu": "cadastros"},
    )


@login_required
def contratos_gestao(request):
    if _is_professor_user(request.user):
        messages.error(request, "Sem permissao para acessar esta area.")
        return redirect("dashboard")

    today = timezone.now().date()
    query = request.GET.get("q", "").strip()
    filtro = request.GET.get("filtro", "ativos")

    qs = models.Contrato.objects.select_related(
        "cdAluno", "cdPlano", "cdUnidade", "cdProfissional"
    )

    if query:
        qs = qs.filter(
            Q(cdAluno__dsNome__icontains=query)
            | Q(cdContrato__icontains=query)
            | Q(cdPlano__dsPlano__icontains=query)
        )

    # Filtros por situacao do vencimento
    if filtro == "vencidos":
        qs = qs.filter(dtFimContrato__lt=today)
    elif filtro == "vencendo":
        qs = qs.filter(
            dtFimContrato__gte=today,
            dtFimContrato__lte=today + timedelta(days=30),
        )
    elif filtro == "todos":
        pass
    else:  # ativos (vigentes hoje)
        filtro = "ativos"
        qs = qs.filter(dtInicioContrato__lte=today, dtFimContrato__gte=today)

    qs = qs.order_by("dtFimContrato")

    contratos = []
    for contrato in qs:
        dias = (contrato.dtFimContrato - today).days
        if dias < 0:
            situacao, badge = "Vencido", "danger"
        elif dias <= 7:
            situacao, badge = "Vence esta semana", "danger"
        elif dias <= 30:
            situacao, badge = "Vence em breve", "warning"
        else:
            situacao, badge = "Vigente", "success"
        contratos.append(
            {
                "obj": contrato,
                "dias_restantes": dias,
                "situacao": situacao,
                "badge": badge,
            }
        )

    # Resumo dos contadores (independente do filtro de listagem)
    base = models.Contrato.objects.all()
    resumo = {
        "ativos": base.filter(dtInicioContrato__lte=today, dtFimContrato__gte=today).count(),
        "vencendo": base.filter(
            dtFimContrato__gte=today, dtFimContrato__lte=today + timedelta(days=30)
        ).count(),
        "vencidos": base.filter(dtFimContrato__lt=today).count(),
        "total": base.count(),
    }

    context = {
        "contratos": contratos,
        "resumo": resumo,
        "query": query,
        "filtro": filtro,
        "today": today,
        "breadcrumbs": [
            ("Home", reverse("dashboard")),
            ("Contratos", reverse("contratos_list")),
            ("Gestao", "#"),
        ],
        "active_menu": "contratos",
    }
    return render(request, "contratos/gestao.html", context)


@login_required
def contrato_agenda(request, pk):
    contrato = get_object_or_404(models.Contrato, pk=pk)
    plano = contrato.cdPlano
    agenda_semanas = 4 if contrato.recorrencia == "SEMANAL" else None
    aulas_por_semana = 1 if contrato.recorrencia == "SEMANAL" else (plano.aulas_por_semana or 1)
    agenda_fim = contrato.dtFimContrato
    if agenda_semanas:
        agenda_fim = min(agenda_fim, contrato.dtInicioContrato + timedelta(days=27))
    aulas = models.AulaSessao.objects.filter(
        unidade=contrato.cdUnidade,
        data__range=(contrato.dtInicioContrato, agenda_fim),
    ).select_related("unidade", "tipoServico").order_by("data", "horaInicio", "tipoServico_id")

    profissionais = list(models.Profissional.objects.all())
    prof_ids = [prof.id for prof in profissionais]
    duracao = contrato.cdUnidade.duracao_aula_minutos or 50
    funcionamento, horarios_configurados = _load_horarios_ativos(contrato.cdUnidade_id)

    slots = {}
    aulas_by_key = {(aula.data, aula.horaInicio, aula.horaFim, aula.profissional_id, aula.tipoServico_id): aula for aula in aulas}
    capacidade_padrao = contrato.cdUnidade.capacidade or 0

    dates_by_weekday = {i: [] for i in range(7)}
    current = contrato.dtInicioContrato
    while current <= agenda_fim:
        dates_by_weekday[current.weekday()].append(current)
        current += timedelta(days=1)

    blocks_by_prof = {}
    for prof in profissionais:
        blocks_by_prof[prof.id] = _load_bloqueios(
            contrato.cdUnidade_id,
            None,
            prof.id,
            contrato.dtInicioContrato,
            agenda_fim,
        )

    for weekday, windows in funcionamento.items():
        if not dates_by_weekday.get(weekday):
            continue
        sample_date = dates_by_weekday[weekday][0]
        time_slots = _generate_slots_for_date(sample_date, windows, duracao)
        for inicio, fim in time_slots:
            allowed_profs = []
            for prof in profissionais:
                prof_ok = True
                blocks = blocks_by_prof.get(prof.id, [])
                for day in dates_by_weekday[weekday]:
                    if _is_slot_blocked(blocks, day, inicio, fim):
                        prof_ok = False
                        break
                    aula = aulas_by_key.get((day, inicio, fim, prof.id, plano.cdTipoServico_id))
                    if aula:
                        reservadas = models.Reserva.objects.filter(aulaSessao=aula, status="RESERVADA").count()
                        cap = aula.capacidade_efetiva()
                    else:
                        reservadas = 0
                        cap = capacidade_padrao
                    if reservadas >= (cap or 0):
                        prof_ok = False
                        break
                if prof_ok:
                    allowed_profs.append(prof.id)
            if allowed_profs:
                slots[(weekday, inicio, fim)] = {
                    "weekday": weekday,
                    "inicio": inicio,
                    "fim": fim,
                    "allowed_profs": allowed_profs,
                }

    if not slots and not horarios_configurados:
        slots = {}
        for aula in aulas:
            reservadas = models.Reserva.objects.filter(aulaSessao=aula, status="RESERVADA").count()
            if reservadas >= aula.capacidade_efetiva():
                continue
            key = (aula.data.weekday(), aula.horaInicio, aula.horaFim)
            slots[key] = {"weekday": key[0], "inicio": key[1], "fim": key[2], "allowed_profs": []}

    if request.method == "POST":

        escolhidas = []
        for idx in range(1, aulas_por_semana + 1):
            valor = request.POST.get(f"slot_{idx}") or ""
            dia_raw = request.POST.get(f"slot_day_{idx}") or ""
            if valor:
                escolhidas.append((idx, valor, dia_raw))
        if len(escolhidas) != aulas_por_semana:
            messages.error(request, f"Selecione exatamente {aulas_por_semana} horarios por semana.")
        else:
            conflitos = []
            faltando_prof = False
            usados = set()
            dias_usados = set()
            for idx, slot_value, dia_raw in escolhidas:
                if slot_value in usados:
                    messages.error(request, "Nao repita o mesmo horario.")
                    return redirect("contratos_agenda", pk=contrato.id)
                usados.add(slot_value)
                try:
                    weekday_raw, inicio, fim = slot_value.split("|")
                    weekday = int(weekday_raw)
                except ValueError:
                    messages.error(request, "Horario invalido na selecao.")
                    return redirect("contratos_agenda", pk=contrato.id)
                if dia_raw:
                    try:
                        dia_selected = int(dia_raw)
                    except ValueError:
                        messages.error(request, "Dia da semana invalido.")
                        return redirect("contratos_agenda", pk=contrato.id)
                    if dia_selected != weekday:
                        messages.error(request, "Dia e horario nao conferem.")
                        return redirect("contratos_agenda", pk=contrato.id)
                if weekday in dias_usados:
                    messages.error(request, "Selecione apenas um horario por dia.")
                    return redirect("contratos_agenda", pk=contrato.id)
                dias_usados.add(weekday)
                prof_id = request.POST.get(f"prof_for_{idx}") or ""
                try:
                    prof_id = int(prof_id)
                except ValueError:
                    prof_id = None
                if not prof_id:
                    faltando_prof = True
                    continue
                try:
                    inicio_time = datetime.strptime(inicio, "%H:%M").time()
                    fim_time = datetime.strptime(fim, "%H:%M").time()
                except ValueError:
                    messages.error(request, "Horario invalido na selecao.")
                    return redirect("contratos_agenda", pk=contrato.id)
                horario_key = (weekday, inicio_time, fim_time)
                if horario_key not in slots:
                    messages.error(request, "Horario invalido para o dia selecionado.")
                    return redirect("contratos_agenda", pk=contrato.id)
                slot_payload = slots.get(horario_key, {})
                allowed = set(slot_payload.get("allowed_profs") or [])
                if allowed and prof_id not in allowed:
                    messages.error(request, "Professor invalido para o horario selecionado.")
                    return redirect("contratos_agenda", pk=contrato.id)
                current = contrato.dtInicioContrato
                while current <= agenda_fim:
                    if current.weekday() == weekday:
                        blocks = blocks_by_prof.get(prof_id, [])
                        if _is_slot_blocked(blocks, current, inicio_time, fim_time):
                            conflitos.append(f"Bloqueio em {current} {inicio}")
                            current = current + timedelta(days=1)
                            continue
                        # A reserva do contrato deve ser somente para o servico do
                        # plano (ex: Pilates). Antes o sistema criava uma reserva
                        # para cada servico oferecido no horario, gerando aulas
                        # duplicadas de outro servico (ex: Alongamento).
                        servico_id = plano.cdTipoServico_id
                        if servico_id:
                            try:
                                aula = models.AulaSessao.objects.filter(
                                    unidade=contrato.cdUnidade,
                                    tipoServico_id=servico_id,
                                    profissional_id=prof_id,
                                    data=current,
                                    horaInicio=inicio_time,
                                    horaFim=fim_time,
                                ).first()
                                if not aula:
                                    aula = models.AulaSessao.objects.create(
                                        unidade=contrato.cdUnidade,
                                        tipoServico_id=servico_id,
                                        profissional_id=prof_id,
                                        data=current,
                                        horaInicio=inicio_time,
                                        horaFim=fim_time,
                                    )
                                else:
                                    aula.profissional_id = prof_id
                                    aula.save(update_fields=["profissional"])
                                if not models.Reserva.objects.filter(aluno=contrato.cdAluno, aulaSessao=aula).exists():
                                    services.create_reserva(contrato.cdAluno, aula, status="RESERVADA")
                            except Exception:
                                conflitos.append(f"Sem vaga em {current} {inicio}")
                    current = current + timedelta(days=1)
            if faltando_prof:
                messages.error(request, "Selecione o professor em todos os horarios.")
                return redirect("contratos_agenda", pk=contrato.id)
            if conflitos:
                messages.warning(request, "Conflitos ao reservar: " + "; ".join(conflitos[:5]))
            else:
                messages.success(request, "Agenda criada com sucesso.")
                return redirect("alunos_detail", pk=contrato.cdAluno_id)

    weekday_labels = {
        0: "Segunda",
        1: "Terca",
        2: "Quarta",
        3: "Quinta",
        4: "Sexta",
        5: "Sabado",
        6: "Domingo",
    }
    slots_by_day = {key: [] for key in weekday_labels}
    seen_slots = set()
    for item in sorted(slots.values(), key=lambda x: (x["weekday"], x["inicio"])):
        key = (item["weekday"], item["inicio"], item["fim"])
        if key in seen_slots:
            continue
        seen_slots.add(key)
        slots_by_day[item["weekday"]].append(
            {
                "label": f'{item["inicio"].strftime("%H:%M")} - {item["fim"].strftime("%H:%M")}',
                "value": f'{item["weekday"]}|{item["inicio"].strftime("%H:%M")}|{item["fim"].strftime("%H:%M")}',
                "allowed_profs": ",".join(
                    [str(pid) for pid in (item.get("allowed_profs") or [])]
                ),
            }
        )

    slot_options = []
    for item in sorted(slots.values(), key=lambda x: (x["weekday"], x["inicio"])):
        slot_options.append(
            {
                "label": f'{weekday_labels[item["weekday"]]} {item["inicio"].strftime("%H:%M")} - {item["fim"].strftime("%H:%M")}',
                "value": f'{item["weekday"]}|{item["inicio"].strftime("%H:%M")}|{item["fim"].strftime("%H:%M")}',
                "allowed_profs": item.get("allowed_profs") or [],
            }
        )

    context = {
        "contrato": contrato,
        "aluno": contrato.cdAluno,
        "aulas_por_semana": aulas_por_semana,
        "agenda_semanas": agenda_semanas,
        "slot_indices": list(range(1, aulas_por_semana + 1)),
        "slots_by_day": slots_by_day,
        "weekday_labels": weekday_labels,
        "slot_options": slot_options,
        "profissionais": profissionais,
        "horarios_configurados": horarios_configurados,
        "breadcrumbs": [("Home", reverse("dashboard")), ("Contratos", reverse("contratos_list")), ("Agenda", "#")],
        "active_menu": "cadastros",
    }
    return render(request, "contratos/agenda.html", context)


@login_required
def email_config_view(request):
    cfg = models.EmailConfiguracao.objects.order_by("-dtCadastro").first()
    if request.method == "POST":
        data = request.POST.copy()
        data = _inject_cd_value(models.EmailConfiguracao, data)
        form = forms.EmailConfiguracaoForm(data, instance=cfg)
        if form.is_valid():
            form.save()
            messages.success(request, "Configuracao de email salva.")
            return redirect("email_config")
        messages.error(request, "Verifique os erros do formulario.")
    else:
        form = forms.EmailConfiguracaoForm(instance=cfg)
    return render(
        request,
        "configuracoes/email.html",
        {
            "form": form,
            "title": "Configuracao de Email",
            "breadcrumbs": [("Home", reverse("dashboard")), ("Configuracoes", "#"), ("Email", "#")],
            "active_menu": "configuracoes",
        },
    )


def aluno_app_view(request):
    return render(request, "app/index.html")


def aluno_app_manifest(request):
    from django.http import JsonResponse
    return JsonResponse({
        "name": "Mayris Pilates",
        "short_name": "Mayris",
        "start_url": "/app/",
        "scope": "/app/",
        "display": "standalone",
        "background_color": "#ffffff",
        "theme_color": "#7a1f2b",
        "icons": [
            {"src": "/app/icon.svg", "sizes": "any", "type": "image/svg+xml", "purpose": "any maskable"},
        ],
    })


def aluno_app_sw(request):
    js = (
        "const CACHE='mayris-app-v1';"
        "self.addEventListener('install',e=>{self.skipWaiting();});"
        "self.addEventListener('activate',e=>{self.clients.claim();});"
        "self.addEventListener('fetch',e=>{});"
    )
    resp = HttpResponse(js, content_type="application/javascript")
    resp["Service-Worker-Allowed"] = "/app/"
    return resp


def aluno_app_icon(request):
    svg = (
        "<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 192 192'>"
        "<rect width='192' height='192' rx='36' fill='#7a1f2b'/>"
        "<text x='96' y='128' font-size='110' text-anchor='middle' fill='#fff' font-family='Georgia'>M</text>"
        "</svg>"
    )
    return HttpResponse(svg, content_type="image/svg+xml")


@login_required
def mayasec_view(request):
    return render(
        request,
        "mayasec.html",
        {
            "title": "MayaSec",
            "breadcrumbs": [("Home", reverse("dashboard")), ("MayaSec", "#")],
            "active_menu": "mayasec",
        },
    )


@login_required
def whatsapp_config_view(request):
    unidades = models.Unidade.objects.order_by("cdUnidade").all()
    if not unidades:
        messages.warning(request, "Cadastre ao menos uma unidade antes de configurar o WhatsApp.")
        return redirect("dashboard")
    unidade_id = request.GET.get("unidade")
    unidade = unidades.filter(pk=unidade_id).first() if unidade_id else unidades.first()
    if not unidade:
        unidade = unidades.first()
    configuracao = models.WhatsappConfiguracao.objects.filter(unidade=unidade).first()
    batch_log = []
    batch_summary = None
    batch_sent = []
    batch_without_phone = []
    batch_rate_limited = []
    batch_failed = []
    batch_skipped = []
    if request.method == "POST":
        action = request.POST.get("action", "").strip()
        if action in {
            "send_aluno_now", "send_professor_now", "send_renovacao_now",
            "send_aniversario_now", "send_vencimento_now", "send_atraso_now", "send_tres_meses_now",
        }:
            if not configuracao:
                messages.warning(request, "Salve a configuracao antes de enviar agora.")
                return redirect(f"{reverse('whatsapp_config')}?unidade={unidade.id}")
            service = WhatsappService()
            hoje = timezone.localdate()
            if action == "send_aluno_now":
                resumo = _send_class_reminders(service, configuracao, hoje + timedelta(days=1), force=True)
                qtd = resumo.get("sent", 0)
                batch_log = resumo.get("entries", [])
                batch_sent = [item for item in batch_log if item.get("status") == "sent"]
                batch_without_phone = [item for item in batch_log if item.get("status") == "without_phone"]
                batch_rate_limited = [item for item in batch_log if item.get("status") == "rate_limited"]
                batch_failed = [item for item in batch_log if item.get("status") == "failed"]
                batch_skipped = [item for item in batch_log if item.get("status") == "skipped"]
                batch_summary = {
                    "sent": resumo.get("sent", 0),
                    "eligible_students": resumo.get("eligible_students", 0),
                    "without_phone": resumo.get("without_phone", 0),
                    "rate_limited": resumo.get("rate_limited", 0),
                    "failed": resumo.get("failed", 0),
                    "already_sent": resumo.get("already_sent", 0),
                }
                if qtd:
                    messages.success(
                        request,
                        (
                            f"Aviso ao aluno executado agora. {qtd} mensagem(ns) enviadas. "
                            f"{resumo.get('eligible_students', 0)} aluno(s) elegiveis, "
                            f"{resumo.get('without_phone', 0)} sem telefone e "
                            f"{resumo.get('rate_limited', 0)} pendente(s) por limite e "
                            f"{resumo.get('failed', 0)} falha(s)."
                        ),
                    )
                else:
                    messages.warning(
                        request,
                        (
                            "Nao ha aulas elegiveis para enviar agora. "
                            f"{resumo.get('eligible_students', 0)} aluno(s) encontrados, "
                            f"{resumo.get('without_phone', 0)} sem telefone."
                        ),
                    )
                form = forms.WhatsappConfiguracaoForm(instance=configuracao)
                return render(
                    request,
                    "configuracoes/whatsapp.html",
                    {
                        "form": form,
                        "unidades": unidades,
                        "unidade": unidade,
                        "batch_log": batch_log,
                        "batch_sent": batch_sent,
                        "batch_without_phone": batch_without_phone,
                        "batch_rate_limited": batch_rate_limited,
                        "batch_failed": batch_failed,
                        "batch_skipped": batch_skipped,
                        "batch_summary": batch_summary,
                        "title": "Configuracao de WhatsApp",
                        "breadcrumbs": [("Home", reverse("dashboard")), ("Configuracoes", "#"), ("WhatsApp", "#")],
                        "active_menu": "configuracoes",
                    },
                )
            elif action == "send_professor_now":
                qtd = _send_professor_schedule(service, configuracao, hoje + timedelta(days=1), force=True)
                if qtd:
                    messages.success(request, f"Aviso ao professor executado agora. {qtd} mensagem(ns) enviadas.")
                else:
                    messages.warning(request, "Nao ha agenda de professor elegivel para enviar agora.")
                form = forms.WhatsappConfiguracaoForm(instance=configuracao)
                return render(
                    request,
                    "configuracoes/whatsapp.html",
                    {
                        "form": form,
                        "unidades": unidades,
                        "unidade": unidade,
                        "batch_log": batch_log,
                        "batch_sent": batch_sent,
                        "batch_without_phone": batch_without_phone,
                        "batch_rate_limited": batch_rate_limited,
                        "batch_failed": batch_failed,
                        "batch_skipped": batch_skipped,
                        "batch_summary": batch_summary,
                        "title": "Configuracao de WhatsApp",
                        "breadcrumbs": [("Home", reverse("dashboard")), ("Configuracoes", "#"), ("WhatsApp", "#")],
                        "active_menu": "configuracoes",
                    },
                )
            elif action == "send_aniversario_now":
                qtd = _send_birthdays(service, configuracao, hoje, force=True)
                if qtd:
                    messages.success(request, f"Aniversário: {qtd} mensagem(ns) enviada(s).")
                else:
                    messages.warning(request, "Nenhum aniversariante (com telefone) hoje.")
                return redirect(f"{reverse('whatsapp_config')}?unidade={unidade.id}")
            elif action == "send_vencimento_now":
                qtd = _send_payment_due(service, configuracao, hoje, force=True)
                if qtd:
                    messages.success(request, f"Vencimento próximo: {qtd} mensagem(ns) enviada(s).")
                else:
                    messages.warning(request, "Nenhuma mensalidade vencendo amanhã (em aberto).")
                return redirect(f"{reverse('whatsapp_config')}?unidade={unidade.id}")
            elif action == "send_atraso_now":
                qtd = _send_payment_overdue(service, configuracao, hoje, force=True)
                if qtd:
                    messages.success(request, f"Mensalidade em atraso: {qtd} mensagem(ns) enviada(s).")
                else:
                    messages.warning(request, "Nenhuma mensalidade em atraso (em aberto).")
                return redirect(f"{reverse('whatsapp_config')}?unidade={unidade.id}")
            elif action == "send_tres_meses_now":
                qtd = _send_three_months(service, configuracao, hoje, force=True)
                if qtd:
                    messages.success(request, f"Acompanhamento 3 meses: {qtd} mensagem(ns) enviada(s).")
                else:
                    messages.warning(request, "Nenhum contrato completando 90 dias hoje.")
                return redirect(f"{reverse('whatsapp_config')}?unidade={unidade.id}")
            elif action == "send_renovacao_now":
                qtd = _send_contract_renewals(service, configuracao, hoje + timedelta(days=7), force=True)
                if qtd:
                    messages.success(request, f"Aviso de renovacao executado agora. {qtd} mensagem(ns) enviadas.")
                else:
                    messages.warning(request, "Nao ha contratos elegiveis para renovacao agora.")
                form = forms.WhatsappConfiguracaoForm(instance=configuracao)
                return render(
                    request,
                    "configuracoes/whatsapp.html",
                    {
                        "form": form,
                        "unidades": unidades,
                        "unidade": unidade,
                        "batch_log": batch_log,
                        "batch_sent": batch_sent,
                        "batch_without_phone": batch_without_phone,
                        "batch_rate_limited": batch_rate_limited,
                        "batch_failed": batch_failed,
                        "batch_skipped": batch_skipped,
                        "batch_summary": batch_summary,
                        "title": "Configuracao de WhatsApp",
                        "breadcrumbs": [("Home", reverse("dashboard")), ("Configuracoes", "#"), ("WhatsApp", "#")],
                        "active_menu": "configuracoes",
                    },
                )
        form = forms.WhatsappConfiguracaoForm(request.POST, instance=configuracao)
        if form.is_valid():
            cfg = form.save(commit=False)
            cfg.unidade = unidade
            cfg.save()
            messages.success(request, "Configuracoes de WhatsApp salvas.")
            return redirect(f"{reverse('whatsapp_config')}?unidade={unidade.id}")
        messages.error(request, "Corrija os erros antes de salvar.")
    else:
        form = forms.WhatsappConfiguracaoForm(instance=configuracao)
    return render(
          request,
          "configuracoes/whatsapp.html",
          {
            "form": form,
            "unidades": unidades,
            "unidade": unidade,
            "batch_log": batch_log,
            "batch_sent": batch_sent,
            "batch_without_phone": batch_without_phone,
            "batch_rate_limited": batch_rate_limited,
            "batch_failed": batch_failed,
            "batch_skipped": batch_skipped,
            "batch_summary": batch_summary,
            "title": "Configuracao de WhatsApp",
            "breadcrumbs": [("Home", reverse("dashboard")), ("Configuracoes", "#"), ("WhatsApp", "#")],
            "active_menu": "configuracoes",
          },
      )


@login_required
def whatsapp_historico_view(request):
    qs = (
        models.AlunoWhatsappMessage.objects.select_related("aluno", "aluno__cdUnidade")
        .order_by("-enviado_em")
    )
    f_status = request.GET.get("status", "").strip()
    f_tipo = request.GET.get("tipo", "").strip()
    f_unidade = request.GET.get("unidade", "").strip()
    f_inicio = request.GET.get("inicio", "").strip()
    f_fim = request.GET.get("fim", "").strip()
    f_busca = request.GET.get("q", "").strip()
    if f_status:
        qs = qs.filter(status=f_status)
    if f_tipo:
        qs = qs.filter(tipo=f_tipo)
    if f_unidade:
        qs = qs.filter(aluno__cdUnidade_id=f_unidade)
    if f_inicio:
        try:
            qs = qs.filter(enviado_em__date__gte=datetime.strptime(f_inicio, "%Y-%m-%d").date())
        except ValueError:
            pass
    if f_fim:
        try:
            qs = qs.filter(enviado_em__date__lte=datetime.strptime(f_fim, "%Y-%m-%d").date())
        except ValueError:
            pass
    if f_busca:
        qs = qs.filter(Q(aluno__dsNome__icontains=f_busca) | Q(telefone__icontains=f_busca))

    total = qs.count()
    enviadas = qs.filter(status="sent").count()
    falhas = total - enviadas

    paginator = Paginator(qs, 50)
    page = paginator.get_page(request.GET.get("page"))

    querystring = request.GET.copy()
    querystring.pop("page", None)

    context = {
        "mensagens": page,
        "page_obj": page,
        "total": total,
        "enviadas": enviadas,
        "falhas": falhas,
        "tipos": models.WhatsappMessageType.choices,
        "unidades": models.Unidade.objects.order_by("cdUnidade").all(),
        "filtros": {
            "status": f_status,
            "tipo": f_tipo,
            "unidade": f_unidade,
            "inicio": f_inicio,
            "fim": f_fim,
            "q": f_busca,
        },
        "querystring": querystring.urlencode(),
        "breadcrumbs": [("Home", reverse("dashboard")), ("Configuracoes", "#"), ("Historico WhatsApp", "#")],
        "active_menu": "configuracoes",
    }
    return render(request, "configuracoes/whatsapp_historico.html", context)


def _whatsapp_unidade_selecionada(request):
    unidades = models.Unidade.objects.order_by("cdUnidade").all()
    unidade_id = request.GET.get("unidade") or request.POST.get("unidade")
    unidade = unidades.filter(pk=unidade_id).first() if unidade_id else unidades.first()
    if not unidade:
        unidade = unidades.first()
    return unidades, unidade


@login_required
def aviso_aluno_config_view(request):
    bloqueio = _professor_block(request)
    if bloqueio:
        return bloqueio
    unidades, unidade = _whatsapp_unidade_selecionada(request)
    if not unidade:
        messages.warning(request, "Cadastre ao menos uma unidade antes de configurar o WhatsApp.")
        return redirect("dashboard")
    configuracao = models.WhatsappConfiguracao.objects.filter(unidade=unidade).first()
    if request.method == "POST":
        form = forms.AvisoAlunoConfigForm(request.POST, instance=configuracao)
        if form.is_valid():
            cfg = form.save(commit=False)
            cfg.unidade = unidade
            cfg.save()
            messages.success(request, "Configuracao do Aviso ao aluno salva.")
            return redirect(f"{reverse('aviso_aluno_config')}?unidade={unidade.id}")
        messages.error(request, "Verifique os erros do formulario.")
    else:
        form = forms.AvisoAlunoConfigForm(instance=configuracao)
    return render(
        request,
        "configuracoes/aviso_aluno.html",
        {
            "form": form,
            "unidades": unidades,
            "unidade": unidade,
            "title": "Aviso ao aluno",
            "breadcrumbs": [("Home", reverse("dashboard")), ("Configuracoes", "#"), ("Aviso ao aluno", "#")],
            "active_menu": "configuracoes",
        },
    )


@login_required
def aviso_aluno_preview_api(request):
    if _is_professor_user(request.user):
        return JsonResponse({"error": "Sem permissao."}, status=403)
    _unidades, unidade = _whatsapp_unidade_selecionada(request)
    config = models.WhatsappConfiguracao.objects.filter(unidade=unidade).first() if unidade else None
    if not config:
        return JsonResponse({"error": "Salve a configuracao desta unidade primeiro."}, status=400)
    target = timezone.localdate() + timedelta(days=1)
    service = WhatsappService()
    grupos = reservas_lembrete_por_aluno(config, target)
    items = []
    for _aluno_id, reservas in grupos.items():
        aluno = reservas[0].aluno
        telefone = service.get_aluno_phone(aluno)
        dedupe = _log_key("student_reminder", config.unidade_id, target, aluno.id)
        if not telefone:
            status = "sem_telefone"
        elif _already_sent(dedupe):
            status = "ja_enviado"
        else:
            status = "pendente"
        items.append({"aluno_id": aluno.id, "nome": aluno.dsNome, "telefone": telefone or "-", "status": status})
    items.sort(key=lambda x: x["nome"])
    return JsonResponse({"target": target.strftime("%d/%m/%Y"), "unidade": unidade.dsUnidade, "items": items})


@login_required
@require_POST
def aviso_aluno_enviar_um_api(request):
    if _is_professor_user(request.user):
        return JsonResponse({"status": "failed", "error": "Sem permissao."}, status=403)
    payload = _parse_json_body(request)
    aluno_id = payload.get("aluno_id")
    unidade = models.Unidade.objects.filter(pk=payload.get("unidade")).first() or models.Unidade.objects.first()
    config = models.WhatsappConfiguracao.objects.filter(unidade=unidade).first() if unidade else None
    if not config:
        return JsonResponse({"status": "failed", "error": "Configuracao nao encontrada."}, status=400)
    target = timezone.localdate() + timedelta(days=1)
    grupos = reservas_lembrete_por_aluno(config, target)
    try:
        reservas = grupos.get(int(aluno_id))
    except (TypeError, ValueError):
        reservas = None
    if not reservas:
        return JsonResponse({"status": "failed", "error": "Aluno sem aula amanha."}, status=400)
    aluno = reservas[0].aluno
    service = WhatsappService()
    telefone = service.get_aluno_phone(aluno)
    if not telefone:
        return JsonResponse({"status": "sem_telefone", "telefone": "-"})
    dedupe = _log_key("student_reminder", config.unidade_id, target, aluno.id)
    if _already_sent(dedupe):
        return JsonResponse({"status": "ja_enviado", "telefone": telefone})
    mensagem = montar_mensagem_lembrete(config, reservas, target)
    resp = service.send(aluno, telefone, mensagem, WhatsappMessageType.AUTOMATED_REMINDER)
    if _is_session_down(resp):
        return JsonResponse({"status": "session_down", "telefone": telefone, "error": "Sessao do WhatsApp desconectada."})
    if resp.get("error"):
        return JsonResponse({"status": "failed", "telefone": telefone, "error": resp.get("error")})
    _store_log(
        dedupe_key=dedupe,
        tipo=WhatsappMessageType.AUTOMATED_REMINDER,
        unidade=config.unidade,
        aluno=aluno,
        data_referencia=target,
        telefone=telefone,
        mensagem=mensagem,
        response_payload=json.dumps(resp, ensure_ascii=False),
    )
    return JsonResponse({"status": "sent", "telefone": telefone})


@login_required
def comunicado_geral_view(request):
    bloqueio = _professor_block(request)
    if bloqueio:
        return bloqueio
    unidades, unidade = _whatsapp_unidade_selecionada(request)
    if not unidade:
        messages.warning(request, "Cadastre ao menos uma unidade antes de enviar comunicados.")
        return redirect("dashboard")
    return render(
        request,
        "configuracoes/comunicado_geral.html",
        {
            "unidades": unidades,
            "unidade": unidade,
            "title": "Comunicado Geral",
            "breadcrumbs": [("Home", reverse("dashboard")), ("Configuracoes", "#"), ("Comunicado Geral", "#")],
            "active_menu": "configuracoes",
        },
    )


@login_required
def comunicado_geral_alunos_api(request):
    if _is_professor_user(request.user):
        return JsonResponse({"error": "Sem permissao."}, status=403)
    _unidades, unidade = _whatsapp_unidade_selecionada(request)
    if not unidade:
        return JsonResponse({"error": "Selecione uma unidade."}, status=400)
    status_filtro = (request.GET.get("status") or "ATIVO").upper()
    if status_filtro not in {"ATIVO", "INATIVO", "AMBOS"}:
        status_filtro = "ATIVO"
    service = WhatsappService()
    alunos = models.Aluno.objects.filter(cdUnidade=unidade).order_by("dsNome")
    if status_filtro != "AMBOS":
        alunos = alunos.filter(status=status_filtro)
    items = []
    for aluno in alunos:
        telefone = service.get_aluno_phone(aluno)
        items.append(
            {
                "aluno_id": aluno.id,
                "nome": aluno.dsNome,
                "telefone": telefone or "-",
                "status": "pendente" if telefone else "sem_telefone",
            }
        )
    return JsonResponse({"unidade": unidade.dsUnidade, "status": status_filtro, "items": items})


@login_required
@require_POST
def comunicado_geral_enviar_um_api(request):
    if _is_professor_user(request.user):
        return JsonResponse({"status": "failed", "error": "Sem permissao."}, status=403)
    payload = _parse_json_body(request)
    aluno_id = payload.get("aluno_id")
    mensagem = (payload.get("mensagem") or "").strip()
    unidade = models.Unidade.objects.filter(pk=payload.get("unidade")).first() or models.Unidade.objects.first()
    if not mensagem:
        return JsonResponse({"status": "failed", "error": "Mensagem vazia."}, status=400)
    aluno = models.Aluno.objects.filter(pk=aluno_id).first()
    if not aluno:
        return JsonResponse({"status": "failed", "error": "Aluno nao encontrado."}, status=400)
    service = WhatsappService()
    telefone = service.get_aluno_phone(aluno)
    if not telefone:
        return JsonResponse({"status": "sem_telefone", "telefone": "-"})
    texto = mensagem.replace("{aluno}", aluno.dsNome)
    if unidade:
        texto = texto.replace("{unidade}", unidade.dsUnidade)
    resp = service.send(aluno, telefone, texto, WhatsappMessageType.MANUAL)
    if _is_session_down(resp):
        return JsonResponse({"status": "session_down", "telefone": telefone, "error": "Sessao do WhatsApp desconectada."})
    if resp.get("error"):
        return JsonResponse({"status": "failed", "telefone": telefone, "error": resp.get("error")})
    return JsonResponse({"status": "sent", "telefone": telefone})


def totalpass_config_view(request):
    unidades = models.Unidade.objects.order_by("cdUnidade").all()
    if not unidades:
        messages.warning(request, "Cadastre ao menos uma unidade antes de configurar o TotalPass.")
        return redirect("dashboard")
    unidade_id = request.GET.get("unidade")
    unidade = unidades.filter(pk=unidade_id).first() if unidade_id else unidades.first()
    if not unidade:
        unidade = unidades.first()
    configuracao = models.TotalpassConfiguracao.objects.filter(unidade=unidade).first()
    if request.method == "POST":
        form = forms.TotalpassConfiguracaoForm(request.POST, instance=configuracao)
        if form.is_valid():
            cfg = form.save(commit=False)
            cfg.unidade = unidade
            cfg.save()
            messages.success(request, "Configuracoes do TotalPass salvas.")
            return redirect(f"{reverse('totalpass_config')}?unidade={unidade.id}")
        messages.error(request, "Corrija os erros antes de salvar.")
    else:
        form = forms.TotalpassConfiguracaoForm(instance=configuracao)
    return render(
        request,
        "configuracoes/totalpass.html",
        {
            "form": form,
            "unidades": unidades,
            "unidade": unidade,
            "title": "Configuracao TotalPass",
            "breadcrumbs": [("Home", reverse("dashboard")), ("Configuracoes", "#"), ("TotalPass", "#")],
            "active_menu": "configuracoes",
        },
    )


def contrato_assinar(request, token):
    try:
        contrato_id = services.validar_token_contrato(token)
    except Exception:
        return render(request, "contratos/assinatura.html", {"token_invalido": True})
    contrato = get_object_or_404(models.Contrato, pk=contrato_id)
    if request.method == "POST":
        if contrato.status in ("ASSINADO", "ASSINADO_DIGITALMENTE"):
            return render(
                request,
                "contratos/assinatura_sucesso.html",
                {"contrato": contrato, "ja_assinado": True},
            )
        assinatura_nome = request.POST.get("assinatura_nome", "").strip()
        assinatura_documento = request.POST.get("assinatura_documento", "").strip()
        assinatura_data = request.POST.get("assinatura_data", "").strip()
        assinatura_ip = request.META.get("REMOTE_ADDR")
        if assinatura_nome and assinatura_data.startswith("data:image/"):
            try:
                from django.core.files.base import ContentFile
                import base64

                header, data = assinatura_data.split(",", 1)
                content = ContentFile(base64.b64decode(data), name=f"contrato_{contrato.id}.png")
                contrato.assinatura_imagem = content
            except Exception:
                assinatura_data = ""
        contrato.assinatura_nome = assinatura_nome
        contrato.assinatura_documento = assinatura_documento
        contrato.assinatura_ip = assinatura_ip
        contrato.status = "ASSINADO_DIGITALMENTE"
        contrato.assinado_em = timezone.now()
        contrato.save()
        already_notified = models.AlunoWhatsappMessage.objects.filter(
            contrato=contrato, tipo=WhatsappMessageType.CONTRACT_LINK
        ).exists()
        if not already_notified:
            try:
                service = WhatsappService()
                telefone = service.get_aluno_phone(contrato.cdAluno)
                if telefone:
                    link = _contrato_assinatura_link(contrato, request=request)
                    mensagem = (
                        f"Olá {contrato.cdAluno.dsNome}! "
                        f"O contrato #{contrato.cdContrato} foi assinado. "
                        f"Você pode acessá-lo em {link}"
                    )
                    resp = service.send(
                        contrato.cdAluno,
                        telefone,
                        mensagem,
                        WhatsappMessageType.CONTRACT_LINK,
                        contrato=contrato,
                    )
                    if resp.get("error"):
                        messages.warning(request, "Contrato assinado, mas não foi possível enviar o aviso via WhatsApp.")
                else:
                    messages.warning(request, "Contrato assinado, mas o aluno não possui telefone válido.")
            except Exception:
                logger.exception("Erro ao enviar contrato assinado pelo WhatsApp")
        return render(request, "contratos/assinatura_sucesso.html", {"contrato": contrato})
    html = services.render_contrato_html(contrato)
    return render(
        request,
        "contratos/assinatura.html",
        {"contrato": contrato, "contrato_html": html, "token": token},
    )


def contrato_pdf(request, token):
    try:
        contrato_id = services.validar_token_contrato(token)
    except Exception:
        return HttpResponse("Token invalido", status=404)
    contrato = get_object_or_404(models.Contrato, pk=contrato_id)
    pdf = services.render_contrato_pdf(contrato)
    _salvar_documento_contrato(contrato, pdf_bytes=pdf)
    response = HttpResponse(pdf, content_type="application/pdf")
    response["Content-Disposition"] = f'inline; filename="contrato-{contrato.cdContrato}.pdf"'
    return response


@login_required
def contrato_enviar_email(request, pk):
    contrato = get_object_or_404(models.Contrato, pk=pk)
    if request.method != "POST":
        return redirect("alunos_detail", pk=contrato.cdAluno_id)
    if not contrato.cdAluno.dsEmail:
        messages.warning(request, "Aluno sem email cadastrado.")
        return redirect("alunos_detail", pk=contrato.cdAluno_id)
    if services.enviar_contrato_para_assinatura(contrato, request.build_absolute_uri("/")):
        messages.success(request, "Contrato enviado para assinatura por email.")
    else:
        messages.warning(request, "Nao foi possivel enviar o email do contrato.")
    return redirect("alunos_detail", pk=contrato.cdAluno_id)


@login_required
def contrato_assinar_local(request, pk):
    contrato = get_object_or_404(models.Contrato, pk=pk)
    token = services.gerar_token_contrato(contrato)
    return redirect("contrato_assinar", token=token)


@login_required
def contrato_documento(request, pk):
    contrato = get_object_or_404(models.Contrato, pk=pk)
    conteudo = services.render_contrato_html(contrato)
    return render(
        request,
        "contratos/documento.html",
        {
            "contrato": contrato,
            "conteudo": conteudo,
            "active_menu": "contratos",
        },
    )


@login_required
def contrato_assinatura_detalhe(request, pk):
    contrato = get_object_or_404(models.Contrato, pk=pk)
    contrato_html = services.render_contrato_html(contrato)
    return render(
        request,
        "contratos/assinatura_detalhe.html",
        {
            "contrato": contrato,
            "contrato_html": contrato_html,
            "breadcrumbs": [("Home", reverse("dashboard")), ("Alunos", reverse("alunos_list")), ("Assinatura", "#")],
            "active_menu": "cadastros",
        },
    )
